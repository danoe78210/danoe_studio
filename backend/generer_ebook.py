#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generer_ebook.py – Générateur d'ebook EPUB 3 compatible Amazon KDP (v1.6)

v1.6 – FIN DE TRAVAIL PROPRE DANS L'INTERFACE :
- plus de sys.exit() dans le flux normal : le script expose EBOOK_CODE ;
  le thread de l'interface se termine donc toujours proprement et
  l'interface se déverrouille (bandeau K2000 arrêté, boutons actifs) ;
- en ligne de commande directe, le comportement et les messages restent
  identiques.

v1.5 – CORRECTIFS :
- DeprecationWarning supprimée : datetime.now(timezone.utc) ;
- entité « &nbsp; » interdite en XHTML → « &#160; » ;
- validation XML (well-formed) de chaque page AVANT écriture du ZIP ;
- EPUB écrit directement (sans ebooklib) : mimetype + container.xml +
  content.opf + toc.ncx + nav.xhtml garantis.

Règles KDP respectées :
- reflowable (pas de pagination fixe, pas d'en-têtes/pieds) ;
- double TDM : NCX (logique) + nav.xhtml + page TDM HTML visible ;
- lettrine neutralisée ; retrait 1re ligne 0,5 cm ; titres H1/H2 ;
- images JPEG ; couverture 1600x2560 (ratio 1.6:1), auto-générée si absente.

Usage :
    python generer_ebook.py                → génère l'EPUB
    python generer_ebook.py --docx chemin  → DOCX précis
    python generer_ebook.py --cover-only   → couverture seulement

Dépendances : pip install python-docx Pillow openpyxl   (ebooklib inutile)
"""
import os
import re
import sys
import glob
import zipfile
import io
import time
import xml.etree.ElementTree as ET
from datetime import datetime, timezone

try:
    from docx import Document
except ImportError:
    print('   ⚠️  python-docx requis : pip install python-docx')
    raise

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
except ImportError:
    PILImage = None
    print('   ⚠️  Pillow absent : génération de couverture désactivée.')

try:
    import configuration_store as cs
    JSON_OK = True
except Exception:
    cs = None
    JSON_OK = False

if getattr(sys, 'frozen', False):
    BASE = os.path.dirname(sys.executable)
else:
    BASE = os.path.dirname(os.path.abspath(__file__))

DOSSIER_IMAGES = os.path.join(BASE, 'Images')
CHEMIN_CONFIG = os.path.join(BASE, 'Configuration_roman.xlsx')

# v1.6 : code de retour exposé au module appelant (interface ou CLI)
EBOOK_CODE = 0

# ─────────────────────────────────────────────
# 1. CONSTANTES
# ─────────────────────────────────────────────
TITRE = 'Titre complet du roman'
SOUS_TITRE = 'Sous-titre éventuel'
AUTEUR = "Nom de l'auteur (couverture)"
ISBN = 'Numéro ISBN'
ANNEE = 'Année de publication'
EDITEUR = "Maison d'édition / auto-édition"
COPYRIGHT = 'Mention de copyright'
SITE = 'Site web'
DEDICACE = 'Dédicace'
EPIGRAPHE = 'Épigraphe'

CHAMPS_LABELS = [TITRE, SOUS_TITRE, AUTEUR, ISBN, ANNEE, EDITEUR,
                 COPYRIGHT, SITE, DEDICACE, EPIGRAPHE]

CORRESPONDANCES = {
    TITRE: ('titre complet',), SOUS_TITRE: ('sous-titre',), AUTEUR: ("nom de l'auteur",),
    ISBN: ('numéro isbn', 'isbn'), ANNEE: ('année',), EDITEUR: ("maison d'édition",),
    COPYRIGHT: ('mention de copyright', 'copyright'), SITE: ('site web',),
    DEDICACE: ('dédicace',), EPIGRAPHE: ('épigraphe',),
}

JSON_INFOS_KEYS = {
    TITRE: 'titre_complet', SOUS_TITRE: 'sous_titre', AUTEUR: 'auteur',
    ISBN: 'isbn', ANNEE: 'annee_publication', EDITEUR: 'maison_edition',
    COPYRIGHT: 'mention_copyright', SITE: 'site_web',
    DEDICACE: 'dedicace', EPIGRAPHE: 'epigraphe',
}

COVER_W, COVER_H = 1600, 2560

CONTAINER_XML = '''<?xml version="1.0" encoding="UTF-8"?>
<container version="1.0" xmlns="urn:oasis:names:tc:opendocument:xmlns:container">
  <rootfiles>
    <rootfile full-path="OEBPS/content.opf" media-type="application/oebps-package+xml"/>
  </rootfiles>
</container>
'''

CSS_EBOOK = b"""
@charset "UTF-8";
body { font-family: serif; text-align: justify; line-height: 1.5; margin: 1em; }
h1 { text-align: center; page-break-before: always; margin: 2em 0 1em; font-size: 1.8em; }
h2 { text-align: center; page-break-before: always; margin: 1.5em 0 0.8em; font-size: 1.4em; }
h3 { text-align: center; page-break-before: always; margin: 1.2em 0 0.6em; font-size: 1.2em; }
p { text-indent: 0.5cm; margin: 0.3em 0; }
p.first { text-indent: 0; }
p.sep { text-align: center; margin: 1em 0; text-indent: 0; font-size: 1.1em; letter-spacing: 0.3em; }
.title-page { text-align: center; margin: 3em 1em; page-break-after: always; }
.title-page h1 { font-size: 2.2em; margin: 2em 0 0.5em; letter-spacing: 0.05em; page-break-before: never; }
.title-page .subtitle { font-size: 1.3em; margin: 0.5em 0 2em; font-style: italic; }
.title-page .author { margin-top: 3em; font-size: 1.3em; }
.copyright { page-break-after: always; font-size: 0.85em; margin: 2em; text-align: left; line-height: 1.4; }
.dedication, .epigraph { text-align: center; font-style: italic; page-break-after: always; margin: 4em 2em; line-height: 1.6; }
.dedication p, .epigraph p { text-indent: 0; margin: 0.5em 0; }
img { max-width: 90%; display: block; margin: 1em auto; }
.publisher { margin-top: 2.5em; font-size: 1.05em; letter-spacing: 0.08em; }
.preface, .postface, .remerciements, .autres-livres { page-break-before: always; }
.autres-livres ul { list-style: none; padding: 0; } .autres-livres li { margin: 0.4em 0; text-indent: 0; }
.planche { text-align: center; page-break-before: always; }
.planche p { text-indent: 0; }
.fig { text-align: center; font-size: 0.85em; font-style: italic; margin: 0.3em 0 1em; text-indent: 0; }
.toc-page { page-break-before: always; }
.toc-page h1 { text-align: center; page-break-before: never; }
.toc-page ul { list-style: none; padding: 0; }
.toc-page li { margin: 0.4em 0; }
.toc-page a { text-decoration: none; color: inherit; }
.act-page { page-break-before: always; text-align: center; padding-top: 35%; margin-bottom: 1em; }
.act-page h1 { font-size: 2em; margin: 0; page-break-before: never; }
strong { font-weight: bold; }
em { font-style: italic; }
"""


# ─────────────────────────────────────────────
# 2. CONFIGURATION / FICHIERS SOURCES
# ─────────────────────────────────────────────
def _lire_json_brut():
    try:
        import json
        return _normaliser_cles(json.load(open(os.path.join(BASE, 'Configuration_roman.json'), encoding='utf-8')))
    except Exception:
        return None

def _get_infos(d):
    d = d or {}
    if 'informations' in d: return d['informations'] or {}
    if 'informations ' in d: return d['informations '] or {}
    for k, v in d.items():
        if str(k).strip() == 'informations' and isinstance(v, dict): return v
    return {}


def _get_section(d, nom, defaut=None):
    """Recherche tolérante d'une section JSON (clés avec espaces finaux)."""
    if defaut is None: defaut = {}
    if not isinstance(d, dict): return defaut
    if nom in d and d[nom] is not None: return d[nom]
    for k, v in d.items():
        if str(k).strip() == nom and v is not None: return v
    return defaut

def _section_infos(d):
    return _get_section(d, 'informations', {}) or {}

def lire_infos():
    infos = {l: '' for l in CHAMPS_LABELS}
    source = 'défaut'
    ji = _get_infos(_lire_json_brut())
    for label in CHAMPS_LABELS:
        for k, v in ji.items():
            kl = str(k).strip().lower()
            if v and any(kl.startswith(p) for p in CORRESPONDANCES[label]):
                infos[label] = str(v).strip(); source = 'JSON'; break
    if JSON_OK:
        try:
            data = cs.charger_configuration()
            j = _infos_de(data)
            for label, cle in JSON_INFOS_KEYS.items():
                v = j.get(cle, '')
                if v and not infos[label]:
                    infos[label] = str(v).strip()
            if any(infos.values()):
                source = 'JSON'
        except Exception:
            pass
    if source == 'défaut' and os.path.isfile(CHEMIN_CONFIG):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CHEMIN_CONFIG, data_only=True)
            ws = wb['Informations'] if 'Informations' in wb.sheetnames else wb.active
            for row in ws.iter_rows(values_only=True):
                if not row or row[0] is None:
                    continue
                k = str(row[0]).strip().lower()
                v = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                for label in CHAMPS_LABELS:
                    if any(k.startswith(p) for p in CORRESPONDANCES.get(label, ())):
                        infos[label] = v
                        source = 'Excel'
                        break
        except Exception:
            pass
    infos['_source'] = source
    return infos



def lire_annexes():
    """Lit les champs annexes du menu Informations (clés FR ou snake)."""
    ax = {'sommaire': True}
    j = {}
    try:
        import json
        p = os.path.join(BASE, 'Configuration_roman.json')
        if os.path.isfile(p):
            j = _infos_de(json.load(open(p, encoding='utf-8')))
    except Exception:
        j = {}
    def g(*cles):
        for c in cles:
            v = j.get(c)
            if v: return str(v).strip()
        return ''
    ax['editeur'] = g('editeur', 'Éditeur', 'Editeur', 'maison_edition')
    ax['autres_livres'] = g('autres_livres', 'Autres livres du même auteur')
    ax['remerciements'] = g('remerciements', 'Remerciements')
    ax['frontispice'] = g('frontispice', 'Frontispice')
    ax['preface'] = g('preface', 'Préface')
    ax['postface'] = g('postface', 'Postface')
    sv = j.get('sommaire')
    ax['sommaire'] = sv if isinstance(sv, bool) else (str(sv).strip().lower() != 'false' if sv is not None else True)
    return ax

def _md_to_html(txt):
    out = []
    for ln in txt.split('\n'):
        ln = ln.rstrip()
        if not ln.strip() or ln.startswith('# '): continue
        if ln.startswith('## '): out.append('<h2>%s</h2>' % ln[3:].strip())
        else: out.append('<p>%s</p>' % ln)
    return ''.join(out)

def trouver_docx():
    cands = (glob.glob(os.path.join(BASE, '*_KDP.docx'))
             + glob.glob(os.path.join(BASE, 'export', '*_KDP.docx')))
    return max(cands, key=os.path.getmtime) if cands else None


def trouver_couverture():
    if not os.path.isdir(DOSSIER_IMAGES):
        return None
    for nom in sorted(os.listdir(DOSSIER_IMAGES)):
        if nom.lower().startswith('couverture') and nom.lower().endswith(('.jpg', '.jpeg', '.png')):
            return os.path.join(DOSSIER_IMAGES, nom)
    return None


# ─────────────────────────────────────────────
# 3. COUVERTURE (PIL)
# ─────────────────────────────────────────────
def generer_couverture(titre, sous_titre, auteur, sortie):
    if PILImage is None:
        return False
    try:
        img = PILImage.new('RGB', (COVER_W, COVER_H), '#0a0e1a')
        draw = ImageDraw.Draw(img)
        for y in range(COVER_H):
            r = y / COVER_H
            draw.line([(0, y), (COVER_W, y)],
                      fill=(int(10 + r * 15), int(14 + r * 20), int(26 + r * 35)))
        try:
            ft = ImageFont.truetype('arial.ttf', 120)
            fs = ImageFont.truetype('arial.ttf', 60)
            fa = ImageFont.truetype('arial.ttf', 70)
        except Exception:
            ft = fs = fa = ImageFont.load_default()
        draw.rectangle([80, 80, COVER_W - 80, COVER_H - 80], outline='#e0b458', width=4)
        draw.rectangle([100, 100, COVER_W - 100, COVER_H - 100], outline='#4a5070', width=2)
        y = 500
        for mot in (titre or 'Roman').split():
            bb = draw.textbbox((0, 0), mot, font=ft)
            draw.text(((COVER_W - (bb[2] - bb[0])) / 2, y), mot, fill='#f5f5f0', font=ft)
            y += 140
        if sous_titre:
            bb = draw.textbbox((0, 0), sous_titre, font=fs)
            draw.text(((COVER_W - (bb[2] - bb[0])) / 2, y + 80), sous_titre,
                      fill='#9aa3c7', font=fs)
        if auteur:
            bb = draw.textbbox((0, 0), auteur, font=fa)
            draw.text(((COVER_W - (bb[2] - bb[0])) / 2, COVER_H - 300), auteur,
                      fill='#e0b458', font=fa)
        img.save(sortie, 'JPEG', quality=92)
        print(f'   🎨 Couverture générée : {os.path.basename(sortie)}')
        return True
    except Exception as e:
        print(f'   ⚠️  Génération couverture échouée : {e}')
        return False


def couverture_octets():
    """Retourne (octets JPEG, origine) : fichier /Images ou génération PIL."""
    chemin = trouver_couverture()
    if chemin and PILImage is not None:
        try:
            with PILImage.open(chemin) as im:
                buf = io.BytesIO()
                im.convert('RGB').save(buf, 'JPEG', quality=92)
                print(f'   🎨 Couverture : {os.path.basename(chemin)}')
                return buf.getvalue(), os.path.basename(chemin)
        except Exception as e:
            print(f'   ⚠️  Couverture illisible : {e}')
    out = os.path.join(BASE, 'Couverture_ebook.jpg')
    infos = lire_infos()
    if generer_couverture(infos[TITRE], infos[SOUS_TITRE], infos[AUTEUR], out):
        with open(out, 'rb') as f:
            return f.read(), 'Couverture_ebook.jpg'
    return None, None


# ─────────────────────────────────────────────
# 4. IMAGES DU DOCX
# ─────────────────────────────────────────────
def extraire_images(docx_path):
    images = []
    try:
        with zipfile.ZipFile(docx_path, 'r') as z:
            media = sorted(n for n in z.namelist() if n.startswith('word/media/'))
            for i, name in enumerate(media):
                data = z.read(name)
                if PILImage is not None:
                    try:
                        with PILImage.open(io.BytesIO(data)) as im:
                            buf = io.BytesIO()
                            im.convert('RGB').save(buf, 'JPEG', quality=85)
                            images.append({'data': buf.getvalue(),
                                           'nom': f'img_{i + 1:03d}.jpg',
                                           'media': 'image/jpeg'})
                            continue
                    except Exception:
                        pass
                ext = os.path.splitext(name)[1].lstrip('.').lower()
                images.append({'data': data, 'nom': f'img_{i + 1:03d}.{ext}',
                               'media': 'image/png' if ext == 'png' else 'image/jpeg'})
    except Exception as e:
        print(f'   ⚠️  Extraction images impossible : {e}')
    return images


# ─────────────────────────────────────────────
# 5. UTILITAIRES HTML / XHTML
# ─────────────────────────────────────────────
def html_escape(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


def runs_to_html(paragraph):
    parts = []
    for run in paragraph.runs:
        t = html_escape(run.text)
        if not t:
            continue
        if run.bold and run.italic:
            parts.append(f'<strong><em>{t}</em></strong>')
        elif run.bold:
            parts.append(f'<strong>{t}</strong>')
        elif run.italic:
            parts.append(f'<em>{t}</em>')
        else:
            parts.append(t)
    return ''.join(parts)


def contient_image(paragraph):
    xml = paragraph._element.xml
    return 'blip' in xml or 'graphicData' in xml


def page_xhtml(titre, corps):
    return ('<?xml version="1.0" encoding="utf-8"?>\n'
            '<!DOCTYPE html>\n'
            '<html xmlns="http://www.w3.org/1999/xhtml" xml:lang="fr">\n'
            '<head><title>' + html_escape(titre) + '</title>\n'
            '<link rel="stylesheet" type="text/css" href="style/default.css"/>\n'
            '</head>\n<body>\n' + corps + '\n</body>\n</html>').encode('utf-8')


def valider_xml(nom, data):
    """Contrôle well-formed XML d'un document avant écriture."""
    try:
        ET.fromstring(data)
        return None
    except Exception as e:
        return f'{nom} : {e}'


# ─────────────────────────────────────────────
# 6. CONSTRUCTION DE L'EPUB (écriture ZIP directe)
# ─────────────────────────────────────────────
def construire_ebook(docx_path, infos, images, sortie):
    doc = Document(docx_path)

    titre = infos[TITRE] or 'Roman'
    sous_titre = infos[SOUS_TITRE] or ''
    auteur = infos[AUTEUR] or 'Auteur'
    annee = infos[ANNEE] or str(datetime.now().year)
    ident = infos[ISBN] or f'urn:uuid:ebook-{int(time.time())}'
    titre_complet = f'{titre} – {sous_titre}' if sous_titre else titre

    items = []    # {'id','href','media','data','props'}
    spine = []    # idref dans l'ordre de lecture
    toc = []      # (idref, libellé) pour NCX + nav

    def add(iid, href, media, data, props=None):
        items.append({'id': iid, 'href': href, 'media': media,
                      'data': data, 'props': props})

    def add_page(iid, href, titre_page, corps, entree_toc=None):
        add(iid, href, 'application/xhtml+xml', page_xhtml(titre_page, corps))
        spine.append(iid)
        if entree_toc:
            toc.append((iid, entree_toc))

    # ── Ressources fixes ──
    add('css', 'style/default.css', 'text/css', CSS_EBOOK)
    cover_data, cover_orig = couverture_octets()
    if cover_data:
        add('cover-image', 'cover.jpg', 'image/jpeg', cover_data, 'cover-image')
    _ax0 = lire_annexes()
    if _ax0.get('frontispice'):
        _fp = os.path.join(DOSSIER_IMAGES, _ax0['frontispice'])
        if not os.path.isfile(_fp):
            for _ext in ('.png', '.jpg', '.jpeg', '.webp', '.bmp'):
                if os.path.isfile(_fp + _ext): _fp += _ext; break
        if os.path.isfile(_fp) and PILImage is not None:
            try:
                with PILImage.open(_fp) as im:
                    _buf = io.BytesIO(); im.convert('RGB').save(_buf, 'JPEG', quality=88)
                add('frontispice-img', 'images/frontispice.jpg', 'image/jpeg', _buf.getvalue())
                add_page('frontispice', 'frontispice.xhtml', 'Frontispice',
                         '<div class="planche"><p><img src="images/frontispice.jpg" alt="frontispice"/></p></div>')
            except Exception: pass

    # ── Pages liminaires (structure éditoriale française) ──
    
    # 1. Page de garde (blanche)
    add_page('garde', 'garde.xhtml', 'Page de garde', '<div class="page-garde"></div>')
    
    # 2. Faux-titre
    corps_faux_titre = f'<div class="faux-titre"><h1>{html_escape(titre)}</h1></div>'
    add_page('faux-titre', 'faux-titre.xhtml', 'Faux-titre', corps_faux_titre)
    
    # 3. Frontispice
    frontispice = infos.get('frontispice') or infos.get('Frontispice')
    if frontispice:
        img_path = os.path.join(DOSSIER_IMAGES, frontispice)
        if os.path.exists(img_path):
            # Copier l'image dans l'EPUB
            img_data = open(img_path, 'rb').read()
            img_id = f'frontispice-{int(time.time())}'
            add(img_id, f'images/{frontispice}', 'image/jpeg' if frontispice.lower().endswith(('.jpg', '.jpeg')) else 'image/png', img_data)
            corps_frontispice = f'<div class="frontispice"><img src="images/{frontispice}" alt="Frontispice"/></div>'
            add_page('frontispice', 'frontispice.xhtml', 'Frontispice', corps_frontispice)
    
    # 4. Page de titre
    corps_titre = '<div class="title-page"><h1>' + html_escape(titre.upper()) + '</h1>'
    if sous_titre:
        corps_titre += f'<p class="subtitle">{html_escape(sous_titre)}</p>'
    corps_titre += f'<p class="author">{html_escape(auteur)}</p>'
    if infos[EDITEUR]:
        corps_titre += f'<p class="publisher">{html_escape(infos[EDITEUR])}</p>'
    corps_titre += '</div>'
    add_page('title', 'title.xhtml', titre, corps_titre, 'Page de titre')
    
    # 5. Copyright
    copyright_txt = infos[COPYRIGHT] or f'© {annee} {auteur}. Tous droits réservés.'
    lignes = [titre_complet, '', copyright_txt]
    if infos[ISBN]:
        lignes.append('ISBN : ' + infos[ISBN])
    lignes.append(f'Dépôt légal : {annee}')
    if infos[SITE]:
        lignes += ['', infos[SITE]]
    lignes += ['', "Toute reproduction, même partielle, est interdite sans l'autorisation préalable de l'auteur."]
    corps = '<div class="copyright">' + ''.join(
        f'<p style="text-indent:0">{html_escape(l) if l else "&#160;"}</p>'
        for l in lignes) + '</div>'
    add_page('copyright', 'copyright.xhtml', 'Mentions légales', corps)
    
    # 6. Dédicace
    if infos[DEDICACE]:
        corps = '<div class="dedication">' + ''.join(
            f'<p>{html_escape(l)}</p>' for l in infos[DEDICACE].splitlines() if l.strip()) + '</div>'
        add_page('dedicace', 'dedicace.xhtml', 'Dédicace', corps, 'Dédicace')
    
    # 7. Épigraphe
    if infos[EPIGRAPHE]:
        corps = '<div class="epigraph">' + ''.join(
            f'<p>{html_escape(l)}</p>' for l in infos[EPIGRAPHE].splitlines() if l.strip()) + '</div>'
        add_page('epigraphe', 'epigraphe.xhtml', 'Épigraphe', corps, 'Épigraphe')
    
    # 8. Table des matières
    corps = '<div class="toc-page"><h1>Table des matières</h1><ul>'
    for iid, lib in toc:
        href = next(i['href'] for i in items if i['id'] == iid)
        corps += f'<li><a href="{href}">{html_escape(lib)}</a></li>'
    corps += '</ul></div>'
    add_page('tdm', 'tdm.xhtml', 'Table des matières', corps, 'Table des matières')
    
    # 9. Préface
    preface = infos.get('preface') or infos.get('Préface')
    if preface:
        preface_path = os.path.join(DOSSIER_CHAPITRES, f"{preface}.md")
        if os.path.exists(preface_path):
            contenu_preface = lire_fichier_markdown(preface_path)
            corps_preface = '<div class="preface"><h1>Préface</h1>' + convertir_markdown_html(contenu_preface) + '</div>'
            add_page('preface', 'preface.xhtml', 'Préface', corps_preface, 'Préface')
    
    # ── Parcours du manuscrit ──
    print('   📖 Parcours du manuscrit…')
    nb_actes = nb_chapitres = nb_images = 0
    idx_chap = idx_planche = 0
    image_counter = 0
    chapitre_courant = None
    blocs_chapitre = []

    def flush_chapitre():
        nonlocal chapitre_courant, blocs_chapitre
        if chapitre_courant and blocs_chapitre:
            corps = ''.join(blocs_chapitre)
            add_page(chapitre_courant['id'], chapitre_courant['href'],
                     chapitre_courant['titre'], corps, chapitre_courant['titre'])
        chapitre_courant = None
        blocs_chapitre = []

    for para in doc.paragraphs:
        style_name = para.style.name if para.style else ''
        texte = para.text.strip()

        if style_name == 'TitreActe' and texte:
            flush_chapitre()
            nb_actes += 1
            iid = f'acte{nb_actes:03d}'
            add_page(iid, f'{iid}.xhtml', texte,
                     f'<div class="act-page"><h1>{html_escape(texte)}</h1></div>', texte)

        elif style_name == 'TitreChapitre' and texte:
            flush_chapitre()
            nb_chapitres += 1
            idx_chap += 1
            chapitre_courant = {
                'id': f'chap{idx_chap:03d}',
                'href': f'chap{idx_chap:03d}.xhtml',
                'titre': texte
            }
            blocs_chapitre = [f'<h1>{html_escape(texte)}</h1>']

        elif style_name == 'TitreSousChap' and texte:
            if chapitre_courant is not None:
                blocs_chapitre.append(f'<h2>{html_escape(texte)}</h2>')

        elif style_name == 'SeparateurScene':
            if chapitre_courant is not None:
                blocs_chapitre.append('<p class="sep">* * *</p>')

        elif contient_image(para):
            if image_counter < len(images):
                flush_chapitre()
                img = images[image_counter]
                image_counter += 1
                nb_images += 1
                iid_img = f'img{nb_images:03d}'
                add(iid_img, f'images/{img["nom"]}', img['media'], img['data'])
                idx_planche += 1
                pid = f'planche{idx_planche:03d}'
                corps = (f'<div class="planche"><p><img src="images/{img["nom"]}" '
                         f'alt="illustration"/></p></div>')
                add_page(pid, f'{pid}.xhtml', f'Illustration {idx_planche}', corps)

        elif style_name in ('CorpsTexte', 'Normal') and texte:
            if chapitre_courant is not None:
                html_body = runs_to_html(para)
                if html_body:
                    classe = 'first' if len(blocs_chapitre) == 1 else ''
                    blocs_chapitre.append(f'<p class="{classe}">{html_body}</p>')

    flush_chapitre()

    # ── Pages finales ──
    
    # 10. Postface
    postface = infos.get('postface') or infos.get('Postface')
    if postface:
        postface_path = os.path.join(DOSSIER_CHAPITRES, f"{postface}.md")
        if os.path.exists(postface_path):
            contenu_postface = lire_fichier_markdown(postface_path)
            corps_postface = '<div class="postface"><h1>Postface</h1>' + convertir_markdown_html(contenu_postface) + '</div>'
            add_page('postface', 'postface.xhtml', 'Postface', corps_postface, 'Postface')
    
    # 11. Remerciements
    remerciements = infos.get('remerciements') or infos.get('Remerciements')
    if remerciements:
        corps_remerciements = '<div class="remerciements"><h1>Remerciements</h1>' + ''.join(
            f'<p>{html_escape(l)}</p>' for l in remerciements.splitlines() if l.strip()) + '</div>'
        add_page('remerciements', 'remerciements.xhtml', 'Remerciements', corps_remerciements, 'Remerciements')
    
    # ── Page TDM HTML visible (si option Sommaire cochée) ──
    if _ax0.get('sommaire', True):
        corps = '<div class="toc-page"><h1>Table des matières</h1><ul>'
        for iid, lib in toc:
            href = next(i['href'] for i in items if i['id'] == iid)
            corps += f'<li><a href="{href}">{html_escape(lib)}</a></li>'
        corps += '</ul></div>'
        add_page('tdm', 'tdm.xhtml', 'Table des matières', corps, 'Table des matières')
        # Réordonnancement de l'ordre de lecture : épigraphe -> TDM -> préface
        if 'tdm' in spine:
            spine.remove('tdm')
            if 'preface' in spine:
                spine.insert(spine.index('preface'), 'tdm')
            elif 'epigraphe' in spine:
                spine.insert(spine.index('epigraphe') + 1, 'tdm')
    
    # ── nav.xhtml (EPUB 3) ──
    nav = ('<?xml version="1.0" encoding="utf-8"?>\n<!DOCTYPE html>\n'
           '<html xmlns="http://www.w3.org/1999/xhtml" '
           'xmlns:epub="http://www.idpf.org/2007/ops" xml:lang="fr">\n'
           '<head><title>Table des matières</title></head>\n'
           '<body><nav epub:type="toc" id="toc"><h1>Table des matières</h1><ol>')
    for iid, lib in toc:
        href = next(i['href'] for i in items if i['id'] == iid)
        nav += f'<li><a href="{href}">{html_escape(lib)}</a></li>'
    nav += '</ol></nav></body></html>'
    add('nav', 'nav.xhtml', 'application/xhtml+xml', nav.encode('utf-8'), 'nav')

    # ── toc.ncx (navigation logique) ──
    ncx = ('<?xml version="1.0" encoding="utf-8"?>\n'
           '<ncx xmlns="http://www.daisy.org/z3986/2005/ncx/" version="2005-1">\n'
           '<head>'
           f'<meta name="dtb:uid" content="{html_escape(ident)}"/>'
           '<meta name="dtb:depth" content="1"/>'
           '<meta name="dtb:totalPageCount" content="0"/>'
           '<meta name="dtb:maxPageNumber" content="0"/>'
           '</head>'
           f'<docTitle><text>{html_escape(titre_complet)}</text></docTitle>\n'
           '<navMap>')
    for n, (iid, lib) in enumerate(toc, 1):
        href = next(i['href'] for i in items if i['id'] == iid)
        ncx += (f'<navPoint id="np{n}" playOrder="{n}"><navLabel>'
                f'<text>{html_escape(lib)}</text></navLabel>'
                f'<content src="{href}"/></navPoint>')
    ncx += '</navMap></ncx>'
    add('ncx', 'toc.ncx', 'application/x-dtbncx+xml', ncx.encode('utf-8'))

    # ── content.opf ──
    modifie = datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')
    opf = ('<?xml version="1.0" encoding="utf-8"?>\n'
           '<package xmlns="http://www.idpf.org/2007/opf" version="3.0" '
           'unique-identifier="book-id" xml:lang="fr">\n<metadata '
           'xmlns:dc="http://purl.org/dc/elements/1.1/">\n'
           f'<dc:identifier id="book-id">{html_escape(ident)}</dc:identifier>\n'
           f'<dc:title>{html_escape(titre_complet)}</dc:title>\n'
           '<dc:language>fr</dc:language>\n'
           f'<dc:creator>{html_escape(auteur)}</dc:creator>\n'
           f'<dc:date>{annee}</dc:date>\n'
           f'<dc:description>Ebook de {html_escape(auteur)}.</dc:description>\n'
           f'<meta property="dcterms:modified">{modifie}</meta>\n')
    if cover_data:
        opf += '<meta name="cover" content="cover-image"/>\n'
    opf += '</metadata>\n<manifest>\n'
    for i in items:
        props = f' properties="{i["props"]}"' if i.get('props') else ''
        opf += f'<item id="{i["id"]}" href="{i["href"]}" media-type="{i["media"]}"{props}/>\n'
    opf += '</manifest>\n<spine toc="ncx">\n'
    for idref in spine:
        opf += f'<itemref idref="{idref}"/>\n'
    opf += '</spine>\n</package>'

    # ── validation XML de tous les documents avant écriture ──
    erreurs = []
    for i in items:
        if i['media'] in ('application/xhtml+xml', 'application/x-dtbncx+xml'):
            err = valider_xml(i['href'], i['data'])
            if err:
                erreurs.append(err)
    err = valider_xml('content.opf', opf.encode('utf-8'))
    if err:
        erreurs.append(err)
    if erreurs:
        for e in erreurs:
            print(f'   ⚠️  XML invalide → {e}')
        raise ValueError('Documents XML invalides : génération interrompue.')

    # ── Écriture ZIP (mimetype en premier, non compressé) ──
    with zipfile.ZipFile(sortie, 'w') as zf:
        zi = zipfile.ZipInfo('mimetype')
        zi.compress_type = zipfile.ZIP_STORED
        zf.writestr(zi, 'application/epub+zip')
        zf.writestr('META-INF/container.xml', CONTAINER_XML, zipfile.ZIP_DEFLATED)
        zf.writestr('OEBPS/content.opf', opf, zipfile.ZIP_DEFLATED)
        for i in items:
            zf.writestr('OEBPS/' + i['href'], i['data'], zipfile.ZIP_DEFLATED)

    return {'chapitres': nb_chapitres, 'actes': nb_actes,
            'images': nb_images, 'pages': len(spine)}


# ─────────────────────────────────────────────
# 7. AUTO-CONTRÔLE
# ─────────────────────────────────────────────
def verifier_epub(chemin):
    ok = True
    try:
        taille = os.path.getsize(chemin) / (1024 * 1024)
        print(f'   ✅ Taille : {taille:.1f} Mo' if taille <= 50
              else f'   ⚠️  Fichier volumineux ({taille:.1f} Mo)')
        ok = ok and taille <= 50

        with zipfile.ZipFile(chemin, 'r') as z:
            noms = z.namelist()

            if noms and noms[0] == 'mimetype' \
                    and z.read('mimetype') == b'application/epub+zip':
                print('   ✅ mimetype conforme (1er, non compressé)')
            else:
                print('   ⚠️  mimetype non conforme')
                ok = False

            if 'OEBPS/toc.ncx' in noms:
                print('   ✅ TDM NCX présente')
            else:
                print('   ⚠️  TDM NCX absente')
                ok = False

            if 'OEBPS/nav.xhtml' in noms:
                print('   ✅ Nav EPUB 3 présent')
            else:
                print('   ⚠️  Nav EPUB 3 absent')
                ok = False

            if 'OEBPS/cover.jpg' in noms:
                print('   ✅ Couverture présente')
            else:
                print('   ⚠️  Couverture absente')
                ok = False

            imgs = [n for n in noms if n.startswith('OEBPS/images/')]
            print(f'   ✅ {len(imgs)} image(s) intégrée(s)' if imgs
                  else '   ⚠️  Aucune image intégrée')
            ok = ok and bool(imgs)

            nb_xml_bad = 0
            for n in noms:
                if n.endswith(('.xhtml', '.ncx', '.opf')):
                    if valider_xml(n, z.read(n)):
                        nb_xml_bad += 1
            if nb_xml_bad:
                print(f'   ⚠️  {nb_xml_bad} document(s) XML invalide(s)')
                ok = False
            else:
                print('   ✅ Tous les documents XML sont valides')
    except Exception as e:
        print(f'   ⚠️  Vérification impossible : {e}')
        ok = False
    return ok


# ─────────────────────────────────────────────
# 8. POINT D'ENTRÉE
# ─────────────────────────────────────────────
def main(docx_override=None):
    print('=' * 58)
    print('📱  GÉNÉRATION EBOOK KDP (EPUB 3)')
    print('=' * 58)
    t0 = time.time()

    docx = docx_override or trouver_docx()
    if not docx or not os.path.isfile(docx):
        print('   ❌ Aucun fichier *_KDP.docx trouvé.')
        print('   ℹ️  Lancez d\'abord « Générer le livre » dans l\'interface.')
        return 1
    print(f'   📄 Source : {os.path.basename(docx)}')

    infos = lire_infos()
    print(f'   📚 Configuration : {infos.get("_source", "?")} '
          f'({infos[TITRE] or "—"})')

    print('   🖼️  Extraction des illustrations…')
    images = extraire_images(docx)
    print(f'   ✅ {len(images)} image(s) extraite(s)')

    slug_titre = re.sub(r'[^\w]+', '_', (infos[TITRE] or 'roman').strip())
    if infos[SOUS_TITRE]:
        slug_titre += '_' + re.sub(r'[^\w]+', '_', infos[SOUS_TITRE].strip())
    dossier_sortie = os.path.join(BASE, 'export')
    os.makedirs(dossier_sortie, exist_ok=True)
    sortie = os.path.join(dossier_sortie, f'{slug_titre}_ebook.epub')

    print('   📖 Construction de l\'EPUB…')
    try:
        stats = construire_ebook(docx, infos, images, sortie)
    except ValueError as e:
        print(f'   ❌ {e}')
        return 1
    print(f'   📊 Actes : {stats["actes"]} · Chapitres : {stats["chapitres"]} · '
          f'Images : {stats["images"]}')

    print('   🔎 Auto-contrôle…')
    ok = verifier_epub(sortie)

    duree = time.time() - t0
    if ok:
        print('✅ Ebook généré avec succès !')
        print(f'   📁 {sortie}')
        print(f'   ⏱️  Durée : {duree:.1f} s')
        print('   ℹ️  À uploader sur KDP comme « Contenu de l\'ebook ».')
        return 0
    print('⚠️  Ebook généré avec des avertissements.')
    return 1


# ── v1.6 : PLUS DE sys.exit() ici ──
# Le code de retour est exposé via EBOOK_CODE : le thread de l'interface
# se termine donc toujours proprement et l'interface se déverrouille.
if __name__ == '__main__':
    if '--docx' in sys.argv:
        idx = sys.argv.index('--docx')
        if idx + 1 < len(sys.argv):
            EBOOK_CODE = main(sys.argv[idx + 1])
    elif '--cover-only' in sys.argv:
        couverture_octets()
        EBOOK_CODE = 0
    else:
        EBOOK_CODE = main()
def _normaliser_cles(d):
    """Recadre les clés (espaces finaux du JSON du menu Informations)."""
    if not isinstance(d, dict):
        return d
    return {str(k).strip(): (_normaliser_cles(v) if isinstance(v, dict) else v)
            for k, v in d.items()}

def _infos_de(d):
    """Retourne le sous-dictionnaire 'informations' normalisé (tolérant)."""
    d = _normaliser_cles(d or {})
    v = d.get('informations')
    return v if isinstance(v, dict) else {}
