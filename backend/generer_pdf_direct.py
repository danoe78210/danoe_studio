#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generer_pdf_direct.py – Moteur PDF KDP SANS Word (v3.0)
v3.0 – ANTI-LayoutError :
- nettoyage élargi de TOUS les espaces insécables / caractères invisibles ;
- padding de 2 pt dans les Frame (évite le cas « largeur == cadre ») ;
- si un LayoutError survient quand même : repasse automatique en mode
  sécurisé (wordWrap CJK + padding 4 pt) qui coupe n'importe quel bloc.
Dépendances : pip install reportlab pypdf Pillow openpyxl
Usage : python generer_pdf_direct.py
"""
import os, re, sys, glob, io, json
BASE = os.path.dirname(os.path.abspath(__file__))
CHEMIN_CONFIG = os.path.join(BASE, 'Configuration_roman.xlsx')
CHEMIN_CONFIG_JSON = os.path.join(BASE, 'Configuration_roman.json')
DOSSIER_IMAGES = os.path.join(BASE, 'Images')
DOSSIER_CHAPITRES = os.path.join(BASE, 'Chapitres')
DOSSIER_CACHE_HD = os.path.join(BASE, '_cache_HD')
SANS_TITRE = ('1.1',)

from xml.sax.saxutils import escape
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import (BaseDocTemplate, PageTemplate, Frame, Paragraph,
                                Spacer, PageBreak, Image as RLImage)
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from pypdf import PdfReader, PdfWriter
import regles as _R

FORMATS_LIVRE = [('5 x 8 po', 12.7, 20.32), ('5.06 x 7.81 po', 12.85, 19.84),
                 ('5.25 x 8 po', 13.34, 20.32), ('5.5 x 8.5 po', 13.97, 21.59),
                 ('6 x 9 po', 15.24, 22.86), ('6.14 x 9.21 po', 15.6, 23.4),
                 ('7 x 10 po', 17.78, 25.4), ('8 x 10 po', 20.32, 25.4),
                 ('8.5 x 8.5 po', 21.59, 21.59), ('8.5 x 11 po', 21.59, 27.94),
                 ('A4', 21.0, 29.7)]
KDP_MARGES = [(150, .375, .250), (300, .5, .313), (500, .625, .375),
              (700, .75, .5), (828, .875, .625)]

def nettoyer(v): return str(v).strip().strip('`').strip('*').strip() if v is not None else ''

def _lire_json():
    try:
        with open(CHEMIN_CONFIG_JSON, encoding='utf-8') as f:
            return _normaliser_cles(json.load(f))
    except Exception:
        return None

# ── lecture config (JSON prioritaire, repli Excel) ──
def lire_style():
    s = {'format': '7 x 10 po', 'police_corps': 'Aptos', 'taille_corps': 11.0,
         'police_titres': 'Cinzel', 'taille_acte': 16.0, 'taille_chap1': 14.0,
         'taille_chap2': 13.0, 'taille_sous': 12.0, 'interligne': 1.0}
    j = _lire_json()
    st = _get_section(j or {}, 'style', {}) or {}
    if st:
        s['format'] = str(st.get('format_livre') or s['format'])
        s['police_corps'] = str(st.get('police_corps') or s['police_corps'])
        s['taille_corps'] = float(str(st.get('taille_corps_pt', 11)).replace(',', '.'))
        s['police_titres'] = str(st.get('police_titres') or s['police_titres'])
        s['taille_acte'] = float(st.get('taille_titres_acte_pt', 16))
        s['taille_chap1'] = float(st.get('taille_chapitre_ligne1_pt', 14))
        s['taille_chap2'] = float(st.get('taille_chapitre_ligne2_pt', 13))
        s['taille_sous'] = float(st.get('taille_sous_chapitre_pt', 12))
        s['interligne'] = float(str(st.get('interligne_corps', 1)).replace(',', '.'))
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CHEMIN_CONFIG, data_only=True)
            if 'Style' in wb.sheetnames:
                vals = {}
                for row in wb['Style'].iter_rows(values_only=True):
                    if row and row[0]: vals[nettoyer(row[0]).lower()] = row[1]
                def g(p):
                    for k, v in vals.items():
                        if k.startswith(p) and v not in (None, ''): return v
                s['format'] = str(g('format') or s['format'])
                s['taille_corps'] = float(str(g('taille du corps') or 11).replace(',', '.'))
                s['interligne'] = float(str(g('interligne') or 1).replace(',', '.'))
        except Exception: pass
    lab = s['format'].lower()
    for nom, w, h in FORMATS_LIVRE:
        if nom.lower() in lab: s['largeur_cm'], s['hauteur_cm'] = w, h; break
    else:
        m = re.search(r'(\d+(?:\.\d+)?)\s*x\s*(\d+(?:\.\d+)?)', lab)
        w, h = (float(m.group(1)), float(m.group(2))) if m else (7, 10)
        if 'cm' in lab: s['largeur_cm'], s['hauteur_cm'] = w, h
        else: s['largeur_cm'], s['hauteur_cm'] = w * 2.54, h * 2.54
    s['largeur_po'], s['hauteur_po'] = s['largeur_cm'] / 2.54, s['hauteur_cm'] / 2.54
    return s

def _get_infos(d):
    d = d or {}
    if 'informations' in d: return d['informations'] or {}
    if 'informations ' in d: return d['informations '] or {}
    for k, v in d.items():
        if str(k).strip() == 'informations' and isinstance(v, dict): return v
    return {}


def _get_section(d, nom, defaut=None):
    """Recherche tolérante d'une section JSON (clés avec espaces finaux)."""
    if defaut is None: defaut = {}
    if not isinstance(d, dict): return defaut
    if nom in d and d[nom] is not None: return d[nom]
    for k, v in d.items():
        if str(k).strip() == nom and v is not None: return v
    return defaut

def _section_infos(d):
    return _get_section(d, 'informations', {}) or {}

def lire_infos():
    def _norm(d):
        if not isinstance(d, dict):
            return d
        return {str(k).strip(): (_norm(v) if isinstance(v, dict) else v)
                for k, v in d.items()}
    infos = {}
    j = _norm(_lire_json() or {})
    ji = j.get('informations')
    ji = ji if isinstance(ji, dict) else {}
    for cle, val in ji.items():
        if val:
            infos[cle.lower()] = nettoyer(val)
    if not infos:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CHEMIN_CONFIG, data_only=True)
            ws = wb['Informations']
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    infos[nettoyer(row[0]).lower()] = nettoyer(row[1])
        except Exception:
            pass
    return infos

def lire_annexes():
    def _norm(d):
        if not isinstance(d, dict):
            return d
        return {str(k).strip(): (_norm(v) if isinstance(v, dict) else v)
                for k, v in d.items()}
    ax = {'sommaire': True}
    j = _norm(_lire_json() or {})
    ji = j.get('informations')
    ji = ji if isinstance(ji, dict) else {}
    def g(*cles):
        for c in cles:
            v = ji.get(c)
            if v:
                return str(v).strip()
        return ''
    ax['editeur'] = g('editeur', 'Éditeur', 'Editeur', 'maison_edition')
    ax['autres_livres'] = g('autres_livres', 'Autres livres du même auteur')
    ax['remerciements'] = g('remerciements', 'Remerciements')
    ax['frontispice'] = g('frontispice', 'Frontispice')
    ax['preface'] = g('preface', 'Préface')
    ax['postface'] = g('postface', 'Postface')
    sv = ji.get('sommaire')
    ax['sommaire'] = sv if isinstance(sv, bool) else (str(sv).strip().lower() != 'false' if sv is not None else True)
    return ax

def _md_to_html(txt):
    out = []
    for ln in txt.split('\n'):
        ln = ln.rstrip()
        if not ln.strip() or ln.startswith('# '): continue
        if ln.startswith('## '): out.append('<h2>%s</h2>' % ln[3:].strip())
        else: out.append('<p>%s</p>' % ln)
    return ''.join(out)

def lire_organisation():
    j = _lire_json()
    chap = _get_section(j or {}, 'chapitres', []) or []
    if chap:
        flux = []
        for it in chap:
            t = it.get('type', '')
            if t == 'image' and it.get('image'):
                flux.append({'type': 'image', 'image': it['image'], 'legende': it.get('legende', '')})
            elif t == 'acte' and it.get('acte'):
                flux.append({'type': 'acte', 'acte': it['acte']})
            elif t == 'chapitre' and it.get('fichier_source'):
                l1 = it.get('chapitre_ligne1', '') or ''; l2 = it.get('chapitre_ligne2', '') or ''
                flux.append({'type': 'chapitre', 'fichier': it['fichier_source'],
                             'titre': (l1 + ' ' + l2).strip() if l2 else l1})
        if flux: return flux
    if not os.path.isfile(CHEMIN_CONFIG):
        return []
    from openpyxl import load_workbook
    wb = load_workbook(CHEMIN_CONFIG, data_only=True)
    ws = wb['Chapitres']
    cols = {'fichier': 0, 'acte': 1, 'l1': 2, 'l2': 3, 'image': 4, 'legende': 5}
    flux = []
    for row in ws.iter_rows(values_only=True):
        cells = [nettoyer(c) for c in row]
        if not any(cells): continue
        low = [c.lower() for c in cells]
        if low[0].startswith('fichier'):
            for i, n in enumerate(low):
                if n.startswith('fichier'): cols['fichier'] = i
                elif n.startswith('acte'): cols['acte'] = i
                elif 'ligne 1' in n: cols['l1'] = i
                elif 'ligne 2' in n: cols['l2'] = i
                elif n.startswith('image'): cols['image'] = i
                elif 'égende' in n: cols['legende'] = i
            continue
        def g(c): return cells[cols[c]] if cols[c] < len(cells) else ''
        if g('fichier').startswith('---'): continue
        if g('image') and not g('fichier') and not g('acte'):
            flux.append({'type': 'image', 'image': g('image'), 'legende': g('legende')}); continue
        if not g('fichier') and g('acte'):
            flux.append({'type': 'acte', 'acte': g('acte')}); continue
        if g('fichier'):
            t = (g('l1') + ' ' + g('l2')).strip() if g('l2') else g('l1')
            flux.append({'type': 'chapitre', 'fichier': g('fichier'), 'titre': t})
    return flux

# ── corrections typo ──
CORRECTIONS_COMMUNES = [('\u2060', ''), ('\ufeff', ''), ('--- ', '\u2014 '),
                        (' \u2026', '\u2026'), ("'", '\u2019'), ('pu is', 'puis'),
                        ('. . ', '. '), ('.À', '. À'), ('cératopsiens ,', 'cératopsiens,'),
                        ("de s'approchaient", "de s'approcher"), ('ils émettait', 'ils émettaient')]
# v2.0 : espaces SÉCABLES autour des guillemets (évite les blocs incassables)
REGEX_PENSEES = [(re.compile(r'"([^"]+)"'), '« \\1 »')]
CORR_PAR_PREFIXE = {'1.1': [], '2.1': [], '2.2': []}
META_RE = re.compile(r'^(titre|œuvre|oeuvre|auteur|pages_pdf|nombre_de_pages|source|'
                     r'version|d[ée]p[ôo]t_l[ée]gal|langue)\s*:', re.IGNORECASE)

def _nettoie_espaces(t):
    """v3.0 : supprime TOUS les espaces insécables / caractères invisibles (anti-LayoutError)."""
    for c in ('\u00a0', '\u202f', '\u2007', '\u2008', '\u2009', '\u200a'):
        t = t.replace(c, ' ')
    for c in ('\u200b', '\u200c', '\u200d', '\u2060', '\ufeff'):
        t = t.replace(c, '')
    t = t.replace('\u2011', '-')
    return t

def charger_chapitre(fichier, skip_titre):
    cand = glob.glob(os.path.join(DOSSIER_CHAPITRES, fichier + '*.md'))
    if not cand: return None
    texte = open(cand[0], encoding='utf-8').read()
    for a, b in CORRECTIONS_COMMUNES: texte = texte.replace(a, b)
    for a, b in _R.corrections_pour(fichier): texte = texte.replace(a, b)
    for pref, corr in CORR_PAR_PREFIXE.items():
        if fichier.startswith(pref):
            for a, b in corr: texte = texte.replace(a, b)
    texte = _nettoie_espaces(texte)
    while '  ' in texte: texte = texte.replace('  ', ' ')
    items = []
    for l in [x.strip() for x in texte.splitlines()]:
        if not l or META_RE.match(l) or l == skip_titre: continue
        if l.startswith('## '): items.append(('h2', l[3:].strip()))
        elif l.startswith('# '): items.append(('h1', l[2:].strip()))
        elif l in ('---', '***', '___'): items.append(('sep', None))
        else:
            if not l.startswith('\u2014'):
                for rx, rp in REGEX_PENSEES: l = rx.sub(rp, l)
            items.append(('p', l))
    while items and items[0][0] == 'sep': items.pop(0)
    return items

# ── images HD N&B ──
def image_haute_definition(chemin, largeur_cm):
    try:
        from PIL import Image as PIL
    except Exception: return chemin
    try:
        os.makedirs(DOSSIER_CACHE_HD, exist_ok=True)
        cache = os.path.join(DOSSIER_CACHE_HD,
                             os.path.splitext(os.path.basename(chemin))[0] +
                             f'_HD{int(os.path.getmtime(chemin))}.png')
        if not os.path.isfile(cache):
            with PIL.open(chemin) as im:
                im.load(); pw, ph = im.size
                cible = int(round((largeur_cm / 2.54) * 300))
                if cible > pw:
                    f = cible / pw
                    im = im.resize((cible, int(ph * f)), PIL.LANCZOS)
                im.convert('L').save(cache, format='PNG', dpi=(300, 300))
        return cache
    except Exception: return chemin

# ── polices TTF (repli Times) ──
def registre_polices():
    fdir = os.path.join(os.environ.get('WINDIR', r'C:\Windows'), 'Fonts')
    def pick(c):
        for n in c:
            p = os.path.join(fdir, n)
            if os.path.isfile(p): return p
    corps, corps_b = pick(['Aptos.ttf', 'times.ttf']), pick(['Aptos-Bold.ttf', 'timesbd.ttf'])
    corps_i, corps_bi = pick(['Aptos-Italic.ttf', 'timesi.ttf']), pick(['Aptos-BoldItalic.ttf', 'timesbi.ttf'])
    titres = pick(['Cinzel.ttf', 'Cinzel-Bold.ttf', 'timesbd.ttf'])
    F = {}
    if corps and corps_b and corps_i and corps_bi:
        pdfmetrics.registerFont(TTFont('Corps', corps)); pdfmetrics.registerFont(TTFont('CorpsB', corps_b))
        pdfmetrics.registerFont(TTFont('CorpsI', corps_i)); pdfmetrics.registerFont(TTFont('CorpsBI', corps_bi))
        F = {'c': 'Corps', 'b': 'CorpsB', 'i': 'CorpsI', 'bi': 'CorpsBI'}
    else:
        F = {'c': 'Times-Roman', 'b': 'Times-Bold', 'i': 'Times-Italic', 'bi': 'Times-BoldItalic'}
    if titres:
        pdfmetrics.registerFont(TTFont('Titres', titres)); F['t'] = 'Titres'
    else:
        F['t'] = F['b']
    return F

# ── gabarit (v3.0 : padding anti-fuzz dans les Frame) ──
class DocKDP(BaseDocTemplate):
    def __init__(self, buf, W, H, marges, params):
        super().__init__(buf, pagesize=(W, H))
        self.W, self.H, self.p = W, H, params
        g, o, b, t = marges
        w_txt = W - g - o
        pad = params.get('pad', 2.0)
        self.addPageTemplates([
            PageTemplate(id='impair',
                         frames=[Frame(g, b, w_txt, H - b - t,
                                       leftPadding=pad, rightPadding=pad, topPadding=0, bottomPadding=0)],
                         onPage=self._dessin),
            PageTemplate(id='pair',
                         frames=[Frame(o, b, w_txt, H - b - t,
                                       leftPadding=pad, rightPadding=pad, topPadding=0, bottomPadding=0)],
                         onPage=self._dessin)])

    def handle_pageBegin(self):
        n = getattr(self, 'page', 0) + 1 + self.p['offset']
        self.pageTemplate = self.pageTemplates[0 if n % 2 == 1 else 1]
        super().handle_pageBegin()

    def _dessin(self, canvas, doc):
        n_abs = doc.page + self.p['offset']
        canvas.saveState()
        if self.p['entete'] and doc.page > 1:
            canvas.setFont(self.p['F']['i'], 9)
            canvas.drawCentredString(self.W / 2, self.H - 1.0 * cm, self.p['entete'])
            canvas.setLineWidth(0.5)
            canvas.line(self.W / 2 - 3 * cm, self.H - 1.0 * cm - 4,
                        self.W / 2 + 3 * cm, self.H - 1.0 * cm - 4)
        if n_abs >= self.p['debut_num']:
            canvas.setFont(self.p['F']['c'], 10)
            canvas.drawCentredString(self.W / 2, 1.0 * cm, str(n_abs - self.p['debut_num'] + 1))
        canvas.restoreState()

def rendre(story, W, H, marges, params, offset):
    params = dict(params); params['offset'] = offset
    buf = io.BytesIO()
    DocKDP(buf, W, H, marges, params).build(story)
    return buf

def slug(s): return re.sub(r'[^\w]+', '_', s.strip(), flags=re.UNICODE).strip('_') or 'roman'

def generer(safe=False):
    STYLE = lire_style(); INFOS = lire_infos(); ORG = lire_organisation()
    F = registre_polices()
    TC = STYLE['taille_corps']; IL = STYLE['interligne']
    W, H = STYLE['largeur_po'] * 72, STYLE['hauteur_po'] * 72
    lead = TC * 1.2 * IL
    ww = 'CJK' if safe else None
    st_corps = ParagraphStyle('c', fontName=F['c'], fontSize=TC, leading=lead,
                              alignment=TA_JUSTIFY, spaceBefore=TC * 0.8, wordWrap=ww)
    st_debut = ParagraphStyle('d', parent=st_corps, firstLineIndent=0.5 * cm, spaceBefore=0)
    st_acte = ParagraphStyle('a', fontName=F['t'], fontSize=STYLE['taille_acte'],
                             alignment=TA_CENTER, spaceBefore=12, spaceAfter=12)
    st_ch1 = ParagraphStyle('h1', fontName=F['t'], fontSize=STYLE['taille_chap1'], alignment=TA_CENTER)
    st_ch2 = ParagraphStyle('h2', fontName=F['t'], fontSize=STYLE['taille_chap2'],
                            alignment=TA_CENTER, spaceBefore=6)
    st_sous = ParagraphStyle('s', fontName=F['b'], fontSize=STYLE['taille_sous'],
                             alignment=TA_CENTER, spaceBefore=12, spaceAfter=12)
    st_sep = ParagraphStyle('sep', fontName=F['c'], fontSize=TC, alignment=TA_CENTER,
                            spaceBefore=TC, spaceAfter=TC)
    st_leg = ParagraphStyle('leg', fontName=F['i'], fontSize=9, alignment=TA_CENTER, spaceBefore=6)
    st_lim = ParagraphStyle('lim', fontName=F['c'], fontSize=9, leading=12)
    st_toc = ParagraphStyle('toc', fontName=F['c'], fontSize=10, leading=16)

    def para(texte, style, gras_debut=False):
        runs = re.findall(r'(\*\*[^*]+?\*\*|\*[^*]+?\*|_[^_]+?_|[^*]+)', texte)
        html = ''
        for r in runs:
            if r.startswith('**'): html += f'<b>{escape(r[2:-2])}</b>'
            elif r.startswith(('*', '_')): html += f'<i>{escape(r[1:-1])}</i>'
            else: html += escape(r)
        if gras_debut and html:
            m = re.match(r'(<[^>]+>)*', html)
            reste = html[m.end():]
            if reste:
                html = (m.group(0) + f'<font name="{F["b"]}" size="{TC + 3}">{reste[0]}</font>' + reste[1:])
        return Paragraph(html, style)

    # ── estimation marges KDP ──
    mots = 0
    for b in ORG:
        if b['type'] == 'chapitre':
            it = charger_chapitre(re.match(r'(\d+\.\d+)', b['fichier']).group(1), b['titre'])
            mots += sum(len(t.split()) for k, t in (it or []) if k == 'p')
    pages_est = 120.0
    for _ in range(6):
        for maxi, g, o in KDP_MARGES:
            if pages_est <= maxi: gut, out = g, o; break
        else: gut, out = .875, .625
        pages_est = 12 + mots / (((STYLE['largeur_po'] - gut - out) / 0.58) *
                                 ((STYLE['hauteur_po'] - 1.5) / (0.21 * IL)))
    for maxi, g, o in KDP_MARGES:
        if pages_est <= maxi: gut, out = g, o; break
    else: gut, out = .875, .625
    m_sym = gut + 0.125; MARGES = (m_sym * 72, m_sym * 72, 1.9 * cm, 1.9 * cm)
    PAD = 4.0 if safe else 2.0
    print(f'   📏 Marges KDP : gouttière {gut:.3f} po · extérieure {out:.3f} po (~{int(pages_est)} pages)'
          + ('  [mode sécurisé]' if safe else ''))


    # ── construction des segments ──
    segments = []
    _ax = lire_annexes()
    _tit = INFOS.get('titre complet du roman', '') or 'Titre'
    _sous = INFOS.get('sous-titre éventuel', '') or ''
    _aut = INFOS.get("nom de l'auteur (couverture)", '') or 'Auteur'
    _ann = INFOS.get('année de publication', '2026')
    _larg_utile = STYLE['largeur_cm'] * cm - 4 * cm
    _haut_utile = H - 3.8 * cm
    lim = []
    # 1. Page de garde (blanche)
    lim += [Spacer(1, 1 * cm), PageBreak()]
    # 2. Faux-titre (titre seul, petits caractères)
    lim += [Spacer(1, 7 * cm), Paragraph(escape(_tit), st_lim), PageBreak()]
    # 3. Frontispice (image)
    if _ax.get('frontispice'):
        _fp = os.path.join(BASE, 'Images', _ax['frontispice'])
        if not os.path.isfile(_fp):
            for _ext in ('.png', '.jpg', '.jpeg', '.webp', '.bmp'):
                if os.path.isfile(_fp + _ext): _fp += _ext; break
        if os.path.isfile(_fp):
            try:
                _img = RLImage(_fp, width=12 * cm, height=16 * cm); _img.hAlign = 'CENTER'
                lim += [Spacer(1, 2 * cm), _img]
            except Exception: pass
        lim.append(PageBreak())
    # 4. Page de titre (titre, sous-titre, auteur, éditeur)
    lim.append(Paragraph(escape(_tit.upper()), st_acte))
    if _sous: lim.append(Paragraph(escape(_sous), st_ch2))
    lim += [Spacer(1, 2 * cm), Paragraph(escape(_aut), st_ch2)]
    if _ax.get('editeur'):
        lim += [Spacer(1, 1 * cm), Paragraph(escape(_ax['editeur']), st_lim)]
    lim.append(PageBreak())
    # 5. Copyright (verso : mentions, ISBN, dépôt légal, édition)
    _cop = [Paragraph(escape((_tit + ' – ' + _sous) if _sous else _tit), st_lim)]
    _cop.append(Spacer(1, 0.6 * cm))
    _cop.append(Paragraph(escape(INFOS.get('mention de copyright', '') or ('© ' + _ann + ' ' + _aut + '. Tous droits réservés.')), st_lim))
    _cop.append(Spacer(1, 0.3 * cm))
    _meta = ' – '.join(x for x in [INFOS.get('édition', ''), _ann] if x)
    if _meta: _cop.append(Paragraph(escape(_meta), st_lim))
    if INFOS.get('isbn'): _cop.append(Paragraph('ISBN : ' + escape(INFOS['isbn']), st_lim))
    _cop.append(Paragraph('Dépôt légal : ' + _ann, st_lim))
    if _ax.get('editeur'): _cop.append(Paragraph(escape(_ax['editeur']), st_lim))
    _cop.append(Spacer(1, 0.6 * cm))
    for _t in ['Toute reproduction, même partielle, est interdite sans l’autorisation',
               'préalable de l’auteur, conformément aux dispositions de la législation',
               'en vigueur sur la propriété intellectuelle.']:
        _cop.append(Paragraph(escape(_t), st_lim))
    if INFOS.get('site web'):
        _cop += [Spacer(1, 0.3 * cm), Paragraph(escape(INFOS['site web']), st_lim)]
    _h_cop = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _cop)
    lim.append(Spacer(1, max(0, _haut_utile - _h_cop)))
    lim += _cop
    lim.append(PageBreak())
    # 6. Dédicace (page impaire dédiée)
    if INFOS.get('dédicace'):
        _ded = [Paragraph(escape(x), st_ch2) for x in INFOS['dédicace'].splitlines() if x.strip()]
        _h_d = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _ded)
        lim.append(Spacer(1, max(0, (_haut_utile - _h_d) / 2)))
        lim += _ded
        lim.append(PageBreak())
    # 7. Épigraphe
    if INFOS.get('épigraphe'):
        _epi = [Paragraph(escape(x), st_ch2) for x in INFOS['épigraphe'].splitlines() if x.strip()]
        _h_e = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _epi)
        lim.append(Spacer(1, max(0, (_haut_utile - _h_e) / 2)))
        lim += _epi
        lim.append(PageBreak())
    segments.append({'type': 'lim', 'story': lim, 'entete': None})
    # 8. Préface (segment séparé ; la TDM sera insérée juste avant)
    if _ax.get('preface'):
        _pc = os.path.join(BASE, 'Chapitres', _ax['preface'] + '.md')
        if os.path.isfile(_pc):
            _st = [Paragraph('Préface', st_acte)]
            for _ln in open(_pc, encoding='utf-8').read().split('\n'):
                _ln = _ln.strip()
                if _ln and not _ln.startswith('# '): _st.append(Paragraph(escape(_ln), st_lim))
            segments.append({'type': 'preface', 'story': _st, 'entete': None})


    for b in ORG:
        if b['type'] == 'image':
            ch = None
            for ext in ('', '.png', '.jpg'):
                p = os.path.join(DOSSIER_IMAGES, b['image'] + ext)
                if os.path.isfile(p): ch = p; break
            st = []
            if ch:
                max_l = max(5.0, STYLE['largeur_cm'] - 3.0)
                ch_hd = image_haute_definition(ch, max_l)
                from PIL import Image as PIL
                with PIL.open(ch_hd) as im: pw, ph = im.size
                w_cm = min(max_l, 12.0); h_cm = w_cm * ph / pw
                max_h = STYLE['hauteur_cm'] - 3.8
                if h_cm > max_h: h_cm, w_cm = max_h, max_h * pw / ph
                st.append(Spacer(1, max(0, (H - 3.8 * cm - h_cm * cm) / 2)))
                img = RLImage(ch_hd, width=w_cm * cm, height=h_cm * cm); img.hAlign = 'CENTER'; st.append(img)
                if b.get('legende'): st.append(Paragraph(escape(b['legende']), st_leg))
            segments.append({'type': 'image', 'story': st, 'entete': None})
        elif b['type'] == 'acte':
            segments.append({'type': 'acte', 'titre': b['acte'], 'entete': None,
                             'story': [Spacer(1, 4 * cm), Paragraph(escape(b['acte']), st_acte)]})
        else:
            pref = re.match(r'(\d+\.\d+)', b['fichier']).group(1)
            sans = pref in SANS_TITRE
            items = charger_chapitre(pref, b['titre']) or []
            if items and items[0][0] in ('h1', 'h2') and not sans: items.pop(0)
            while items and items[0][0] == 'sep': items.pop(0)
            st = []
            if not sans:
                if ' : ' in b['titre']:
                    l1, l2 = b['titre'].split(' : ', 1)
                    st += [Spacer(1, 3 * cm), Paragraph(escape(l1 + ' :'), st_ch1),
                           Paragraph(escape(l2), st_ch2)]
                else:
                    st += [Spacer(1, 3 * cm), Paragraph(escape(b['titre']), st_ch1)]
            premier = True
            for k, t in items:
                if k == 'h1': st += [PageBreak(), Paragraph(escape(t), st_ch1)]; premier = True
                elif k == 'h2': st += [PageBreak(), Paragraph(escape(t), st_sous)]; premier = True
                elif k == 'sep': st.append(Paragraph('--- ✦ ---', st_sep))
                else:
                    st.append(para(t, st_debut if premier else st_corps, gras_debut=premier))
                    premier = False
            segments.append({'type': 'chapitre', 'titre': b['titre'], 'entete': b['titre'], 'story': st})

        # ── UNE SEULE passe de rendu + fusion (offset connu en avançant) ──
    writer = PdfWriter()
    running = 1
    debut_num = None
    entrees = []
    for s in segments:
        if running % 2 == 0:
            writer.add_blank_page(width=W, height=H); running += 1
        if s['type'] == 'chapitre' and debut_num is None:
            debut_num = running
        if s['type'] in ('acte', 'chapitre'):
            entrees.append((s['titre'], running))
        params = {'F': F, 'entete': s['entete'], 'debut_num': debut_num or 10 ** 9, 'pad': PAD}
        buf = rendre(s['story'], W, H, MARGES, params, running - 1)
        pages = PdfReader(buf)
        for p in pages.pages:
            writer.add_page(p)
        running += len(pages.pages)
    # ── TDM en fin de volume ──
    if running % 2 == 0:
        writer.add_blank_page(width=W, height=H); running += 1
    toc = [Spacer(1, 2 * cm), Paragraph('Table des matières', st_ch1), Spacer(1, 1 * cm)]
    for t, p in entrees:
        pts = max(3, 70 - len(t))
        toc.append(Paragraph(f'{escape(t)} {"." * pts} {p if p < (debut_num or 1) else p - (debut_num or 1) + 1}', st_toc))
    buf = rendre(toc, W, H, MARGES, {'F': F, 'entete': None, 'debut_num': 10 ** 9, 'pad': PAD}, running - 1)
    for p in PdfReader(buf).pages:
        writer.add_page(p)
    titre = INFOS.get('titre complet du roman', '') or 'Roman'
    writer.add_metadata({'/Title': titre, '/Author': INFOS.get("nom de l'auteur (couverture)", ''),
                         '/Creator': 'generer_pdf_direct v3.0'})
    _doss = os.path.join(BASE, 'export'); os.makedirs(_doss, exist_ok=True)
    nom = os.path.join(_doss, slug(titre) + '_KDP.pdf')
    with open(nom, 'wb') as f:
        writer.write(f)
    mb = PdfReader(nom).pages[0].mediabox
    ok = abs(float(mb.width) - W) <= 2 and abs(float(mb.height) - H) <= 2
    print(f'   {"✅" if ok else "⚠️"} MediaBox : {float(mb.width):.0f} x {float(mb.height):.0f} pt '
          f'(attendu {W:.0f} x {H:.0f})')
    print(f'✅ PDF KDP prêt ({len(PdfReader(nom).pages)} pages) : {nom}')

def _convert_word_pdf(docx_path, pdf_path):
    import win32com.client
    word = win32com.client.Dispatch('Word.Application'); word.Visible = False
    try: word.DisplayAlerts = 0; word.AutomationSecurity = 3
    except Exception: pass
    d = word.Documents.Open(os.path.abspath(docx_path), False, True, False)
    try: d.ExportAsFixedFormat(os.path.abspath(pdf_path), 17, False, 0)
    finally: d.Close(False); word.Quit()

def generer_depuis_word():
    import glob, subprocess, time
    c = glob.glob(os.path.join(BASE, '*_KDP.docx')) + glob.glob(os.path.join(BASE, 'export', '*_KDP.docx'))
    if not c: return False
    dx = max(c, key=os.path.getmtime); pf = os.path.splitext(dx)[0] + '.pdf'
    for _ in range(5):
        try:
            if os.path.exists(pf): os.remove(pf)
            break
        except Exception: time.sleep(0.3)
    print('   🖨 Conversion Word→PDF (vos modifs incluses)…')
    try:
        r = subprocess.run([sys.executable, os.path.abspath(__file__), '--convert', dx, pf], timeout=300, capture_output=True, text=True)
        if r.returncode == 0 and os.path.isfile(pf):
            print('✅ PDF KDP prêt (converti depuis Word) : ' + pf); return True
        return False
    except subprocess.TimeoutExpired:
        print('   ⚠️ Word trop lent (>180 s) → génération directe.'); return False

def main():
    if '--direct' not in sys.argv[1:] and generer_depuis_word():
        return

    try:
        generer(safe=False)
    except Exception as e:
        print(f'⚠️ Erreur de mise en page ({e}). Nouvelle passe en mode sécurisé…')
        generer(safe=True)

if __name__ == '__main__':
    if '--convert' in sys.argv[1:]:
        _i = sys.argv.index('--convert')
        _convert_word_pdf(sys.argv[_i + 1], sys.argv[_i + 2])
    else:
        main()
def _normaliser_cles(d):
    """Recadre les clés (espaces finaux du JSON du menu Informations)."""
    if not isinstance(d, dict):
        return d
    return {str(k).strip(): (_normaliser_cles(v) if isinstance(v, dict) else v)
            for k, v in d.items()}

def _infos_de(d):
    """Retourne le sous-dictionnaire 'informations' normalisé (tolérant)."""
    d = _normaliser_cles(d or {})
    v = d.get('informations')
    return v if isinstance(v, dict) else {}
