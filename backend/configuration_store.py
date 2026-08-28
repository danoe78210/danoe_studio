#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
configuration_store.py – Stockage JSON de la configuration (v1.3)

v1.3 : ajout de POLICES_KDP (polices acceptées/recommandées par Amazon KDP)
       et de la clé de style « police_lettrine » (1ʳᵉ lettre des chapitres).
v1.1 : catalogue des formats KDP + dimensions + marges officielles.
"""
import os
import re
import json
import shutil

BASE = os.path.dirname(os.path.abspath(__file__))
CHEMIN_CONFIG_JSON = os.path.join(BASE, 'Configuration_roman.json')
CHEMIN_CONFIG_EXCEL = os.path.join(BASE, 'Configuration_roman.xlsx')

# ─────────────────────────────────────────────
# Catalogue des formats Amazon KDP
# ─────────────────────────────────────────────
FORMATS_KDP = [
    {"label": "6 x 9 po",     "cm": "15,24 x 22,86 cm", "type": "Broché standard",
     "usage": "Le plus populaire : romans, essais, guides"},
    {"label": "5.5 x 8.5 po", "cm": "13,97 x 21,59 cm", "type": "Compact",
     "usage": "Livres courts, carnets, fiction légère"},
    {"label": "5 x 8 po",     "cm": "12,7 x 20,32 cm",  "type": "Petit format",
     "usage": "Livres de poche, nouvelles, journaux intimes"},
    {"label": "7 x 10 po",    "cm": "17,78 x 25,4 cm",  "type": "Grand format",
     "usage": "Livres illustrés, manuels, ouvrages éducatifs"},
    {"label": "8.5 x 8.5 po", "cm": "21,59 x 21,59 cm", "type": "Format carré",
     "usage": "Livres enfants, albums photo, livre créatif"},
    {"label": "8 x 10 po",    "cm": "20,32 x 25,4 cm",  "type": "Grand carré",
     "usage": "Livres images, art, recettes, projets visuels"},
    {"label": "8.5 x 11 po",  "cm": "21,59 x 27,94 cm", "type": "Grand carré",
     "usage": "Livres images, art, recettes, projets visuels"},
]

# ─────────────────────────────────────────────
# v1.3 : polices acceptées / recommandées par Amazon KDP
# (familles supportées par Kindle + polices standard embarquables
#  pour le livre broché). Liste extensible si besoin.
# ─────────────────────────────────────────────
POLICES_KDP = [
    'Aptos',
    'Arial',
    'Baskerville Old Face',
    'Book Antiqua',
    'Bookerly',
    'Caecilia',
    'Calibri',
    'Cambria',
    'Candara',
    'Cinzel',
    'Constantia',
    'Cormorant Garamond',
    'Crimson Text',
    'EB Garamond',
    'Garamond',
    'Georgia',
    'Goudy Old Style',
    'Helvetica',
    'Ibarra Real Nova',
    'Libre Baskerville',
    'Literata',
    'Lora',
    'Merriweather',
    'Minion Pro',
    'Noto Serif',
    'Palatino Linotype',
    'Playfair Display',
    'Rockwell',
    'Segoe UI',
    'Source Serif Pro',
    'Spectral',
    'Times New Roman',
    'Trebuchet MS',
    'Verdana',
]

CONFIG_VIDE = {
    "informations": {
        "titre_complet": "", "sous_titre": "", "auteur": "", "preface_postface": "",
        "isbn": "", "depot_legal": "", "annee_publication": "", "maison_edition": "",
        "mention_copyright": "", "edition": "", "site_web": "", "avertissement": "",
        "dedicace": "", "epigraphe": ""
    },
    "ia": {
        "mode": "auto", "url_ollama": "", "modele_ollama": "", "cle_api_ollama": "",
        "cle_api_openai": "", "modele_openai": "", "url_api_openai": ""
    },
    "style": {
        "format_livre": "7 x 10 po",
        "police_corps": "Aptos",
        "taille_corps_pt": 11,
        "police_titres": "Cinzel",
        "taille_titres_acte_pt": 14,
        "taille_chapitre_ligne1_pt": 12,
        "taille_chapitre_ligne2_pt": 11,
        "taille_sous_chapitre_pt": 11,
        "interligne_corps": 1,
        "police_lettrine": "Cinzel"          # v1.3
    },
    "chapitres": [],
    "bible": {"personnages": [], "lieux": [], "objets": [], "chronologie": []}
}


# ─────────────────────────────────────────────
# Aides formats / marges KDP
# ─────────────────────────────────────────────
def dimensions_format_po(valeur):
    """Retourne (largeur, hauteur) en pouces à partir d'un libellé de format."""
    defaut = (7.0, 10.0)
    label = str(valeur or '').lower()
    if not label:
        return defaut
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)', label)
    if not m:
        return defaut
    w = float(m.group(1).replace(',', '.'))
    h = float(m.group(2).replace(',', '.'))
    if 'cm' in label and 'po' not in label:
        w, h = w / 2.54, h / 2.54
    return w, h


def infos_format_kdp(valeur):
    """Retourne type d'ouvrage + usage recommandé pour un libellé de format."""
    w, h = dimensions_format_po(valeur)
    for f in FORMATS_KDP:
        fw, fh = dimensions_format_po(f["label"])
        if abs(fw - w) < 0.05 and abs(fh - h) < 0.05:
            return {"type": f["type"], "usage": f["usage"], "cm": f["cm"]}
    return {"type": "Format personnalisé",
            "usage": "Hors catalogue KDP standard : vérifiez la conformité avant publication.",
            "cm": ""}


def marges_kdp_po(nb_pages):
    """Marges minimales Amazon KDP (broché, sans fond perdu)."""
    try:
        n = int(nb_pages)
    except Exception:
        n = 150
    if n <= 150:
        gouttiere = 0.375
    elif n <= 300:
        gouttiere = 0.5
    elif n <= 500:
        gouttiere = 0.625
    elif n <= 700:
        gouttiere = 0.75
    else:
        gouttiere = 0.875
    return {"interieure_po": gouttiere, "exterieure_po": 0.25}


# ─────────────────────────────────────────────
# Migration Excel → JSON
# ─────────────────────────────────────────────
def _lire_excel():
    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG_EXCEL, data_only=True)
        data = json.loads(json.dumps(CONFIG_VIDE))

        if 'Informations' in wb.sheetnames:
            map_infos = {"Titre complet du roman": "titre_complet", "Sous-titre éventuel": "sous_titre",
                         "Nom de l'auteur (couverture)": "auteur", "Préface / Postface": "preface_postface",
                         "Numéro ISBN": "isbn", "Dépôt légal": "depot_legal", "Année de publication": "annee_publication",
                         "Maison d'édition / auto-édition": "maison_edition", "Mention de copyright": "mention_copyright",
                         "Édition": "edition", "Site web": "site_web", "Avertissement": "avertissement",
                         "Dédicace": "dedicace", "Épigraphe": "epigraphe"}
            for row in wb['Informations'].iter_rows(min_row=1, values_only=True):
                if row and row[0] and str(row[0]).strip() in map_infos:
                    data["informations"][map_infos[str(row[0]).strip()]] = str(row[1]).strip() if row[1] else ""

        if 'IA' in wb.sheetnames:
            map_ia = {"Mode IA": "mode", "URL Ollama": "url_ollama", "Modèle Ollama": "modele_ollama",
                      "Clé API Ollama": "cle_api_ollama", "Clé API OpenAI": "cle_api_openai",
                      "Modèle OpenAI": "modele_openai", "URL API OpenAI": "url_api_openai"}
            for row in wb['IA'].iter_rows(min_row=1, values_only=True):
                if row and row[0] and str(row[0]).strip() in map_ia:
                    data["ia"][map_ia[str(row[0]).strip()]] = str(row[1]).strip() if row[1] else ""

        if 'Style' in wb.sheetnames:
            map_style = {"Format du livre": "format_livre", "Police du corps de texte": "police_corps",
                         "Taille du corps (pt)": "taille_corps_pt", "Police des titres": "police_titres",
                         "Taille des titres d'acte (pt)": "taille_titres_acte_pt",
                         "Taille chapitre - ligne 1 (pt)": "taille_chapitre_ligne1_pt",
                         "Taille chapitre - ligne 2 (pt)": "taille_chapitre_ligne2_pt",
                         "Taille sous-chapitre (pt)": "taille_sous_chapitre_pt",
                         "Interligne du corps": "interligne_corps"}
            for row in wb['Style'].iter_rows(min_row=1, values_only=True):
                if row and row[0] and str(row[0]).strip() in map_style:
                    cle = map_style[str(row[0]).strip()]
                    val = row[1]
                    if val is not None:
                        if "taille" in cle or "interligne" in cle:
                            try:
                                val = float(str(val).replace(',', '.'))
                            except Exception:
                                pass
                        else:
                            val = str(val).strip()
                        data["style"][cle] = val

        if 'Chapitres' in wb.sheetnames:
            en_tete = False
            for row in wb['Chapitres'].iter_rows(min_row=1, values_only=True):
                if not en_tete:
                    if row and row[0] and "fichier" in str(row[0]).lower():
                        en_tete = True
                        continue
                    else:
                        en_tete = True
                if not row or not any(row):
                    continue
                vals = [str(c).strip() if c is not None else "" for c in row]
                while len(vals) < 6:
                    vals.append("")
                fichier, acte, l1, l2, image, legende = vals[:6]

                if image and not fichier:
                    chap = {"type": "image", "image": image, "legende": legende}
                elif acte and not fichier and not image:
                    chap = {"type": "acte", "acte": acte}
                elif fichier:
                    chap = {"type": "chapitre", "fichier_source": fichier,
                            "chapitre_ligne1": l1, "chapitre_ligne2": l2}
                else:
                    continue
                data["chapitres"].append(chap)

        wb.close()
        return data
    except Exception as e:
        print(f"⚠️ Erreur migration Excel : {e}")
        return None


# ─────────────────────────────────────────────
# Lecture / écriture JSON
# ─────────────────────────────────────────────
def charger_configuration():
    if not os.path.isfile(CHEMIN_CONFIG_JSON):
        if os.path.isfile(CHEMIN_CONFIG_EXCEL):
            data = _lire_excel()
            if data:
                sauvegarder_configuration(data)
                print("✅ Configuration JSON créée à partir de Configuration_roman.xlsx.")
                return data
        sauvegarder_configuration(CONFIG_VIDE)
        return json.loads(json.dumps(CONFIG_VIDE))

    try:
        with open(CHEMIN_CONFIG_JSON, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for k, v in CONFIG_VIDE.items():
            if k not in data:
                data[k] = v
            elif isinstance(v, dict):
                for sk, sv in v.items():
                    if sk not in data[k]:
                        data[k][sk] = sv
        return data
    except Exception:
        return json.loads(json.dumps(CONFIG_VIDE))


def sauvegarder_configuration(data):
    try:
        if os.path.isfile(CHEMIN_CONFIG_JSON):
            shutil.copy2(CHEMIN_CONFIG_JSON, CHEMIN_CONFIG_JSON + '.backup')
        with open(CHEMIN_CONFIG_JSON, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        print(f"⚠️ Erreur sauvegarde JSON : {e}")
        return False


def lire_format_actuel():
    return str(charger_configuration().get('style', {}).get('format_livre', ''))


def lire_chapitres_pour_traduction():
    data = charger_configuration()
    titres, mapping = [], {}
    for chap in data.get('chapitres', []):
        if chap.get('type') == 'chapitre':
            titre, fichier = chap.get('chapitre_ligne1', ''), chap.get('fichier_source', '')
            if titre and fichier and titre not in mapping:
                titres.append(titre)
                mapping[titre] = fichier
    return titres, mapping