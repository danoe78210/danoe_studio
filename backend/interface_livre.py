#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
interface_livre.py – Danoë Studio (v1.26.5)
Interface professionnelle customtkinter : thème sombre, barre latérale,
cartes statistiques, console temps réel, infobulles, rendu DPI-aware.

- v1.26.5 : AMÉLIORATION MAJEURE de l'effet K2000 (50 FPS, glow réaliste, 
            traînée rouge/orange/jaune, gestion robuste des dimensions).
- v1.26.4 : Correction format preferredVariants (fr-FR, fr-CA, etc.).
- v1.26.3 : Correction contrainte API LanguageTool (language=auto + preferredVariants).
- v1.26.2 : Suppression de la fonctionnalité "Cohérence IA".
- v1.26.1 : Correction critique CTkComboBox et sécurisation API.
- v1.26 : Menu de configuration du correcteur.
- v1.25 : Correcteur Orthographique et Grammatical intégré.
- v1.24 : Détection proactive 'pywin32', Regex parser_stats assouplie.
- v1.23 : Configuration JSON-first, repli Excel, UI tolérante.

Dépendances UI : pip install customtkinter requests
Modules requis : configuration_store.py, configuration_ui.py, spellchecker.py
"""

import os
import re
import sys
import glob
import math
import queue
import runpy
import shutil
import tempfile
import threading
import time
import json
import tkinter as tk
from tkinter import font as tkfont, messagebox
import requests
import spellchecker

# ── Rendu net sous Windows (DPI awareness) ──
try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(1)
except Exception:
    pass

try:
    import customtkinter as ctk
except ImportError:
    import tkinter as _tk
    _tk.Tk().withdraw()
    messagebox.showerror('Dépendance manquante', 'Cette interface requiert customtkinter.\nLancez : pip install customtkinter')
    raise SystemExit(1)

# ── Modules de configuration JSON (repli Excel si absents) ──
try:
    import configuration_store as cst
    CONFIG_JSON_OK = True
except Exception:
    cst = None
    CONFIG_JSON_OK = False

try:
    import configuration_ui as cui
    CONFIG_UI_OK = True
except Exception:
    cui = None
    CONFIG_UI_OK = False

# ─────────────────────────────────────────
# Chemins
# ─────────────────────────────────────────
if getattr(sys, 'frozen', False):
    BASE = os.path.dirname(sys.executable)
    _MEI = getattr(sys, '_MEIPASS', BASE)
    SCRIPT = os.path.join(_MEI, 'generer_roman.py')
    IA_SCRIPT = os.path.join(_MEI, 'IA_Roman.py')
    EBOOK_SCRIPT = os.path.join(_MEI, 'generer_ebook.py')
else:
    BASE = os.path.dirname(os.path.abspath(__file__))
    SCRIPT = os.path.join(BASE, 'generer_roman.py')
    IA_SCRIPT = os.path.join(BASE, 'IA_Roman.py')
    EBOOK_SCRIPT = os.path.join(BASE, 'generer_ebook.py')

CHEMIN_CONFIG = os.path.join(BASE, 'Configuration_roman.xlsx')
CHEMIN_LOGO = os.path.join(BASE, 'logo_danoe.png')
CONFIG_CORRECTEUR = os.path.join(BASE, 'correcteur_config.json')
DOSSIER_IMAGES = os.path.join(BASE, 'Images')
DOSSIER_CHAPITRES = os.path.join(BASE, 'Chapitres')
DOSSIER_TRADUCTIONS = os.path.join(BASE, 'Traductions')
VERSION_INTERFACE = '1.26.5'

# ── Palette « studio » ──
BG = '#101322'
PANEL = '#171b2e'
CARD = '#1b2036'
BORD = '#262c49'
TXT = '#e8eaf6'
MUTED = '#6b7194'
ACCENT = '#3d6bff'
GOLD = '#e0b458'
OK = '#3ddc97'
WARN = '#ff6b6b'
CONSOLE = '#121524'

# ─────────────────────────────────────────
# Config Excel (onglets IA + Bible) — repli historique
# ─────────────────────────────────────────
def assurer_onglets_config():
    if not os.path.isfile(CHEMIN_CONFIG):
        return
    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG)
        modif = False
        if 'IA' not in wb.sheetnames:
            ws = wb.create_sheet('IA')
            ws.append(['Paramètre', 'Valeur'])
            ws.append(['Mode IA', 'auto'])
            ws.append(['URL Ollama', 'https://ollama.com'])
            ws.append(['Modèle Ollama', 'llama3.2:3b'])
            ws.append(['Clé API Ollama', ''])
            ws.append(['Clé API OpenAI', ''])
            ws.append(['Modèle OpenAI', ''])
            ws.append(['URL API OpenAI', ''])
            ws.column_dimensions['A'].width = 28
            ws.column_dimensions['B'].width = 60
            modif = True

        if 'Bible' not in wb.sheetnames:
            ws = wb.create_sheet('Bible')
            ws.append(['BIBLE DE SÉRIE', '', '', '', ''])
            ws.append([])
            ws.append(['PERSONNAGES', '', '', '', ''])
            ws.append(['Nom', 'Traits', 'Relations', 'Statut', 'Chapitres'])
            ws.append([])
            ws.append(['LIEUX', ''])
            ws.append(['Lieu', 'Chapitres'])
            ws.append([])
            ws.append(['OBJETS', ''])
            ws.append(['Objet', 'Chapitres'])
            ws.append([])
            ws.append(['CHRONOLOGIE', '', ''])
            ws.append(['Chapitre', 'Événement', 'Repère temporel'])
            for col, w in zip('ABCDE', (22, 30, 30, 14, 24)):
                ws.column_dimensions[col].width = w
            modif = True

        if modif:
            wb.save(CHEMIN_CONFIG)
        wb.close()
    except Exception as e:
        print(f'⚠️ Onglets Excel non créés : {e}')

def lire_format_actuel():
    if CONFIG_JSON_OK:
        try:
            fmt = cst.lire_format_actuel()
            if fmt: return fmt
        except Exception: pass
    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG, data_only=True, read_only=True)
        if 'Style' in wb.sheetnames:
            for row in wb['Style'].iter_rows(values_only=True):
                if row and row[0] and str(row[0]).strip().lower().startswith('format'):
                    return str(row[1]).strip() if row[1] else ''
    except Exception: pass
    return ''

class Redirecteur:
    def __init__(self, file):
        self._file = file
    def write(self, s):
        if s and s.strip():
            self._file.put(s)
            return len(s) if s else 0
    def flush(self): pass

# ─────────────────────────────────────────
# EXPORT PDF KDP
# ─────────────────────────────────────────
WD_EXPORT_PDF = 17
WD_OPTIMIZE_PRINT = 1

def detecter_dernier_docx_kdp():
    cands = glob.glob(os.path.join(BASE, '*_KDP.docx'))
    return max(cands, key=os.path.getmtime) if cands else None

def _dimensions_format_kdp():
    defaut = (7.0, 10.0)
    label = (lire_format_actuel() or '').lower()
    if not label: return defaut
    m = re.search(r'(\d+(?:[.,]\d+)?)\s*x\s*(\d+(?:[.,]\d+)?)', label)
    if not m: return defaut
    w = float(m.group(1).replace(',', '.'))
    h = float(m.group(2).replace(',', '.'))
    if 'cm' in label and 'po' not in label: w, h = w / 2.54, h / 2.54
    return w, h

def verifier_pdf_kdp(chemin_pdf, journal=print):
    w_po, h_po = _dimensions_format_kdp()
    ok = True
    try:
        with open(chemin_pdf, 'rb') as f: data = f.read(300000)
        m = re.search(rb'/MediaBox\s*\[\s*([\d.]+)\s+([\d.]+)\s+([\d.]+)\s+([\d.]+)\s*\]', data)
        if m:
            x0, y0, x1, y1 = map(float, m.groups())
            w_pt, h_pt = x1 - x0, y1 - y0
            if abs(w_pt - w_po * 72) <= 2 and abs(h_pt - h_po * 72) <= 2:
                journal(f' ✅ Format de page conforme : {w_po:g} x {h_po:g} po')
            else:
                ok = False
                journal(f' ⚠️ Format inattendu : {w_pt / 72:.2f} x {h_pt / 72:.2f} po (attendu {w_po:g} x {h_po:g} po)')
        else:
            journal(' ℹ️ MediaBox non lue : format hérité fidèlement du Word.')
    except Exception as e:
        journal(f' ⚠️ Vérification impossible : {e}')
    return ok

def _exporter_pdf_word(tmp_docx, out_pdf, journal):
    import win32com.client as win32
    word = win32.Dispatch('Word.Application')
    word.Visible = False
    word.DisplayAlerts = 0
    try:
        doc = word.Documents.Open(os.path.abspath(tmp_docx))
        try: doc.DoNotCompressImages = True
        except Exception: pass
        journal(' 🖤 Texte forcé en noir pur…')
        doc.Content.Font.Color = 0
        for sec in doc.Sections:
            for coll in (sec.Headers, sec.Footers):
                for hf in coll:
                    try: hf.Range.Font.Color = 0
                    except Exception: pass
        pages = doc.ComputeStatistics(2)
        journal(f' 📑 Export HD de {pages} pages (qualité impression)…')
        doc.ExportAsFixedFormat(
            OutputFileName=os.path.abspath(out_pdf), ExportFormat=WD_EXPORT_PDF,
            OpenAfterExport=False, OptimizeFor=WD_OPTIMIZE_PRINT, IncludeDocProps=True,
            DocStructureTags=True, BitmapMissingFonts=False, UseISO19005_1=False,
        )
        doc.Close(False)
        return pages
    finally:
        try: word.Quit()
        except Exception: pass

def generer_pdf_kdp(chemin_docx=None, chemin_pdf=None, journal=print):
    try:
        import pythoncom
        pythoncom.CoInitialize()
    except Exception: pass
    try:
        import win32com.client  # noqa
    except ImportError:
        raise RuntimeError('pywin32 requis : lancez « pip install pywin32 ».')
    
    if chemin_docx is None: chemin_docx = detecter_dernier_docx_kdp()
    if chemin_docx is None: raise FileNotFoundError('Aucun document _KDP.docx trouvé.')
    if chemin_pdf is None:
        base, _ = os.path.splitext(chemin_docx)
        chemin_pdf = base + '.pdf'

    journal(f'🖨 Préparation de l\'export PDF depuis : {os.path.basename(chemin_docx)}')
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_docx = os.path.join(tmp_dir, 'master.docx')
        shutil.copy2(chemin_docx, tmp_docx)
        pages = _exporter_pdf_word(tmp_docx, chemin_pdf, journal)
    verifier_pdf_kdp(chemin_pdf, journal)
    return chemin_pdf, pages

# ─────────────────────────────────────────
# UI : Scanner K2000 Amélioré (v1.26.5)
# ─────────────────────────────────────────
class ScannerK2000(tk.Canvas):
    """
    Effet K2000 amélioré : scanner lumineux avec traînée, effet de lueur (glow),
    et animation fluide à 50 FPS. S'adapte automatiquement à la largeur du canvas.
    """
    def __init__(self, parent, height=30, bg=CONSOLE, width=400):
        super().__init__(parent, height=height, width=width, bg=bg, 
                         highlightthickness=0, borderwidth=0)
        self._actif = False
        self._tete = 0.0  # Position flottante pour interpolation fluide
        self._sens = 1
        self._id = None
        self._largeur = width
        self._hauteur = height
        
        # Forcer le calcul des dimensions réelles après le rendu
        self.bind('<Configure>', self._on_resize)
        
    def _on_resize(self, event):
        """Met à jour les dimensions quand le canvas est redimensionné."""
        self._largeur = event.width
        self._hauteur = event.height
        
    def demarrer(self):
        """Démarre l'animation depuis le bord gauche."""
        self._actif = True
        self._tete = 0.0
        self._sens = 1
        # Forcer un refresh des dimensions avant de démarrer
        self.update_idletasks()
        self._largeur = self.winfo_width()
        if self._largeur < 50:
            self._largeur = 400  # Fallback si le canvas n'est pas encore rendu
        self._animer()

    def arreter(self):
        """Arrête l'animation et efface le canvas."""
        self._actif = False
        if self._id:
            self.after_cancel(self._id)
            self._id = None
        self.delete('all')

    def _couleur(self, intensite):
        """
        Calcule une couleur rouge/orange/jaune en fonction de l'intensité.
        Intensité 1.0 = rouge vif, 0.0 = orange foncé (presque éteint).
        """
        # Dégradé : rouge vif -> orange -> jaune pâle
        r = int(255)
        g = int(80 + 175 * intensite)  # 80 (rouge) à 255 (jaune)
        b = int(50 * intensite)         # Bleu très faible pour l'effet chaud
        return f'#{r:02x}{g:02x}{b:02x}'

    def _animer(self):
        """Boucle d'animation principale à ~50 FPS."""
        if not self._actif:
            return
            
        self.delete('all')
        
        # S'assurer que les dimensions sont valides
        if self._largeur < 50:
            self._largeur = self.winfo_width()
            if self._largeur < 50:
                self._largeur = 400
        
        w = self._largeur
        h = self._hauteur
        cy = h // 2
        
        # Paramètres de l'animation
        vitesse = 8.0  # Pixels par frame (plus fluide que 6)
        trainee_longueur = 120  # Longueur de la traînée en pixels
        pas = 3  # Espacement entre les segments de la traînée
        
        # ── DESSIN DE LA TRAÎNÉE ──
        n_segments = int(trainee_longueur // pas)
        
        for i in range(n_segments):
            distance = i * pas
            x = self._tete - (distance * self._sens)
            
            if x < -pas or x > w + pas:
                continue
            
            intensite = 1.0 - (distance / trainee_longueur)
            if intensite <= 0.05:
                continue
            
            hauteur_segment = 4 + 18 * (intensite ** 1.5)
            couleur = self._couleur(intensite)
            
            # Cercle de lueur (plus grand, plus transparent)
            if intensite > 0.3:
                glow_size = hauteur_segment * 1.8
                self.create_oval(
                    x - glow_size/2, cy - glow_size/2,
                    x + glow_size/2, cy + glow_size/2,
                    fill=self._couleur(intensite * 0.4), outline=''
                )
            
            # Segment principal
            self.create_rectangle(
                x - pas/2, cy - hauteur_segment/2,
                x + pas/2, cy + hauteur_segment/2,
                fill=couleur, outline=''
            )
        
        # ── DESSIN DE LA TÊTE LUMINEUSE ──
        # Cercle externe (lueur diffuse)
        self.create_oval(
            self._tete - 20, cy - 20, self._tete + 20, cy + 20,
            fill=self._couleur(0.15), outline=''
        )
        # Cercle moyen (lueur moyenne)
        self.create_oval(
            self._tete - 14, cy - 14, self._tete + 14, cy + 14,
            fill=self._couleur(0.4), outline=''
        )
        # Cercle interne (lueur forte)
        self.create_oval(
            self._tete - 9, cy - 9, self._tete + 9, cy + 9,
            fill=self._couleur(0.7), outline=''
        )
        # Cœur brillant (blanc/jaune pâle)
        self.create_oval(
            self._tete - 5, cy - 5, self._tete + 5, cy + 5,
            fill='#ffddaa', outline=''
        )
        # Point central ultra-brillant
        self.create_oval(
            self._tete - 2, cy - 2, self._tete + 2, cy + 2,
            fill='#ffffff', outline=''
        )
        
        # ── MISE À JOUR DE LA POSITION ──
        self._tete += vitesse * self._sens
        
        # Rebond aux extrémités avec marge
        marge = 25
        if self._tete >= w - marge:
            self._tete = w - marge
            self._sens = -1
        elif self._tete <= marge:
            self._tete = marge
            self._sens = 1
        
        # Prochaine frame dans 20ms (~50 FPS)
        self._id = self.after(20, self._animer)

class InfoBulle:
    def __init__(self, widget, texte):
        self.widget = widget
        self.texte = texte
        self.tipwindow = None
        self.id = None
        self.widget.bind('<Enter>', self.enter)
        self.widget.bind('<Leave>', self.leave)

    def enter(self, event=None): self.schedule()
    def leave(self, event=None): self.unschedule(); self.hidetip()
    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(400, self.showtip)
    def unschedule(self):
        if self.id:
            self.widget.after_cancel(self.id)
            self.id = None
    def showtip(self):
        if self.tipwindow or not self.texte: return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f'+{x}+{y}')
        tk.Label(tw, text=self.texte, justify='left', background='#232741', foreground='#e8eaf6', relief='solid', borderwidth=1, font=('Segoe UI', 9)).pack(ipadx=6, ipady=4)
    def hidetip(self):
        tw = self.tipwindow
        self.tipwindow = None
        if tw: tw.destroy()

# ─────────────────────────────────────────
# Application Principale
# ─────────────────────────────────────────
class StudioApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        ctk.set_appearance_mode('dark')
        ctk.set_default_color_theme('dark-blue')
        self.title(f'Danoë Studio — Générateur universel de roman · v{VERSION_INTERFACE}')
        self.geometry('1180x820')
        self.minsize(980, 700)
        self.configure(fg_color=BG)
        
        self.file_sortie = queue.Queue()
        self.sortie = []
        self.boutons = []
        self.combos = []
        self.processus_en_cours = False
        self.logo_img = None
        self.fenetre_config = None
        self._map_titre_fichier = {}
        
        # Détection proactive de pywin32
        self.pywin32_ok = False
        try:
            import win32com.client  # noqa: F401
            self.pywin32_ok = True
        except ImportError: pass

        # Chargement de la config du correcteur
        self.config_correcteur = self._charger_config_correcteur()

        try: assurer_onglets_config()
        except Exception: pass

        self._police_valeur = tkfont.Font(family='Segoe UI', size=15, weight='bold')
        self._construire_ui()
        self.after(120, self._vider_console)
        self._log(f' ✅ Danoë Studio v{VERSION_INTERFACE} prêt.')
        if CONFIG_JSON_OK:
            self._log(' 🧩 Configuration JSON active : Configuration_roman.json')
        else:
            self._log(' ⚠️ configuration_store.py introuvable → repli Excel.')

    def _charger_config_correcteur(self):
        config_defaut = {
            'level': 'default',
            'preferredVariants': 'fr-FR',
            'motherTongue': 'Aucune'
        }
        if os.path.isfile(CONFIG_CORRECTEUR):
            try:
                with open(CONFIG_CORRECTEUR, 'r', encoding='utf-8') as f:
                    loaded = json.load(f)
                    return {**config_defaut, **loaded}
            except Exception: pass
        return config_defaut

    def _sauvegarder_config_correcteur(self, config):
        try:
            with open(CONFIG_CORRECTEUR, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=4)
            self.config_correcteur = config
            self._log(' ⚙️ Paramètres du correcteur sauvegardés.')
        except Exception as e:
            messagebox.showerror('Erreur', f'Impossible de sauvegarder la configuration :\n{e}')

    def _construire_ui(self):
        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self._en_tete()
        self._barre_laterale()
        self._zone_principale()
        self._barre_etat()

    def _en_tete(self):
        ent = ctk.CTkFrame(self, fg_color=PANEL, corner_radius=0, height=64)
        ent.grid(row=0, column=0, columnspan=2, sticky='ew')
        ent.grid_propagate(False)
        if os.path.isfile(CHEMIN_LOGO):
            try:
                img = tk.PhotoImage(file=CHEMIN_LOGO)
                f = max(1, img.height() // 40)
                if f > 1: img = img.subsample(f, f)
                self.logo_img = ctk.CTkImage(light_image=img, dark_image=img, size=(40, 40))
                ctk.CTkLabel(ent, image=self.logo_img, text='').pack(side='left', padx=(16, 10), pady=12)
            except Exception: pass
        ctk.CTkLabel(ent, text='Danoë Studio', font=('Segoe UI', 20, 'bold'), text_color=GOLD).pack(side='left', padx=(0, 8))
        ctk.CTkLabel(ent, text='Générateur universel de roman · KDP', font=('Segoe UI', 11), text_color=MUTED).pack(side='left', pady=12)

        fmt = lire_format_actuel() or '—'
        chip = ctk.CTkFrame(ent, fg_color=CARD, corner_radius=12, border_width=1, border_color=BORD)
        chip.pack(side='right', padx=16, pady=16)
        self.chip_format_label = ctk.CTkLabel(chip, text=f'📐 {fmt}', font=('Segoe UI', 11, 'bold'), text_color=TXT)
        self.chip_format_label.pack(padx=12, pady=5)

        ver = ctk.CTkFrame(ent, fg_color=CARD, corner_radius=12, border_width=1, border_color=BORD)
        ver.pack(side='right', padx=(0, 8), pady=16)
        ctk.CTkLabel(ver, text=f'v{VERSION_INTERFACE}', font=('Segoe UI', 10, 'bold'), text_color=MUTED).pack(padx=10, pady=5)

    def _titre_section(self, parent, texte):
        ctk.CTkLabel(parent, text=texte, font=('Segoe UI', 10, 'bold'), text_color=MUTED, anchor='w').pack(fill='x', padx=16, pady=(14, 2))

    def _btn(self, parent, texte, methode, tooltip, accent=False):
        b = ctk.CTkButton(parent, text=texte, command=getattr(self, methode), anchor='w', height=38, corner_radius=8,
                          font=('Segoe UI', 12, 'bold'), fg_color=(ACCENT, '#2f56d6') if accent else ('#232741', '#232741'),
                          hover_color=('#2f56d6', '#343a63') if accent else ('#2b3050', '#343a63'), text_color=TXT)
        b.pack(fill='x', padx=10, pady=3)
        InfoBulle(b, tooltip)
        self.boutons.append((b, texte))
        return b

    def _barre_laterale(self):
        lat = ctk.CTkFrame(self, fg_color=PANEL, corner_radius=0, width=250)
        lat.grid(row=1, column=0, sticky='ns')
        lat.grid_propagate(False)

        self._titre_section(lat, 'PRODUCTION')
        self._btn(lat, '▶ Générer le livre', 'lancer', 'Construit le .docx complet ; statistiques dans la console.', accent=True)
        
        tooltip_pdf = 'Export Word haute qualité du master .docx N&B 300 dpi (conforme KDP).' if self.pywin32_ok else '⚠️ Fonction désactivée : installez la dépendance via "pip install pywin32".'
        btn_pdf = self._btn(lat, '🖨 PDF KDP noir & blanc', 'generer_pdf_kdp_gui', tooltip_pdf)
        if not self.pywin32_ok: btn_pdf.configure(state='disabled', fg_color='#3a3f55', text_color='#8b92b3')

        self._btn(lat, '📱 Ebook KDP (EPUB)', 'generer_ebook_gui', 'Génère un ebook EPUB 3 reflowable compatible Amazon KDP.')
        self._btn(lat, '📄 Ouvrir le .docx', 'ouvrir_docx', 'Ouvre le dernier document généré dans Word.')

        self._titre_section(lat, 'CONFIGURATION')
        self._btn(lat, '⚙ Configuration du livre', 'ouvrir_configuration', 'Ouvre l\'éditeur de configuration JSON.')
        self._btn(lat, '🖼 Dossier des images', 'ouvrir_dossier_images', 'Ouvre Images\\ pour insérer une illustration.')
        self._btn(lat, '📁 Dossier des chapitres', 'ouvrir_dossier_chapitres', 'Ouvre Chapitres\\ pour insérer un chapitre (.md).')

        self._titre_section(lat, 'INTELLIGENCE')
        self._btn(lat, '🤖 Résumés IA', 'generer_resumes', 'Génère Résumé.md à partir des chapitres.')

        self._titre_section(lat, 'TRADUCTION')
        ctk.CTkLabel(lat, text='Cible de traduction', font=('Segoe UI', 9, 'bold'), text_color=MUTED, anchor='w').pack(fill='x', padx=16, pady=(8, 2))
        cadre_cible = ctk.CTkFrame(lat, fg_color='transparent')
        cadre_cible.pack(fill='x', padx=10, pady=(0, 3))
        self.cible_traduction_var = ctk.StringVar(value='Tout le dossier')
        self.combo_traduction = ctk.CTkComboBox(cadre_cible, values=self.lister_cibles_traduction(), variable=self.cible_traduction_var, state='readonly', width=175)
        self.combo_traduction.pack(side='left', fill='x', expand=True)
        InfoBulle(self.combo_traduction, 'Choisis "Tout le dossier" ou un chapitre spécifique.')
        self.combos.append(self.combo_traduction)
        
        refresh = ctk.CTkButton(cadre_cible, text='↻', width=34, height=28, command=self.rafraichir_cibles_traduction)
        refresh.pack(side='right', padx=(6, 0))
        InfoBulle(refresh, 'Actualise la liste des chapitres.')
        self.boutons.append((refresh, '↻'))
        self._btn(lat, '🌐 Traduction EN', 'traduire_en', 'Traduit le dossier ou le chapitre sélectionné en anglais.')

        # ── Section CORRECTION ──
        self._titre_section(lat, 'CORRECTION')
        self._btn(lat, '⚙️ Paramètres du correcteur', 'ouvrir_parametres_correcteur',
                  'Configurer le niveau de vérification, la variante orthographique et la langue maternelle.')
        self._btn(lat, '📝 Correcteur Orthographique', 'lancer_correcteur',
                  'Vérifie l\'orthographe et la grammaire du chapitre sélectionné ci-dessus.')

    def _carte(self, parent, titre):
        f = ctk.CTkFrame(parent, corner_radius=10, fg_color=CARD, border_width=1, border_color=BORD)
        f.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(f, text=titre, font=('Segoe UI', 9, 'bold'), text_color=MUTED, anchor='w').grid(row=0, column=0, sticky='w', padx=10, pady=(8, 0))
        v = ctk.CTkLabel(f, text='—', font=('Segoe UI', 15, 'bold'), text_color=TXT, anchor='w')
        v.grid(row=1, column=0, sticky='w', padx=10, pady=(0, 8))
        return v

    def _zone_principale(self):
        main = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        main.grid(row=1, column=1, sticky='nsew', padx=14, pady=12)
        main.grid_columnconfigure(0, weight=1)
        main.grid_rowconfigure(1, weight=1)

        cartes = ctk.CTkFrame(main, fg_color=BG)
        cartes.grid(row=0, column=0, sticky='ew', pady=(0, 10))
        for i in range(6): cartes.grid_columnconfigure(i, weight=1, uniform='c')
        
        self.c_ouvrage = self._carte(cartes, 'OUVRAGE'); self.c_ouvrage.grid(row=0, column=0, sticky='ew', padx=(0, 5))
        self.c_format = self._carte(cartes, 'FORMAT'); self.c_format.grid(row=0, column=1, sticky='ew', padx=5)
        self.c_mots = self._carte(cartes, 'MOTS'); self.c_mots.grid(row=0, column=2, sticky='ew', padx=5)
        self.c_pages = self._carte(cartes, 'PAGES'); self.c_pages.grid(row=0, column=3, sticky='ew', padx=5)
        self.c_chap = self._carte(cartes, 'CHAPITRES'); self.c_chap.grid(row=0, column=4, sticky='ew', padx=5)
        self.c_illus = self._carte(cartes, 'ILLUSTRATIONS'); self.c_illus.grid(row=0, column=5, sticky='ew', padx=(5, 0))

        cadre = ctk.CTkFrame(main, fg_color=PANEL, corner_radius=10, border_width=1, border_color=BORD)
        cadre.grid(row=1, column=0, sticky='nsew')
        cadre.grid_rowconfigure(2, weight=1)
        cadre.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(cadre, text='CONSOLE DE PRODUCTION', font=('Segoe UI', 10, 'bold'), text_color=MUTED, anchor='w').grid(row=0, column=0, sticky='w', padx=12, pady=(10, 4))
        
        # v1.26.5 : Scanner K2000 avec largeur explicite
        self.k2000 = ScannerK2000(cadre, height=30, bg=CONSOLE, width=800)
        self.k2000.grid(row=1, column=0, sticky='ew', padx=8, pady=(0, 6))
        self.k2000.grid_remove()

        self.console = ctk.CTkTextbox(cadre, font=('Consolas', 14), fg_color=CONSOLE, text_color='#c9cde0', corner_radius=8, border_width=0, state='disabled', wrap='none')
        self.console.grid(row=2, column=0, sticky='nsew', padx=8, pady=(0, 8))
        self.console.tag_config('ok', foreground=OK)
        self.console.tag_config('warn', foreground=WARN)
        self.console.tag_config('sep', foreground='#4a5070')
        self.console.tag_config('head', foreground=GOLD)
        self.console.tag_config('line', foreground='#c9cde0')

    def _barre_etat(self):
        bar = ctk.CTkFrame(self, fg_color=PANEL, corner_radius=0, height=34)
        bar.grid(row=2, column=0, columnspan=2, sticky='ew')
        bar.grid_propagate(False)
        self.statut = ctk.CTkLabel(bar, text='Prêt.', font=('Segoe UI', 10), text_color=MUTED, anchor='w')
        self.statut.pack(side='left', padx=16)
        self.chrono = ctk.CTkLabel(bar, text='', font=('Segoe UI', 10), text_color=MUTED, anchor='e')
        self.chrono.pack(side='right', padx=16)

    def _vider_console(self):
        try:
            while True: self._log(self.file_sortie.get_nowait())
        except queue.Empty: pass
        self.after(120, self._vider_console)

    def _log(self, ligne):
        if ligne.strip().startswith('💬'): return
        self.sortie.append(ligne)
        if ligne.startswith('⚠️'): tag = 'warn'
        elif ligne.startswith('✅'): tag = 'ok'
        elif ligne.startswith(('=', '-')): tag = 'sep'
        elif ligne.startswith(('📊', '📖', '🎭', '📚', '🖼', '🔤', '🧠', '🎨', '📄', '🌐', '⏱', '📝', '🪙', '🧩', '📱', '🖨', '⚙️')): tag = 'head'
        else: tag = 'line'
        self.console.configure(state='normal')
        self.console.insert('end', ligne + '\n', tag)
        self.console.see('end')
        self.console.configure(state='disabled')

    def _elaguer(self, texte, largeur_max=150):
        if self._police_valeur.measure(texte) <= largeur_max: return texte
        while texte and self._police_valeur.measure(texte + '…') > largeur_max: texte = texte[:-1]
        return texte + '…'

    def _maj_cartes(self, s):
        self.c_ouvrage.configure(text=self._elaguer(s.get('ouvrage', '—')))
        self.c_format.configure(text=self._elaguer(lire_format_actuel() or '—'))
        self.c_mots.configure(text=self._elaguer(s.get('mots', '—')))
        self.c_pages.configure(text=self._elaguer(s.get('pages', '—')))
        self.c_chap.configure(text=self._elaguer(s.get('chapitres', '—')))
        self.c_illus.configure(text=self._elaguer(s.get('illustrations', '—')))
        if hasattr(self, 'chip_format_label'):
            self.chip_format_label.configure(text=f'📐 {lire_format_actuel() or "—"}')

    def _verrouiller(self):
        self.processus_en_cours = True
        for b, _t in self.boutons: b.configure(state='disabled')
        for combo in self.combos:
            try: combo.configure(state='disabled')
            except Exception: pass

    def _reactiver(self):
        self.processus_en_cours = False
        for b, t in self.boutons: b.configure(state='normal', text=t)
        for combo in self.combos:
            try: combo.configure(state='readonly')
            except Exception: pass
        try:
            self.k2000.arreter()
            self.k2000.grid_remove()
        except Exception: pass

    def _occupe(self):
        if self.processus_en_cours:
            messagebox.showinfo('Occupé', 'Une autre opération est déjà en cours.')
            return True
        return False

    def set_statut(self, texte, couleur=MUTED):
        self.statut.configure(text=texte, text_color=couleur)

    def ouvrir_configuration(self):
        if not CONFIG_JSON_OK or not CONFIG_UI_OK:
            messagebox.showerror('Modules manquants', 'configuration_store.py et/ou configuration_ui.py sont introuvables.')
            return
        if self.fenetre_config is None or not self.fenetre_config.winfo_exists():
            self.fenetre_config = cui.FenetreConfiguration(self)
        else:
            self.fenetre_config.focus()

    def _lire_chapitres_config(self):
        titres, mapping = [], {}
        if CONFIG_JSON_OK:
            try:
                titres, mapping = cst.lire_chapitres_pour_traduction()
                if titres: return titres, mapping
            except Exception: pass
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CHEMIN_CONFIG, data_only=True, read_only=True)
            if 'Chapitres' not in wb.sheetnames:
                wb.close()
                return titres, mapping
            ws = wb['Chapitres']
            idx_fichier, idx_ligne1, en_tete_vue = 0, 2, False
            for row in ws.iter_rows(values_only=True):
                cellules = [str(c).strip() if c is not None else '' for c in row]
                if not any(cellules): continue
                basses = [c.lower() for c in cellules]
                if not en_tete_vue and basses and basses[0].startswith('fichier'):
                    for i, nom in enumerate(basses):
                        if nom.startswith('fichier'): idx_fichier = i
                        elif 'ligne 1' in nom: idx_ligne1 = i
                    en_tete_vue = True
                    continue
                if not en_tete_vue: continue
                fichier = cellules[idx_fichier] if idx_fichier < len(cellules) else ''
                titre = cellules[idx_ligne1] if idx_ligne1 < len(cellules) else ''
                if not fichier or fichier.startswith('---') or not titre: continue
                if titre not in mapping: titres.append(titre)
                mapping[titre] = fichier
            wb.close()
            return titres, mapping
        except Exception:
            return titres, mapping

    def lister_cibles_traduction(self):
        titres, mapping = self._lire_chapitres_config()
        self._map_titre_fichier = mapping
        if titres: return ['Tout le dossier'] + titres
        valeurs = ['Tout le dossier']
        if os.path.isdir(DOSSIER_CHAPITRES):
            fichiers = [f for f in os.listdir(DOSSIER_CHAPITRES) if f.lower().endswith('.md') and not f.startswith('.') and f.lower() not in {'résumé.md', 'resume.md'}]
            def cle(n):
                m = re.match(r'^\s*(\d+)', n)
                return (int(m.group(1)) if m else 9999, n.lower())
            fichiers.sort(key=cle)
            valeurs.extend(fichiers)
        return valeurs

    def rafraichir_cibles_traduction(self, silencieux=False):
        valeurs = self.lister_cibles_traduction()
        actuel = self.cible_traduction_var.get()
        self.combo_traduction.configure(values=valeurs)
        if actuel not in valeurs: self.cible_traduction_var.set(valeurs[0] if valeurs else 'Tout le dossier')
        if not silencieux: self.set_statut('📂 Liste des chapitres actualisée.', OK)

    # ─────────────────────────────────────────
    # Actions Principales
    # ─────────────────────────────────────────
    def lancer(self):
        if self._occupe(): return
        if not os.path.isfile(SCRIPT):
            messagebox.showerror('Script introuvable', 'Fichier introuvable :\n' + SCRIPT)
            return
        self._demarrer('▶ Génération en cours…', self.executer)

    def generer_pdf_kdp_gui(self):
        if self._occupe(): return
        if not self.pywin32_ok:
            messagebox.showwarning('Dépendance manquante', 'L\'export PDF nécessite la bibliothèque pywin32.\nVeuillez lancer : pip install pywin32')
            return
        if not detecter_dernier_docx_kdp():
            messagebox.showinfo('Document absent', "Aucun .docx généré pour l'instant.\nLancez d'abord une génération.")
            return
        self._demarrer('🖨 Export PDF KDP en cours…', self.executer_pdf)

    def generer_ebook_gui(self):
        if self._occupe(): return
        if not detecter_dernier_docx_kdp():
            messagebox.showinfo('Document absent', "Aucun .docx généré pour l'instant.")
            return
        if not os.path.isfile(EBOOK_SCRIPT):
            messagebox.showerror('Script introuvable', f'Fichier introuvable :\n{EBOOK_SCRIPT}')
            return
        self._demarrer('📱 Génération ebook EPUB en cours…', self.executer_ebook)

    def generer_resumes(self):
        if self._occupe(): return
        if not os.path.isfile(IA_SCRIPT):
            messagebox.showerror('Script introuvable', 'Fichier introuvable :\n' + IA_SCRIPT)
            return
        self._demarrer('🤖 Résumés IA en cours…', self.executer_ia)

    def traduire_en(self):
        if self._occupe(): return
        if not os.path.isfile(IA_SCRIPT):
            messagebox.showerror('Script introuvable', 'Fichier introuvable :\n' + IA_SCRIPT)
            return
        self.rafraichir_cibles_traduction(silencieux=True)
        cible = self.cible_traduction_var.get().strip() or 'Tout le dossier'
        if cible == 'Tout le dossier':
            args, message = ['--traduire'], '🌐 Traduction de tout le dossier en cours…'
        else:
            fichier = self._map_titre_fichier.get(cible, cible)
            if not os.path.isfile(os.path.join(DOSSIER_CHAPITRES, fichier)):
                messagebox.showerror('Fichier introuvable', f'Le chapitre est introuvable :\n{fichier}')
                return
            args, message = ['--traduire-chapitre', fichier], f'🌐 Traduction de « {cible} » en cours…'
        self._demarrer(message, self.executer_traduction, args=(args,))

    def _demarrer(self, message, cible, args=()):
        self._verrouiller()
        self.set_statut(message, GOLD)
        self.chrono.configure(text='⏳ en cours…')
        
        # v1.26.5 : Afficher le scanner et forcer le rendu avant de démarrer
        self.k2000.grid()
        self.update_idletasks()  # Force le calcul des dimensions du canvas
        
        self.k2000.demarrer()
        threading.Thread(target=cible, args=args, daemon=True).start()

    def ouvrir_dossier_images(self): self.ouvrir_dossier(DOSSIER_IMAGES)
    def ouvrir_dossier_chapitres(self): self.ouvrir_dossier(DOSSIER_CHAPITRES)

    def ouvrir_docx(self):
        cand = glob.glob(os.path.join(BASE, '*_KDP.docx'))
        if cand: os.startfile(max(cand, key=os.path.getmtime))
        else: messagebox.showinfo('Document absent', "Aucun .docx généré pour l'instant.")

    def ouvrir_dossier(self, dossier):
        try:
            os.makedirs(dossier, exist_ok=True)
            os.startfile(dossier)
            self.set_statut(f'📂 Ouvert : {os.path.basename(dossier)}', OK)
        except Exception as e:
            messagebox.showerror('Ouverture impossible', str(e))

    # ─────────────────────────────────────────
    # Exécution des scripts
    # ─────────────────────────────────────────
    def _run_script(self, chemin, cle_code=None, args=None):
        t0 = time.time()
        code = 0
        ancien, ancien_argv = sys.stdout, sys.argv
        if args is not None: sys.argv = [chemin] + list(args)
        sys.stdout = Redirecteur(self.file_sortie)
        try:
            module = runpy.run_path(chemin, run_name='__main__')
            if cle_code and module.get(cle_code, 0) != 0: code = 1
        except SystemExit as e:
            code = e.code if isinstance(e.code, int) else (0 if e.code is None else 1)
        except Exception as e:
            print(f'\n⚠️ Erreur : {e}')
            code = 1
        finally:
            sys.stdout, sys.argv = ancien, ancien_argv
        return code, time.time() - t0

    def executer(self):
        code, duree = self._run_script(SCRIPT)
        self.after(0, self.fin, code, duree)

    def executer_pdf(self):
        t0, code, ancien = time.time(), 0, sys.stdout
        sys.stdout = Redirecteur(self.file_sortie)
        try: generer_pdf_kdp()
        except Exception as e:
            print(f'\n⚠️ Erreur PDF : {e}')
            code = 1
        finally: sys.stdout = ancien
        self.after(0, self.fin_pdf, code, time.time() - t0)

    def executer_ebook(self):
        code, duree = self._run_script(EBOOK_SCRIPT, 'EBOOK_CODE')
        self.after(0, self.fin_ebook, code, duree)

    def executer_ia(self):
        code, duree = self._run_script(IA_SCRIPT, 'IA_ROMAN_CODE')
        self.after(0, self.fin_ia, code, duree)

    def executer_traduction(self, args):
        code, duree = self._run_script(IA_SCRIPT, 'IA_TRADUCTION_CODE', args=args)
        self.after(0, self.fin_traduction, code, duree)

    # ─────────────────────────────────────────
    # Fins de traitement
    # ─────────────────────────────────────────
    def fin(self, code, duree):
        self._reactiver()
        self.chrono.configure(text=f'{duree:.1f} s')
        if code == 0:
            self._maj_cartes(parser_stats('\n'.join(self.sortie)))
            self.set_statut('✅ Terminé — document prêt.', OK)
        else:
            self.set_statut(f'⚠️ Terminé avec erreurs ({duree:.1f} s).', WARN)

    def fin_pdf(self, code, duree):
        self._reactiver()
        self.chrono.configure(text=f'{duree:.1f} s')
        self.set_statut('✅ PDF KDP prêt.' if code == 0 else f'⚠️ Erreur export PDF ({duree:.1f} s).', OK if code == 0 else WARN)

    def fin_ebook(self, code, duree):
        self._reactiver()
        self.chrono.configure(text=f'{duree:.1f} s')
        self.set_statut('✅ Ebook EPUB prêt.' if code == 0 else f'⚠️ Erreur génération ebook ({duree:.1f} s).', OK if code == 0 else WARN)

    def fin_ia(self, code, duree):
        self._reactiver()
        self.chrono.configure(text=f'{duree:.1f} s')
        self.set_statut('✅ Résumés générés.' if code == 0 else f'⚠️ Erreur résumés ({duree:.1f} s).', OK if code == 0 else WARN)

    def fin_traduction(self, code, duree):
        self._reactiver()
        self.chrono.configure(text=f'{duree:.1f} s')
        self.set_statut('✅ Traduction terminée.' if code == 0 else f'⚠️ Erreur traduction ({duree:.1f} s).', OK if code == 0 else WARN)

    # ─────────────────────────────────────────
    # PARAMÉTRAGE DU CORRECTEUR (v1.26.4+)
    # ─────────────────────────────────────────
    def ouvrir_parametres_correcteur(self):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Paramètres du Correcteur")
        dialog.geometry("450x380")
        dialog.transient(self)
        dialog.grab_set()
        
        x = self.winfo_rootx() + (self.winfo_width() // 2) - 225
        y = self.winfo_rooty() + (self.winfo_height() // 2) - 190
        dialog.geometry(f"+{x}+{y}")

        ctk.CTkLabel(dialog, text="⚙️ Configuration du Correcteur", font=('Segoe UI', 14, 'bold'), text_color=GOLD).pack(pady=(15, 10))

        # 1. Niveau
        ctk.CTkLabel(dialog, text="Niveau de vérification :", font=('Segoe UI', 11, 'bold'), anchor='w').pack(fill='x', padx=30, pady=(10, 0))
        current_level = self.config_correcteur.get('level', 'default')
        level_display = "Exigeant (Style, répétitions, clarté)" if current_level == 'picky' else "Standard (Orthographe & Grammaire)"
        var_level = ctk.StringVar(value=level_display)
        
        combo_level = ctk.CTkComboBox(dialog, values=[
            "Standard (Orthographe & Grammaire)",
            "Exigeant (Style, répétitions, clarté)"
        ], variable=var_level, state='readonly')
        combo_level.pack(fill='x', padx=30, pady=5)

        # 2. Variante (Format fr-FR, fr-CA, etc.)
        ctk.CTkLabel(dialog, text="Variante orthographique préférée :", font=('Segoe UI', 11, 'bold'), anchor='w').pack(fill='x', padx=30, pady=(15, 0))
        variant_mapping = {
            'fr-FR': "Français (France)",
            'fr-1990': "Rectifications de 1990",
            'fr-CA': "Français (Canada)",
            'fr-CH': "Français (Suisse)",
            'fr-BE': "Français (Belgique)"
        }
        current_variant = self.config_correcteur.get('preferredVariants', 'fr-FR')
        var_variant = ctk.StringVar(value=variant_mapping.get(current_variant, "Français (France)"))
        
        combo_variant = ctk.CTkComboBox(dialog, values=list(variant_mapping.values()), variable=var_variant, state='readonly')
        combo_variant.pack(fill='x', padx=30, pady=5)

        # 3. Langue maternelle
        ctk.CTkLabel(dialog, text="Votre langue maternelle (optionnel) :", font=('Segoe UI', 11, 'bold'), anchor='w').pack(fill='x', padx=30, pady=(15, 0))
        ctk.CTkLabel(dialog, text="Aide à détecter les calques linguistiques (ex: anglicismes).", font=('Segoe UI', 9), text_color=MUTED, anchor='w').pack(fill='x', padx=30)
        
        mother_mapping = {
            'Aucune': "Aucune",
            'en-US': "Anglais (en-US)",
            'es-ES': "Espagnol (es-ES)",
            'de-DE': "Allemand (de-DE)"
        }
        current_mother = self.config_correcteur.get('motherTongue', 'Aucune')
        var_mother = ctk.StringVar(value=mother_mapping.get(current_mother, "Aucune"))
        
        combo_mother = ctk.CTkComboBox(dialog, values=list(mother_mapping.values()), variable=var_mother, state='readonly')
        combo_mother.pack(fill='x', padx=30, pady=5)

        def sauvegarder():
            level_val = 'picky' if "Exigeant" in combo_level.get() else 'default'
            
            variant_val = 'fr-FR'
            for k, v in variant_mapping.items():
                if v == combo_variant.get():
                    variant_val = k
                    break
                    
            mother_val = 'Aucune'
            for k, v in mother_mapping.items():
                if v == combo_mother.get():
                    mother_val = k
                    break
            
            nouvelle_config = {
                'level': level_val,
                'preferredVariants': variant_val,
                'motherTongue': mother_val
            }
            self._sauvegarder_config_correcteur(nouvelle_config)
            dialog.destroy()

        frame_btn = ctk.CTkFrame(dialog, fg_color='transparent')
        frame_btn.pack(fill='x', padx=30, pady=20)
        ctk.CTkButton(frame_btn, text="Annuler", command=dialog.destroy, fg_color='#555').pack(side='left', padx=5)
        ctk.CTkButton(frame_btn, text="Sauvegarder", command=sauvegarder, fg_color=OK, text_color='#000').pack(side='right', padx=5)

        self.wait_window(dialog)

    # ─────────────────────────────────────────
    # CORRECTEUR ORTHOGRAPHIQUE & GRAMMATICAL
    # ─────────────────────────────────────────
    def lancer_correcteur(self):
        if self._occupe(): return
        
        cible = self.cible_traduction_var.get().strip()
        if not cible or cible == 'Tout le dossier':
            messagebox.showwarning('Sélection requise', 'Veuillez sélectionner un chapitre spécifique dans la liste déroulante\navant de lancer le correcteur.')
            return
        
        fichier = self._map_titre_fichier.get(cible, cible)
        chemin_fichier = os.path.join(DOSSIER_CHAPITRES, fichier)
        
        if not os.path.isfile(chemin_fichier):
            messagebox.showerror('Fichier introuvable', f'Le chapitre est introuvable :\n{fichier}')
            return

        self._demarrer(f'📝 Vérification de « {cible} » en cours…', self._executer_correcteur, args=(chemin_fichier, cible))

    def _executer_correcteur(self, chemin_fichier, titre_chapitre):
        try:
            with open(chemin_fichier, 'r', encoding='utf-8') as f:
                texte_original = f.read()
            
            self.file_sortie.put(f"🔍 Analyse de « {titre_chapitre} » en cours...\n")
            
            matches = spellchecker.verifier_texte(texte_original, langue='fr', config=self.config_correcteur)
            
            if not matches:
                self.file_sortie.put(f"✅ Aucune erreur détectée dans « {titre_chapitre} ».\n")
                self.after(0, self.fin_correcteur, 0, 0, titre_chapitre, chemin_fichier, texte_original, [])
                return

            self.file_sortie.put(f"⚠️ {len(matches)} anomalies potentielles détectées.\n")
            self.after(0, self._lancer_boucle_correction_ui, matches, texte_original, chemin_fichier, titre_chapitre)
            
        except Exception as e:
            self.file_sortie.put(f"⚠️ Erreur lors de l'analyse : {e}\n")
            self.after(0, self.fin_correcteur, 1, 0, titre_chapitre, chemin_fichier, "", [])

    def _lancer_boucle_correction_ui(self, matches, texte_original, chemin_fichier, titre_chapitre):
        corrections_a_appliquer = []
        ignore_tout = False
        
        matches.sort(key=lambda x: x['offset'], reverse=True)
        
        for i, match in enumerate(matches):
            if ignore_tout: break
                
            context = match['context']
            texte_erreur = context['text'][context['offset']:context['offset'] + context['length']]
            self.file_sortie.put(f"  → Erreur {i+1}/{len(matches)} : '{texte_erreur}' ({match['shortMessage']})\n")
            
            action, nouvelle_valeur = self._afficher_dialogue_correction(match, texte_original, titre_chapitre)
            
            if action == 'corriger' and nouvelle_valeur:
                corrections_a_appliquer.append({
                    'offset': context['offset'],
                    'length': context['length'],
                    'nouvelle_valeur': nouvelle_valeur
                })
            elif action == 'ignorer_tout':
                ignore_tout = True
                self.file_sortie.put("  ⏭️ Ignorance des corrections restantes activée.\n")
        
        texte_final = spellchecker.appliquer_corrections(texte_original, corrections_a_appliquer)
        
        if texte_final != texte_original:
            try:
                with open(chemin_fichier, 'w', encoding='utf-8') as f:
                    f.write(texte_final)
                self.file_sortie.put(f"💾 Fichier « {titre_chapitre} » sauvegardé avec les modifications.\n")
            except Exception as e:
                self.file_sortie.put(f"⚠️ Erreur de sauvegarde : {e}\n")
        else:
            self.file_sortie.put(f"ℹ️ Aucune modification appliquée à « {titre_chapitre} ».\n")
            
        nb_corrections = len(corrections_a_appliquer)
        self.after(0, self.fin_correcteur, 0, nb_corrections, titre_chapitre, chemin_fichier, texte_final, corrections_a_appliquer)

    def _afficher_dialogue_correction(self, match, texte_original, titre_chapitre):
        dialog = ctk.CTkToplevel(self)
        dialog.title("Correction Orthographique")
        dialog.geometry("550x400")
        dialog.transient(self)
        dialog.grab_set()
        
        x = self.winfo_rootx() + (self.winfo_width() // 2) - 275
        y = self.winfo_rooty() + (self.winfo_height() // 2) - 200
        dialog.geometry(f"+{x}+{y}")

        ctk.CTkLabel(dialog, text=f"Chapitre : {titre_chapitre}", font=('Segoe UI', 12, 'bold'), text_color=GOLD).pack(pady=(15, 5))
        ctk.CTkLabel(dialog, text=match['message'], font=('Segoe UI', 11), text_color=WARN, wraplength=500, justify='left').pack(pady=5, padx=20)
        
        offset = match['offset']
        length = match['length']
        avant = texte_original[max(0, offset-40) : offset]
        erreur = texte_original[offset : offset+length]
        apres = texte_original[offset+length : min(len(texte_original), offset+length+40)]
        contexte_complet = f"...{avant}[{erreur}]{apres}..."
        
        txt_ctx = ctk.CTkTextbox(dialog, height=60, font=('Consolas', 11), fg_color=CARD, text_color=TXT, state='normal')
        txt_ctx.pack(fill='x', padx=20, pady=5)
        txt_ctx.insert('0.0', contexte_complet)
        txt_ctx.configure(state='disabled')

        ctk.CTkLabel(dialog, text="Suggestions :", font=('Segoe UI', 10, 'bold')).pack(anchor='w', padx=20, pady=(10, 0))
        suggestions = [r['value'] for r in match.get('replacements', [])[:5]]
        var_suggestion = ctk.StringVar(value=suggestions[0] if suggestions else "")
        combo = ctk.CTkComboBox(dialog, values=suggestions if suggestions else ["Aucune suggestion automatique"], variable=var_suggestion, state='readonly' if suggestions else 'disabled')
        combo.pack(fill='x', padx=20, pady=5)

        ctk.CTkLabel(dialog, text="Ou modifier manuellement :", font=('Segoe UI', 10)).pack(anchor='w', padx=20, pady=(5, 0))
        entry_manuel = ctk.CTkEntry(dialog, font=('Segoe UI', 11))
        entry_manuel.pack(fill='x', padx=20, pady=5)

        resultat = {'action': 'ignorer', 'valeur': ''}

        def valider_corr():
            resultat['action'] = 'corriger'
            resultat['valeur'] = entry_manuel.get() if entry_manuel.get() else var_suggestion.get()
            dialog.destroy()

        def modifier_manuel():
            resultat['action'] = 'corriger'
            resultat['valeur'] = entry_manuel.get()
            dialog.destroy()

        def ignorer():
            resultat['action'] = 'ignorer'
            dialog.destroy()

        def ignorer_tout():
            resultat['action'] = 'ignorer_tout'
            dialog.destroy()

        frame_btn = ctk.CTkFrame(dialog, fg_color='transparent')
        frame_btn.pack(fill='x', padx=20, pady=15)
        
        ctk.CTkButton(frame_btn, text="Appliquer suggestion", command=valider_corr, fg_color=OK, text_color='#000').pack(side='left', padx=5)
        ctk.CTkButton(frame_btn, text="Modifier manuellement", command=modifier_manuel, fg_color=ACCENT).pack(side='left', padx=5)
        ctk.CTkButton(frame_btn, text="Ignorer", command=ignorer, fg_color='#555').pack(side='left', padx=5)
        ctk.CTkButton(frame_btn, text="Ignorer tout le reste", command=ignorer_tout, fg_color=WARN, text_color='#fff').pack(side='right', padx=5)

        self.wait_window(dialog)
        return resultat['action'], resultat['valeur']

    def fin_correcteur(self, code, nb_corrections, titre, chemin, texte_final, corrections):
        self._reactiver()
        if code == 0:
            self.set_statut(f'✅ Correction terminée : {nb_corrections} modification(s) appliquée(s).', OK)
            self._log(f"📊 SYNTHÈSE CORRECTION : {nb_corrections} erreur(s) corrigée(s) dans « {titre} ».")
        else:
            self.set_statut('⚠️ Erreur lors de la correction.', WARN)


def parser_stats(texte):
    s = {
        'ouvrage': '—', 'actes': '—', 'chapitres': '—', 'dialogues': '—',
        'illustrations': '—', 'mots': '—', 'mots_moy': '—', 'pages': '—',
        'lecture': '—', 'detail': []
    }

    def val(l):
        return l.split(':', 1)[1].strip() if ':' in l else '—'

    for ligne in texte.splitlines():
        l = ligne.strip()
        if not l: continue
        if l.startswith('📖'): s['ouvrage'] = val(l)
        elif l.startswith('🎭'): s['actes'] = val(l)
        elif l.startswith('📚'): s['chapitres'] = val(l)
        elif l.startswith('💬'): s['dialogues'] = val(l)
        elif l.startswith('🖼'): s['illustrations'] = val(l)
        elif l.startswith('🔤'): s['mots'] = val(l)
        elif l.startswith('📊'): s['mots_moy'] = val(l)
        elif l.startswith('📑'): s['pages'] = val(l)
        elif l.startswith('⏱'): s['lecture'] = val(l).replace('(à 230 mots/min)', '').strip()
        elif l.startswith('•'):
            m = re.match(r'•\s*(.*?)\s+([\d\u202f]+)\s*mots', l)
            if m:
                s['detail'].append((m.group(1).strip(), m.group(2)))
    return s


if __name__ == '__main__':
    app = StudioApp()
    app.mainloop()