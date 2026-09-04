#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
IA_Roman.py
===========
Génère un fichier "Résumé.md" à partir des chapitres Markdown contenus dans le dossier "Chapitres".
Ajoute également une fonction de traduction anglaise des chapitres.

Fonctions principales :
1. Génération de résumés de chapitres avec IA.
2. Traduction des chapitres en anglais via IA (tout le dossier ou un chapitre précis).

IA par défaut : Ollama.
Fournisseurs disponibles : Ollama (local ou cloud) et OpenAI.
"""

import os
import re
import sys
import json
import time
import unicodedata
import urllib.request
import urllib.error
from datetime import datetime

# ─────────────────────────────────────────
# Chemins
# ─────────────────────────────────────────
if getattr(sys, "frozen", False):
    BASE = os.path.dirname(sys.executable)
else:
    BASE = os.path.dirname(os.path.abspath(__file__))

DOSSIER_CHAPITRES = os.path.join(BASE, "Chapitres")
CHEMIN_CONFIG = os.path.join(BASE, "Configuration_roman.xlsx")
CHEMIN_RESUME = os.path.join(BASE, "Résumé.md")
DOSSIER_TRADUCTIONS = os.path.join(BASE, "Traductions")
CHEMIN_PROMPT_TRADUCTION = os.path.join(BASE, "Traduction.md")

FEUILLE_IA = "IA"

IA_ROMAN_CODE = 0
IA_TRADUCTION_CODE = 0

# ─────────────────────────────────────────
# Réglages
# ─────────────────────────────────────────
def entier_env(nom, defaut):
    try:
        return int(os.getenv(nom, defaut))
    except Exception:
        return defaut


def float_env(nom, defaut):
    try:
        return float(os.getenv(nom, defaut))
    except Exception:
        return defaut


LONGUEUR_RESUME = entier_env("IA_RESUME_LONGUEUR", 160)
PAUSE_CHAPITRES = float_env("IA_ROMAN_PAUSE", 0.4)
MODE_IA = os.getenv("IA_ROMAN_MODE", "ollama").strip().lower()

OPENAI_BASE_URL = (
    os.getenv("OPENAI_BASE_URL")
    or os.getenv("IA_OPENAI_BASE_URL")
    or "https://api.openai.com/v1"
)

OPENAI_MODEL = (
    os.getenv("OPENAI_MODEL")
    or os.getenv("IA_OPENAI_MODEL")
    or "gpt-4o-mini"
)

OLLAMA_URL = (
    os.getenv("IA_OLLAMA_URL")
    or os.getenv("OLLAMA_URL")
    or "http://localhost:11434"
)

OLLAMA_MODEL = (
    os.getenv("IA_OLLAMA_MODEL")
    or os.getenv("OLLAMA_MODEL")
    or "llama3.1"
)

OLLAMA_API_KEY = (
    os.getenv("IA_OLLAMA_API_KEY")
    or os.getenv("OLLAMA_API_KEY")
    or ""
)

PARAMS_IA = {}
TOKENS_SESSION = 0

_OLLAMA_ERREUR_SIGNAL = False
_OPENAI_ERREUR_SIGNAL = False


def ajouter_tokens(n):
    global TOKENS_SESSION
    try:
        TOKENS_SESSION += int(n or 0)
    except Exception:
        pass


# ─────────────────────────────────────────
# Outils texte
# ─────────────────────────────────────────
STOPWORDS = set("""
le la les un une des de du au aux et ou mais donc or ni car que qui quoi dont
il elle on nous vous ils elles je tu ce cet cette ces mon ma mes ton ta tes
son sa ses notre nos votre vos leur leurs
a au aux avec sans chez dans de des du entre vers pour par sur sous
etre avoir fait faire dit dire aller venir voir savoir pouvoir vouloir devoir falloir
""".split())


def sans_accents(texte):
    return unicodedata.normalize("NFKD", str(texte)).encode("ascii", "ignore").decode("utf-8")


def valeur_texte(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def romain_vers_int(txt):
    valeurs = {"I": 1, "V": 5, "X": 10, "L": 50, "C": 100, "D": 500, "M": 1000}
    txt = str(txt).upper().strip()

    if not txt or any(ch not in valeurs for ch in txt):
        return None

    total = 0
    prev = 0

    for ch in reversed(txt):
        val = valeurs[ch]
        if val < prev:
            total -= val
        else:
            total += val
        prev = val

    return total or None


def ordre_num(v):
    if v is None:
        return 999999

    s = str(v).strip()
    if not s:
        return 999999

    if s.isdigit():
        return int(s)

    r = romain_vers_int(s)
    if r:
        return r

    m = re.search(r"\d+", s)
    if m:
        return int(m.group(0))

    m = re.search(r"\b[IVXLCDM]+\b", s.upper())
    if m:
        r = romain_vers_int(m.group(0))
        if r:
            return r

    return 999999


def nettoyer_nom_fichier(nom):
    base = os.path.splitext(os.path.basename(nom))[0]
    base = re.sub(r"[_\-]+", " ", base)
    return re.sub(r"\s+", " ", base).strip() or nom


def nettoyer_titre_chapitre(titre):
    if not titre:
        return ""

    titre = str(titre).strip()
    titre = re.sub(r"(?i)^chapitre\s*(?:\d+|[ivxlcdm]+)\s*[-–—:]?\s*", "", titre)
    titre = re.sub(r"(?i)^acte\s*(?:\d+|[ivxlcdm]+)\s*[-–—:]?\s*", "", titre)
    return titre.strip()


def extraire_titre_markdown(texte):
    for ligne in texte.splitlines():
        l = ligne.strip()
        if l.startswith("#"):
            return l.lstrip("#").strip()
    return ""


def nettoyer_markdown(texte):
    texte = re.sub(r"```.*?```", " ", texte, flags=re.S)
    texte = re.sub(r"`[^`]*`", " ", texte)
    texte = re.sub(r"!\[[^\]]*\]\([^)]*\)", " ", texte)
    texte = re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", texte)
    texte = re.sub(r"<[^>]+>", " ", texte)
    texte = re.sub(r"^#{1,6}\s*", "", texte, flags=re.M)
    texte = re.sub(r"^>\s?", "", texte, flags=re.M)
    texte = re.sub(r"^\s*[-*+]\s+", "", texte, flags=re.M)
    texte = re.sub(r"[*_~]+", " ", texte)
    return re.sub(r"\s+", " ", texte).strip()


def tronquer_texte(texte, max_caracteres=8000):
    if len(texte) <= max_caracteres:
        return texte

    tete = int(max_caracteres * 0.6)
    queue = max_caracteres - tete - 20
    return texte[:tete] + "\n[...]\n" + texte[-queue:]


def decouper_phrases(texte):
    texte = re.sub(r"\s+", " ", texte)
    phrases = re.split(r"(?<=[.!?…])\s+", texte)
    return [p.strip() for p in phrases if len(p.strip()) > 25]


# ─────────────────────────────────────────
# Paramètres IA stockés dans Excel
# ─────────────────────────────────────────
def normaliser_label_excel(txt):
    if txt is None:
        return ""
    txt = sans_accents(str(txt).strip().lower())
    return re.sub(r"[^a-z0-9]+", " ", txt).strip()


def lignes_standard_ia():
    return [
        ("Mode IA", "ollama"),
        ("URL Ollama", OLLAMA_URL),
        ("Modèle Ollama", OLLAMA_MODEL),
        ("Clé API Ollama", ""),
        ("Clé API OpenAI", ""),
        ("Modèle OpenAI", OPENAI_MODEL),
        ("URL API OpenAI", OPENAI_BASE_URL),
    ]


def supprimer_parametres_obsoletes():
    if not os.path.isfile(CHEMIN_CONFIG):
        return

    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG)
        modification = False

        for nom in list(wb.sheetnames):
            if nom.strip().lower() not in {FEUILLE_IA.lower(), "informations"}:
                continue

            ws = wb[nom]
            lignes_a_supprimer = []

            for row in ws.iter_rows(min_col=1, max_col=2):
                if row[0].value is None:
                    continue

                label = normaliser_label_excel(row[0].value)
                if not label:
                    continue

                if "qwen" in label or "dashscope" in label:
                    lignes_a_supprimer.append(row[0].row)
                    continue

                if label in ("mode ia", "mode", "fournisseur ia", "ia mode"):
                    if len(row) > 1 and row[1].value is not None:
                        if str(row[1].value).strip().lower() in ("qwen", "dashscope"):
                            row[1].value = "ollama"
                            modification = True

            for ligne in sorted(lignes_a_supprimer, reverse=True):
                ws.delete_rows(ligne, 1)
                modification = True

        if modification:
            wb.save(CHEMIN_CONFIG)

        wb.close()

    except Exception:
        pass


def assurer_feuille_ia():
    if not os.path.isfile(CHEMIN_CONFIG):
        return

    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG)
        ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == FEUILLE_IA.lower()), None)

        if ws is None:
            ws = wb.create_sheet(FEUILLE_IA)
            ws.append(["Paramètre", "Valeur"])

            for label, valeur in lignes_standard_ia():
                ws.append([label, valeur])

            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 60
            wb.save(CHEMIN_CONFIG)
            wb.close()
            return

        existants = {
            normaliser_label_excel(r[0].value)
            for r in ws.iter_rows(min_col=1, max_col=2)
            if r[0].value
        }

        ajout = False
        for label, valeur in lignes_standard_ia():
            if normaliser_label_excel(label) not in existants:
                ws.append([label, valeur])
                ajout = True

        if ajout:
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 60
            wb.save(CHEMIN_CONFIG)

        wb.close()

    except Exception:
        pass


def lire_parametres_ia_excel():
    params = {}

    if not os.path.isfile(CHEMIN_CONFIG):
        return params

    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG, data_only=True, read_only=True)
        ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == FEUILLE_IA.lower()), None)

        if ws is None:
            ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == "informations"), None)

        if ws is None:
            wb.close()
            return params

        placeholders = {"", "none", "sk-...", "ta_cle_api", "ta_cle", "votre_cle_api", "votre_cle"}

        for row in ws.iter_rows(values_only=True):
            if not row or row[0] is None:
                continue

            label = normaliser_label_excel(row[0])
            valeur = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

            if valeur.lower() in placeholders:
                continue

            if label in ("mode ia", "mode", "fournisseur ia", "ia mode"):
                params["MODE_IA"] = valeur.lower()

            elif label in ("ollama api key", "cle api ollama", "ollama key", "api key ollama", "cle ollama"):
                params["OLLAMA_API_KEY"] = valeur

            elif label in ("ollama url", "url ollama"):
                params["OLLAMA_URL"] = valeur

            elif label in ("ollama model", "modele ollama", "model ollama"):
                params["OLLAMA_MODEL"] = valeur

            elif label in ("openai api key", "cle api openai", "openai key", "api key openai", "cle openai"):
                params["OPENAI_API_KEY"] = valeur

            elif label in ("openai model", "modele openai", "model openai"):
                params["OPENAI_MODEL"] = valeur

            elif label in ("openai base url", "url api openai", "url openai", "base url openai"):
                params["OPENAI_BASE_URL"] = valeur

        wb.close()

    except Exception:
        pass

    return params


def charger_parametres_ia():
    global PARAMS_IA, MODE_IA, OPENAI_BASE_URL, OPENAI_MODEL, OLLAMA_URL, OLLAMA_MODEL, OLLAMA_API_KEY

    supprimer_parametres_obsoletes()
    assurer_feuille_ia()

    PARAMS_IA = lire_parametres_ia_excel()

    mode_excel = PARAMS_IA.get("MODE_IA", "").strip().lower()

    if mode_excel in ("qwen", "dashscope"):
        MODE_IA = "ollama"
    elif mode_excel:
        MODE_IA = mode_excel

    if MODE_IA not in {"ollama", "openai", "auto", "offline"}:
        MODE_IA = "ollama"

    if PARAMS_IA.get("OPENAI_BASE_URL"):
        OPENAI_BASE_URL = PARAMS_IA["OPENAI_BASE_URL"].strip()

    if PARAMS_IA.get("OPENAI_MODEL"):
        OPENAI_MODEL = PARAMS_IA["OPENAI_MODEL"].strip()

    if PARAMS_IA.get("OLLAMA_URL"):
        OLLAMA_URL = PARAMS_IA["OLLAMA_URL"].strip()

    if PARAMS_IA.get("OLLAMA_MODEL"):
        OLLAMA_MODEL = PARAMS_IA["OLLAMA_MODEL"].strip()

    if PARAMS_IA.get("OLLAMA_API_KEY"):
        OLLAMA_API_KEY = PARAMS_IA["OLLAMA_API_KEY"].strip()


def enregistrer_parametre_ia(label, valeur):
    if not os.path.isfile(CHEMIN_CONFIG):
        print("❌ Configuration_roman.xlsx introuvable.")
        return False

    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG)
        ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == FEUILLE_IA.lower()), None)

        if ws is None:
            ws = wb.create_sheet(FEUILLE_IA)
            ws.append(["Paramètre", "Valeur"])
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 60

        label_norm = normaliser_label_excel(label)
        ligne_cible = next(
            (
                r[0].row
                for r in ws.iter_rows(min_col=1, max_col=2)
                if r[0].value and normaliser_label_excel(r[0].value) == label_norm
            ),
            None
        )

        val_str = "" if valeur is None else str(valeur)

        if ligne_cible:
            ws.cell(row=ligne_cible, column=2).value = val_str
        else:
            ws.append([label, val_str])

        wb.save(CHEMIN_CONFIG)
        wb.close()
        return True

    except Exception as e:
        print(f"❌ Impossible d'enregistrer le paramètre : {e}")
        return False


# ─────────────────────────────────────────
# Lecture du plan depuis Excel
# ─────────────────────────────────────────
def lire_plan_depuis_excel():
    if not os.path.isfile(CHEMIN_CONFIG):
        return []

    try:
        from openpyxl import load_workbook
        wb = load_workbook(CHEMIN_CONFIG, data_only=True, read_only=True)
        ws = next((wb[n] for n in wb.sheetnames if n.strip().lower() == "chapitres"), None)

        if not ws:
            wb.close()
            return []

        lignes = list(ws.iter_rows(values_only=True))
        wb.close()

        header_idx, cols = None, {}

        for i, row in enumerate(lignes[:40]):
            if not row or all(c is None for c in row):
                continue

            norm = [sans_accents(str(c).strip().lower()) if c else "" for c in row]
            current = {}

            for j, c in enumerate(norm):
                if not c:
                    continue

                if ("acte" in c or c == "act") and "acte" not in current:
                    current["acte"] = j

                if ("chapitre" in c or c.startswith("chap")) and "chapitre" not in current:
                    current["chapitre"] = j

                if ("fichier" in c or "file" in c or "nom" in c) and "fichier" not in current:
                    current["fichier"] = j

                if ("titre" in c or "title" in c) and "titre" not in current:
                    current["titre"] = j

            if "fichier" in current and len(current) > 1:
                header_idx, cols = i, current
                break

        if header_idx is None:
            return []

        plan, acte_courant = [], None

        for row in lignes[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            def get(k):
                return row[cols[k]] if k in cols and cols[k] < len(row) else None

            fichier = valeur_texte(get("fichier"))
            if not fichier:
                continue

            if not fichier.lower().endswith(".md"):
                fichier += ".md"

            chemin = os.path.join(DOSSIER_CHAPITRES, os.path.basename(fichier))

            if not os.path.isfile(chemin):
                chemin_alt = os.path.join(BASE, fichier)
                if os.path.isfile(chemin_alt):
                    chemin = chemin_alt
                else:
                    continue

            acte = valeur_texte(get("acte"))
            if acte:
                acte_courant = acte
            else:
                acte = acte_courant or "1"

            plan.append({
                "acte": acte,
                "chapitre": valeur_texte(get("chapitre")),
                "titre": valeur_texte(get("titre")),
                "fichier": os.path.basename(fichier),
                "chemin": chemin,
            })

        return plan

    except Exception:
        return []


# ─────────────────────────────────────────
# Analyse des fichiers Markdown
# ─────────────────────────────────────────
def lister_fichiers_chapitres():
    if not os.path.isdir(DOSSIER_CHAPITRES):
        return []

    exclus = {"résumé.md", "resume.md"}
    fichiers = [
        n for n in os.listdir(DOSSIER_CHAPITRES)
        if n.lower().endswith(".md")
        and n.lower() not in exclus
        and not n.startswith(".")
    ]

    return sorted(
        fichiers,
        key=lambda n: (
            int(m.group(1)) if (m := re.match(r"^\s*(\d+)", n)) else 9999,
            sans_accents(n.lower())
        )
    )


def inferer_metadonnees(nom, contenu):
    base = sans_accents(os.path.splitext(nom)[0].lower())

    acte_m = re.search(r"acte[\s_-]*([0-9]+|[ivxlcdm]+)", base)
    chap_m = re.search(r"chap(?:itre)?[\s_-]*([0-9]+|[ivxlcdm]+)", base)

    acte = acte_m.group(1) if acte_m else "1"
    chapitre = chap_m.group(1) if chap_m else ""

    if not chapitre:
        m = re.match(r"^\s*(\d+)", nom)
        if m:
            chapitre = m.group(1)

    titre = nettoyer_titre_chapitre(extraire_titre_markdown(contenu))
    if not titre:
        titre = nettoyer_nom_fichier(nom)

    titre = re.sub(r"(?i)^acte\s*\w+\s*[-–—]?\s*", "", titre)
    titre = re.sub(r"(?i)^chap(?:itre)?\s*\w+\s*[-–—]?\s*", "", titre).strip()

    return acte, chapitre, titre


def completer_chapitres_manquants(plan):
    compte = {}

    for item in plan:
        acte_key = sans_accents(str(item.get("acte", "1")).lower())

        if not str(item.get("chapitre", "")).strip():
            compte[acte_key] = compte.get(acte_key, 0) + 1
            item["chapitre"] = str(compte[acte_key])

    return plan


def construire_plan():
    plan = lire_plan_depuis_excel()

    if plan:
        return completer_chapitres_manquants(plan)

    plan = []

    for nom in lister_fichiers_chapitres():
        chemin = os.path.join(DOSSIER_CHAPITRES, nom)

        try:
            with open(chemin, encoding="utf-8", errors="ignore") as f:
                contenu = f.read()
        except Exception:
            contenu = ""

        acte, chapitre, titre = inferer_metadonnees(nom, contenu)
        plan.append({
            "acte": acte,
            "chapitre": chapitre,
            "titre": titre,
            "fichier": nom,
            "chemin": chemin
        })

    return completer_chapitres_manquants(plan)


def grouper_par_acte(plan):
    groupes = {}

    for item in plan:
        acte = str(item.get("acte") or "1").strip() or "1"
        groupes.setdefault(acte, []).append(item)

    for acte, items in groupes.items():
        items.sort(key=lambda it: (ordre_num(it.get("chapitre")), sans_accents(it.get("fichier", "").lower())))

    return sorted(groupes.items(), key=lambda kv: (ordre_num(kv[0]), sans_accents(kv[0].lower())))


# ─────────────────────────────────────────
# Appels IA
# ─────────────────────────────────────────
def construire_prompt(titre, texte):
    return f"""Rôle : Tu es un critique littéraire expert en analyse narrative.
Tâche : Résume ce chapitre en 10 à 20 lignes maximum.
Contraintes :
- Style : littéraire, fluide, au présent
- Focus : intrigue principale, personnages clés, enjeux narratifs
- Structure : début → développement → fin (sans formules type "Ce chapitre...")
- Ton : neutre mais engageant, comme une quatrième de couverture
Chapitre : {titre}
Texte du chapitre :
{texte}
Résumé :"""


def construire_url_ollama(base):
    base = (base or "").strip().rstrip("/")

    if base.endswith("/api/generate"):
        return base

    if base.endswith("/api"):
        return base + "/generate"

    return base + "/api/generate"


def appeler_ollama(prompt, system=None, temperature=0.3):
    """Appelle un modèle Ollama avec fallback automatique sur l'API OpenAI-compatible."""
    global _OLLAMA_ERREUR_SIGNAL

    url_native = construire_url_ollama(OLLAMA_URL)

    payload_native = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False,
        "options": {"temperature": temperature}
    }

    if system:
        payload_native["system"] = system

    headers = {"Content-Type": "application/json"}
    if OLLAMA_API_KEY:
        headers["Authorization"] = f"Bearer {OLLAMA_API_KEY}"

    req = urllib.request.Request(
        url_native,
        data=json.dumps(payload_native).encode("utf-8"),
        headers=headers
    )

    try:
        with urllib.request.urlopen(req, timeout=240) as rep:
            data = json.loads(rep.read().decode("utf-8"))
            response = data.get("response", "").strip()

            if response:
                ajouter_tokens((data.get("prompt_eval_count") or 0) + (data.get("eval_count") or 0))
                return response, f"Ollama ({OLLAMA_MODEL})"

    except urllib.error.HTTPError as e:
        corps = ""
        try:
            corps = e.read().decode("utf-8", errors="ignore")[:500]
        except Exception:
            pass

        if not _OLLAMA_ERREUR_SIGNAL:
            print(f"⚠️ Ollama : erreur HTTP {e.code} sur API native")
            print(f"   URL : {url_native}")
            print(f"   Modèle : {OLLAMA_MODEL}")
            if corps:
                print(f"   Serveur : {corps}")
            print("   Tentative avec l'API OpenAI-compatible...")

        base = OLLAMA_URL.rstrip("/")
        url_openai = base + "/v1/chat/completions"

        messages = []
        if system:
            messages.append({"role": "system", "content": system})
        else:
            messages.append({"role": "system", "content": "Tu es un critique littéraire expert en analyse narrative."})

        messages.append({"role": "user", "content": prompt})

        payload_openai = {
            "model": OLLAMA_MODEL,
            "temperature": temperature,
            "messages": messages
        }

        req2 = urllib.request.Request(
            url_openai,
            data=json.dumps(payload_openai).encode("utf-8"),
            headers=headers
        )

        try:
            with urllib.request.urlopen(req2, timeout=240) as rep:
                data = json.loads(rep.read().decode("utf-8"))
                choices = data.get("choices", [])

                if choices:
                    msg = choices[0].get("message", {}).get("content", "").strip()
                    if msg:
                        usage = data.get("usage") or {}
                        ajouter_tokens(
                            usage.get("total_tokens")
                            or (usage.get("prompt_tokens") or 0) + (usage.get("completion_tokens") or 0)
                        )
                        return msg, f"Ollama-OpenAI ({OLLAMA_MODEL})"

        except urllib.error.HTTPError as e2:
            corps2 = ""
            try:
                corps2 = e2.read().decode("utf-8", errors="ignore")[:500]
            except Exception:
                pass

            if not _OLLAMA_ERREUR_SIGNAL:
                print(f"   Échec aussi sur OpenAI-compatible : HTTP {e2.code}")
                if corps2:
                    print(f"   Serveur : {corps2}")
                print()
                print("   💡 Cause probable : le modèle n'est pas disponible en cloud.")
                print("   Solution 1 : dans Excel, remplace le modèle par llama3.2:3b")
                print("   Solution 2 : passe en Ollama local (http://localhost:11434)")

            _OLLAMA_ERREUR_SIGNAL = True

        except Exception as e2:
            if not _OLLAMA_ERREUR_SIGNAL:
                print(f"   Erreur OpenAI-compatible : {type(e2).__name__}")
            _OLLAMA_ERREUR_SIGNAL = True

        return None

    except Exception as e:
        if not _OLLAMA_ERREUR_SIGNAL:
            print(f"⚠️ Ollama indisponible : {type(e).__name__}")
            print(f"   URL : {url_native}")
        _OLLAMA_ERREUR_SIGNAL = True
        return None

    return None


def obtenir_cle_openai():
    return (
        PARAMS_IA.get("OPENAI_API_KEY")
        or os.getenv("OPENAI_API_KEY")
        or os.getenv("IA_OPENAI_API_KEY")
    )


def appeler_openai(prompt, system=None, temperature=0.3):
    global _OPENAI_ERREUR_SIGNAL

    api_key = obtenir_cle_openai()
    if not api_key:
        return None

    base = OPENAI_BASE_URL.rstrip("/")
    url = base if base.endswith("/chat/completions") else base + "/chat/completions"

    payload = {
        "model": OPENAI_MODEL,
        "temperature": temperature,
        "messages": [
            {
                "role": "system",
                "content": system or "Tu es un critique littéraire expert en analyse narrative."
            },
            {
                "role": "user",
                "content": prompt
            }
        ]
    }

    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}"
        }
    )

    try:
        with urllib.request.urlopen(req, timeout=180) as rep:
            data = json.loads(rep.read().decode("utf-8"))
            choices = data.get("choices", [])

            if choices:
                msg = choices[0].get("message", {}).get("content", "").strip()
                if msg:
                    usage = data.get("usage") or {}
                    ajouter_tokens(
                        usage.get("total_tokens")
                        or (usage.get("prompt_tokens") or 0) + (usage.get("completion_tokens") or 0)
                    )
                    return msg, f"OpenAI ({OPENAI_MODEL})"

    except Exception as e:
        if not _OPENAI_ERREUR_SIGNAL:
            print(f"⚠️ OpenAI indisponible : {type(e).__name__}")
        _OPENAI_ERREUR_SIGNAL = True

    return None


def resumer_hors_ligne(texte, longueur_cible=160):
    phrases = decouper_phrases(texte)

    if not phrases:
        return texte[:500].strip()

    mots = re.findall(r"[a-z']+", sans_accents(texte.lower()))
    freq = {}

    for mot in mots:
        if mot in STOPWORDS or len(mot) < 3:
            continue
        freq[mot] = freq.get(mot, 0) + 1

    scored = []

    for i, phrase in enumerate(phrases):
        tokens = re.findall(r"[a-z']+", sans_accents(phrase.lower()))
        if not tokens:
            continue

        score = sum(freq.get(t, 0) for t in tokens if t not in STOPWORDS) / (len(tokens) + 1)
        scored.append((score, i, phrase))

    scored.sort(reverse=True)

    selection, total_mots = [], 0

    for score, i, phrase in scored:
        if len(selection) >= 6:
            break

        selection.append((i, phrase))
        total_mots += len(phrase.split())

        if total_mots >= longueur_cible:
            break

    selection.sort(key=lambda x: x[0])

    return " ".join(p for _, p in selection).strip() or phrases[0]


def resumer_chapitre(titre, texte):
    texte_net = nettoyer_markdown(texte)

    if not texte_net:
        return "Chapitre vide ou impossible à résumer.", "aucun texte"

    prompt = construire_prompt(titre, tronquer_texte(texte_net))

    if MODE_IA in ("ollama", "auto"):
        try:
            res = appeler_ollama(prompt)
            if res:
                return res
        except Exception:
            pass

    if MODE_IA in ("openai", "auto"):
        try:
            res = appeler_openai(prompt)
            if res:
                return res
        except Exception:
            pass

    return resumer_hors_ligne(texte_net, LONGUEUR_RESUME), "Résumé automatique hors IA"


# ─────────────────────────────────────────
# Traduction EN
# ─────────────────────────────────────────
SYSTEME_TRADUCTION = (
    "You are a professional literary translator and English-language editor. "
    "You translate French novels into natural, publishable International English."
)


def charger_prompt_traduction():
    if os.path.isfile(CHEMIN_PROMPT_TRADUCTION):
        try:
            with open(CHEMIN_PROMPT_TRADUCTION, encoding="utf-8", errors="ignore") as f:
                contenu = f.read().strip()
            if contenu:
                return contenu
        except Exception:
            pass

    return (
        "Tu es un traducteur littéraire professionnel. "
        "Traduis le chapitre suivant du français vers un anglais international naturel, "
        "en conservant l'histoire, les personnages, le ton, les émotions, le rythme et les dialogues. "
        "N'ajoute rien, ne supprime rien et ne fais aucun commentaire."
    )


def nom_fichier_depuis_titre(titre):
    titre = str(titre or "").strip()
    titre = re.sub(r"^#{1,6}\s*", "", titre)
    titre = re.sub(r"[*_`]+", "", titre)
    titre = sans_accents(titre)
    titre = re.sub(r"[\\/:*?\"<>|\r\n\t]+", " ", titre)
    titre = re.sub(r"\s+", " ", titre).strip(" .")

    if len(titre) > 120:
        titre = titre[:120].rstrip()

    return titre or "Untitled"


def extraire_titre_traduit(contenu):
    for ligne in contenu.splitlines():
        l = ligne.strip()
        if l.startswith("#"):
            return l.lstrip("#").strip()

    for ligne in contenu.splitlines():
        l = ligne.strip()
        if not l:
            continue

        if l.startswith(('"', "'", "—", "-")):
            return ""

        if len(l) <= 120 and not l.endswith((".", "!", "?", "…")):
            return l

        break

    return ""


def construire_prompt_traduction(texte):
    prompt_base = charger_prompt_traduction()

    return f"""{prompt_base}

---

TEXTE ORIGINAL À TRADUIRE :

{texte}

Restitue uniquement le chapitre traduit, sans note, sans commentaire et sans explication."""


def traduire_titre_chapitre(titre):
    titre = str(titre or "").strip()
    if not titre:
        return ""

    prompt = (
        "Traduis uniquement ce titre de chapitre du français vers l'anglais. "
        "Retourne uniquement le titre traduit, sans guillemets, sans commentaire "
        "et sans ponctuation finale inutile.\n\n"
        f"Titre : {titre}"
    )

    if MODE_IA in ("ollama", "auto"):
        try:
            res = appeler_ollama(prompt, system=SYSTEME_TRADUCTION, temperature=0.1)
            if res:
                return res[0].strip().splitlines()[0].strip()
        except Exception:
            pass

    if MODE_IA in ("openai", "auto"):
        try:
            res = appeler_openai(prompt, system=SYSTEME_TRADUCTION, temperature=0.1)
            if res:
                return res[0].strip().splitlines()[0].strip()
        except Exception:
            pass

    return ""


def traduire_chapitre(titre, texte):
    if not texte.strip():
        return None, "aucun texte"

    prompt = construire_prompt_traduction(texte)

    if MODE_IA in ("ollama", "auto"):
        try:
            res = appeler_ollama(prompt, system=SYSTEME_TRADUCTION, temperature=0.2)
            if res:
                return res
        except Exception:
            pass

    if MODE_IA in ("openai", "auto"):
        try:
            res = appeler_openai(prompt, system=SYSTEME_TRADUCTION, temperature=0.2)
            if res:
                return res
        except Exception:
            pass

    return None, "aucune IA disponible"


def main_traduction(fichiers_cibles=None):
    print("\U0001F310 IA_Roman : traduction des chapitres en anglais")
    print(f"Dossier chapitres : {DOSSIER_CHAPITRES}")

    charger_parametres_ia()
    print(f"Mode IA : {MODE_IA}")

    if MODE_IA == "offline":
        print("❌ Le mode 'offline' ne permet pas de traduire avec l'IA.")
        return 1

    if not os.path.isdir(DOSSIER_CHAPITRES):
        print("❌ Dossier Chapitres introuvable.")
        return 1

    plan = construire_plan()
    if not plan:
        print("❌ Aucun chapitre .md détecté dans le dossier Chapitres.")
        return 1

    if fichiers_cibles:
        cibles = {sans_accents(os.path.basename(f).lower()) for f in fichiers_cibles}
        plan = [
            item for item in plan
            if sans_accents(os.path.basename(item.get("fichier", "")).lower()) in cibles
        ]

        if not plan:
            print("❌ Aucun chapitre cible trouvé.")
            return 1

    print(f"\U0001F4DA Chapitres à traduire : {len(plan)}")
    print(
        f"\U0001F4DD Prompt de traduction : "
        f"{CHEMIN_PROMPT_TRADUCTION if os.path.isfile(CHEMIN_PROMPT_TRADUCTION) else 'intégré (fallback)'}"
    )

    os.makedirs(DOSSIER_TRADUCTIONS, exist_ok=True)
    print(f"Dossier de sortie : {DOSSIER_TRADUCTIONS}")

    noms_utilises = set()
    succes = 0

    for item in plan:
        fichier = item.get("fichier", "")
        print(f"\U0001F4D6 Traduction : {fichier}")

        try:
            with open(item["chemin"], encoding="utf-8", errors="ignore") as f:
                contenu = f.read()
        except Exception as e:
            print(f"   ❌ Lecture impossible : {e}")
            continue

        if not contenu.strip():
            print("   ⚠️ Chapitre vide.")
            continue

        if not item.get("titre"):
            item["titre"] = (
                nettoyer_titre_chapitre(extraire_titre_markdown(contenu))
                or nettoyer_nom_fichier(fichier)
            )

        traduction, mode = traduire_chapitre(item["titre"], contenu)

        if not traduction:
            print(f"   ❌ Traduction impossible ({mode}).")
            continue

        print(f"   \U0001F9E0 Mode : {mode}")

        titre_traduit = extraire_titre_traduit(traduction)

        if not titre_traduit:
            titre_traduit = traduire_titre_chapitre(
                item.get("titre") or nettoyer_nom_fichier(fichier)
            )

        if not titre_traduit:
            titre_traduit = item.get("titre") or "Chapter"

        base_nom = nom_fichier_depuis_titre(titre_traduit)
        nom = base_nom
        compteur = 2

        while nom.lower() in noms_utilises:
            nom = f"{base_nom}_{compteur}"
            compteur += 1

        noms_utilises.add(nom.lower())

        chemin_sortie = os.path.join(DOSSIER_TRADUCTIONS, nom + ".md")

        try:
            with open(chemin_sortie, "w", encoding="utf-8", newline="\n") as f:
                f.write(traduction.strip() + "\n")

            print(f"   ✅ Fichier créé : {chemin_sortie}")
            succes += 1

        except Exception as e:
            print(f"   ❌ Écriture impossible : {e}")

        if PAUSE_CHAPITRES > 0:
            time.sleep(PAUSE_CHAPITRES)

    print(f"\U0001FA99 Tokens utilisés : {TOKENS_SESSION}")

    if succes == 0:
        print("❌ Aucun chapitre traduit.")
        return 1

    print(f"✅ Traduction terminée : {succes} fichier(s) créé(s) dans {DOSSIER_TRADUCTIONS}")
    return 0


# ─────────────────────────────────────────
# Écriture et Main
# ─────────────────────────────────────────
def ecrire_resume(groupes):
    lignes = ["# Résumés du roman", "", f"_Généré le {datetime.now():%d/%m/%Y à %H:%M}_", ""]

    for acte, items in groupes:
        lignes.append(f"## Acte {acte}")
        lignes.append("")

        for item in items:
            chap = item.get("chapitre") or "?"
            titre = item.get("titre") or "Sans titre"
            resume = item.get("resume", "").strip() or "Résumé non disponible."

            lignes.extend([f"### Chapitre {chap} — {titre}", "", resume, ""])

    with open(CHEMIN_RESUME, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lignes))


def main():
    print("\U0001F916 IA_Roman : génération des résumés")
    print(f"Dossier chapitres : {DOSSIER_CHAPITRES}")

    charger_parametres_ia()
    print(f"Mode IA : {MODE_IA}")

    if not os.path.isdir(DOSSIER_CHAPITRES):
        print("❌ Dossier Chapitres introuvable.")
        return 1

    plan = construire_plan()
    if not plan:
        print("❌ Aucun chapitre .md détecté dans le dossier Chapitres.")
        return 1

    print(f"\U0001F4DA Chapitres détectés : {len(plan)}")

    if MODE_IA == "offline":
        print("\U0001F9E0 Mode forcé : résumé automatique hors IA.")
    else:
        print(f"\U0001F9E0 Ollama : {OLLAMA_URL} / modèle {OLLAMA_MODEL}")
        print(f"\U0001F511 Clé API Ollama : {'Détectée' if OLLAMA_API_KEY else 'Non détectée'}")
        print(f"\U0001F511 Clé OpenAI : {'Détectée' if obtenir_cle_openai() else 'Non détectée'}")

    for item in plan:
        fichier = item.get("fichier", "")
        print(f"\U0001F4D6 Lecture : {fichier}")

        try:
            with open(item["chemin"], encoding="utf-8", errors="ignore") as f:
                contenu = f.read()
        except Exception as e:
            item["resume"], item["mode"] = f"Lecture impossible : {e}", "erreur"
            continue

        if not item.get("titre"):
            item["titre"] = nettoyer_titre_chapitre(extraire_titre_markdown(contenu)) or nettoyer_nom_fichier(fichier)

        resume, mode = resumer_chapitre(item["titre"], contenu)
        item["resume"], item["mode"] = resume, mode

        print(f"   \U0001F9E0 Mode : {mode}")

        if PAUSE_CHAPITRES > 0:
            time.sleep(PAUSE_CHAPITRES)

    try:
        ecrire_resume(grouper_par_acte(plan))
    except Exception as e:
        print(f"❌ Impossible d'écrire Résumé.md : {e}")
        return 1

    modes = sorted({item.get("mode", "inconnu") for item in plan})
    print(f"\U0001F9E0 Mode(s) utilisé(s) : {', '.join(modes)}")
    print(f"✅ Fichier créé : {CHEMIN_RESUME}")
    print(f"\U0001FA99 Tokens utilisés : {TOKENS_SESSION}")

    return 0


def cli_set_key(fournisseur):
    import getpass

    print(f"Enregistrement de la clé API {fournisseur} dans Configuration_roman.xlsx (onglet IA).")

    try:
        cle = getpass.getpass(f"Clé API {fournisseur} (masquée) : ").strip()
    except (KeyboardInterrupt, EOFError):
        print("\nOpération annulée.")
        return 1

    if not cle:
        print("❌ Aucune clé saisie.")
        return 1

    if enregistrer_parametre_ia(f"Clé API {fournisseur}", cle):
        print(f"✅ Clé API {fournisseur} enregistrée.")
        return 0

    return 1


if __name__ == "__main__":
    if len(sys.argv) >= 2:
        if sys.argv[1] == "--set-openai-key":
            IA_ROMAN_CODE = cli_set_key("OpenAI")

        elif sys.argv[1] == "--set-ollama-key":
            IA_ROMAN_CODE = cli_set_key("Ollama")

        elif sys.argv[1] == "--init-ia-sheet":
            charger_parametres_ia()
            print("✅ Onglet IA initialisé ou vérifié dans Configuration_roman.xlsx.")
            IA_ROMAN_CODE = 0

        elif sys.argv[1] == "--traduire":
            IA_TRADUCTION_CODE = main_traduction()

        elif sys.argv[1] == "--traduire-chapitre" and len(sys.argv) >= 3:
            IA_TRADUCTION_CODE = main_traduction(sys.argv[2:])

        else:
            IA_ROMAN_CODE = main()
    else:
        IA_ROMAN_CODE = main()

    code = (IA_TRADUCTION_CODE
            if sys.argv[1:2] in (['--traduire'], ['--traduire-chapitre'])
            else IA_ROMAN_CODE)
    raise SystemExit(code)