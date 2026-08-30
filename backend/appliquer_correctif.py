#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""appliquer_correctif.py – FICHIER UNIQUE de correctifs (idempotent).
Règles centralisées + correctifs Word / PDF / EPUB / Flutter + REGISTRE v4.
Usage : python appliquer_correctif.py"""
import os, re, json, shutil, ast

BASE  = os.path.dirname(os.path.abspath(__file__))
ROMAN = os.path.join(BASE, 'generer_roman.py')
PDF   = os.path.join(BASE, 'generer_pdf_direct.py')
EBOOK = os.path.join(BASE, 'generer_ebook.py')
DART  = os.path.join(BASE, '..', 'lib', 'ui', 'home_screen.dart')
JSONR = os.path.join(BASE, 'regles_mise_en_page.json')
PYR   = os.path.join(BASE, 'regles.py')


def bak(p):
    if os.path.isfile(p): shutil.copyfile(p, p + '.bak')


def entre(src, a, b, new):
    if a not in src or b not in src: return src, False
    i = src.index(a); j = src.index(b)
    if j <= i: return src, False
    return src[:i] + new + src[j:], True


def suppr_fonction_dart(src, sig):
    i = src.find(sig)
    if i == -1: return src, False
    j = src.find('{', i)
    if j == -1: return src, False
    d = 0; k = j
    while k < len(src):
        if src[k] == '{': d += 1
        elif src[k] == '}':
            d -= 1
            if d == 0: return src[:i] + src[k+1:], True
        k += 1
    return src, False


def suppr_methode_dart(src, nom):
    """Supprime une méthode/getter/champ Dart (corps {…} ou => …;)."""
    m = re.search(r'^[ \t]*(?:String\s+)?(?:get\s+)?' + re.escape(nom) + r'\s*(?:\([^)]*\))?', src, re.M)
    if not m: return src, False
    i = m.start(); k = m.end()
    while k < len(src) and src[k] in ' \t': k += 1
    if k < len(src) and src[k] == '=':
        e = src.find(';', k)
        if e == -1: return src, False
        return src[:i] + src[e+1:], True
    if k < len(src) and src[k] == '{':
        d = 0; e = k
        while e < len(src):
            if src[e] == '{': d += 1
            elif src[e] == '}':
                d -= 1
                if d == 0: return src[:i] + src[e+1:], True
            e += 1
    return src, False


def inserer_avant_lire_doc(s, bloc):
    for a in ('Future<void> _lireDocument(', 'Future _lireDocument(', 'void _lireDocument('):
        if a in s: return s.replace(a, bloc + a, 1), True
    return s, False


# ═══════════════ 1. SOURCE DE VÉRITÉ ═══════════════
REGLES = {
  "version": "1.14",
  "formats_livre": [["5 x 8 po",12.7,20.32],["5.06 x 7.81 po",12.85,19.84],["5.25 x 8 po",13.34,20.32],
    ["5.5 x 8.5 po",13.97,21.59],["6 x 9 po",15.24,22.86],["6.14 x 9.21 po",15.6,23.4],
    ["7 x 10 po",17.78,25.4],["8 x 10 po",20.32,25.4],["8.5 x 8.5 po",21.59,21.59],
    ["8.5 x 11 po",21.59,27.94],["A4",21.0,29.7]],
  "marges_kdp": {"bareme_po": [[150,0.375],[300,0.5],[500,0.625],[700,0.75],[828,0.875]],
                 "garde_po": 0.125, "symetriques": True},
  "typographie_defaut": {"police_corps":"Aptos","police_titres":"Cinzel","taille_corps_pt":11,
                         "interligne":1.0,"retrait_premiere_ligne_cm":0.5},
  "mise_en_page": {"centrage_horizontal_images": True, "marges_haut_bas_cm": 1.9},
  "corrections_communes": [["⁠",""],["﻿",""],["--- ","— "],[" …","…"],["'","’"],
    ["pu is","puis"],[". . ",". "],[".À",". À"],["cératopsiens ,","cératopsiens,"],
    ["de s'approchaient","de s'approcher"],["ils émettait","ils émettaient"]],
  "corrections_par_chapitre": {
    "1.1": [["E t si","Et si"],["XIIe siècle","XIIᵉ siècle"]],
    "2.1": [["une par inhérente","une part inhérente"],["me pousser à","me pousse à"],
      ["l’Étranger. il observe","l’Étranger. Il observe"],["L’ombre, m’observe.","L’ombre m’observe."],
      ["pas d’atmosphère se dissipe","pas d’atmosphère qui se dissipe"]],
    "2.2": [["Ute tension insupportable","Une tension insupportable"],["Ute chance de naître","Une chance de naître"],
      ["Mon regard se live","Mon regard se lève"],["ma traversé ,","ma traversée,"],["ma traversé,","ma traversée,"],
      ["La Lumiere éclate","La lumière éclate"],["Je le survole.","Je la survole."],
      ["s’harmonise, créé une","s’harmonise, crée une"],["fait créé une","fait crée une"],["Nunael","Nunaël"],
      ["me défier ainsi ?","me défier ainsi ? »"],["Quand il sera trop tard ?","Quand il sera trop tard ? »"],
      ["si la fin est toujours la même ?","si la fin est toujours la même ? »"],
      ["« Mais ils auront existé.","« Mais ils auront existé. »"],["je combats !","je combats ! »"],
      ["même si je ne peux pas l’anéantir.","même si je ne peux pas l’anéantir. »"],["a créé …. Et","a créé… Et"]]},
  "regex_pensees": [["\"([^\"]+)\"","« \\1 »"]]
}

MODULE = '''"""regles.py – règles partagées (source : regles_mise_en_page.json)."""
import os, re, json
BASE = os.path.dirname(os.path.abspath(__file__))
try:
    with open(os.path.join(BASE, 'regles_mise_en_page.json'), encoding='utf-8') as f:
        REGLES = json.load(f)
except Exception:
    REGLES = {}
FORMATS_LIVRE = [(n, float(w), float(h)) for n, w, h in REGLES.get('formats_livre', [['6 x 9 po', 15.24, 22.86]])]
def dimensions_pour_format(label):
    s = str(label).lower().strip().replace(',', '.')
    for nom, w, h in FORMATS_LIVRE:
        if nom.lower() in s or s == nom.lower().replace(' po', ''):
            return w, h
    if 'a4' in s:
        return 21.0, 29.7
    m = re.search(r'(\\d+(?:\\.\\d+)?)\\s*x\\s*(\\d+(?:\\.\\d+)?)', s)
    if m:
        w, h = float(m.group(1)), float(m.group(2))
        if 'mm' in s: return w / 10.0, h / 10.0
        if 'cm' in s: return w, h
        return w * 2.54, h * 2.54
    return 17.78, 25.4
_M = REGLES.get('marges_kdp', {})
BAREME_KDP = [(int(m), float(g)) for m, g in _M.get('bareme_po', [(828, 0.875)])]
GARDE = float(_M.get('garde_po', 0.125))
MARGES_SYMETRIQUES = bool(_M.get('symetriques', True))
def gouttiere_kdp_pour(p):
    for maxi, g in BAREME_KDP:
        if p <= maxi: return g
    return BAREME_KDP[-1][1]
def marge_kdp_pour(p):
    return gouttiere_kdp_pour(p) + GARDE
CORRECTIONS_COMMUNES = [(a, b) for a, b in REGLES.get('corrections_communes', [])]
CORR_PAR_PREFIXE = {k: [(a, b) for a, b in v] for k, v in REGLES.get('corrections_par_chapitre', {}).items()}
REGEX_PENSEES = [(re.compile(p), r) for p, r in REGLES.get('regex_pensees', [])]
def corrections_pour(f):
    for pref, corr in CORR_PAR_PREFIXE.items():
        if str(f).startswith(pref): return corr
    return []
TYPO = REGLES.get('typographie_defaut', {})
MISE_EN_PAGE = REGLES.get('mise_en_page', {})
def __getattr__(name):
    _defauts = {'MARGES_SYMETRIQUES': True, 'GARDE': 0.125, 'TYPO': {}, 'MISE_EN_PAGE': {},
                'BAREME_KDP': [(828, 0.875)], 'CORRECTIONS_COMMUNES': [], 'CORR_PAR_PREFIXE': {},
                'REGEX_PENSEES': [], 'FORMATS_LIVRE': [('6 x 9 po', 15.24, 22.86)]}
    if name in _defauts:
        return _defauts[name]
    raise AttributeError("module 'regles' has no attribute " + repr(name))
'''

LIRE_INFOS_PDF = '''def lire_infos():
    infos = {}
    j = _lire_json()
    for cle, val in ((j or {}).get('informations', {}) or {}).items():
        if val: infos[cle.lower()] = nettoyer(val)
    if not infos:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(CHEMIN_CONFIG, data_only=True)
            ws = wb['Informations']
            for row in ws.iter_rows(values_only=True):
                if row and row[0]: infos[nettoyer(row[0]).lower()] = nettoyer(row[1])
        except Exception:
            pass
    return infos

'''

LIM_NOUVEAU = """_tit = INFOS.get('titre complet du roman', '') or 'Titre'
    _aut = INFOS.get("nom de l'auteur (couverture)", '') or 'Auteur'
    _ann = INFOS.get('année de publication', '2026')
    lim = [Paragraph(escape(_tit.upper()), st_acte),
    Paragraph(escape(INFOS.get('sous-titre éventuel', '') or ''), st_ch2),
    Spacer(1, 2 * cm), Paragraph(escape(_aut), st_ch2), PageBreak()]
    _larg_utile = STYLE['largeur_cm'] * cm - 4 * cm
    _haut_utile = H - 3.8 * cm
    # Page 2 : copyright – texte en BAS de page
    _cop = []
    _cop.append(Paragraph(escape((_tit + ' – ' + INFOS['sous-titre éventuel']) if INFOS.get('sous-titre éventuel') else _tit), st_lim))
    _cop.append(Spacer(1, 0.6 * cm))
    _cop.append(Paragraph(escape(INFOS.get('mention de copyright', '') or ('© ' + _ann + ' ' + _aut + '. Tous droits réservés.')), st_lim))
    _cop.append(Spacer(1, 0.3 * cm))
    _meta = ' – '.join(x for x in [INFOS.get('édition', ''), _ann] if x)
    if _meta: _cop.append(Paragraph(escape(_meta), st_lim))
    if INFOS.get('isbn'): _cop.append(Paragraph('ISBN : ' + escape(INFOS['isbn']), st_lim))
    if INFOS.get('dépôt légal'): _cop.append(Paragraph('Dépôt légal : ' + escape(INFOS['dépôt légal']), st_lim))
    if INFOS.get("maison d'édition"):
        _cop.append(Spacer(1, 0.3 * cm)); _cop.append(Paragraph(escape(INFOS["maison d'édition"]), st_lim))
    _cop.append(Spacer(1, 0.6 * cm))
    for _t in ['Toute reproduction, même partielle, est interdite sans l’autorisation',
    'préalable de l’auteur, conformément aux dispositions de la législation',
    'en vigueur sur la propriété intellectuelle.']:
        _cop.append(Paragraph(escape(_t), st_lim))
    if INFOS.get('site web'):
        _cop.append(Spacer(1, 0.3 * cm)); _cop.append(Paragraph(escape(INFOS['site web']), st_lim))
    _h_cop = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _cop)
    lim.append(Spacer(1, max(0, _haut_utile - _h_cop)))
    lim += _cop
    # Page 3 : dédicace – centrée au milieu de la page
    if INFOS.get('dédicace'):
        _ded = [Paragraph(escape(x), st_ch2) for x in INFOS['dédicace'].splitlines() if x.strip()]
        _h_d = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _ded)
        lim.append(PageBreak())
        lim.append(Spacer(1, max(0, (_haut_utile - _h_d) / 2)))
        lim += _ded
    # Page 4 : épigraphe – centrée au milieu de la page
    if INFOS.get('épigraphe'):
        _epi = [Paragraph(escape(x), st_ch2) for x in INFOS['épigraphe'].splitlines() if x.strip()]
        _h_e = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _epi)
        lim.append(PageBreak())
        lim.append(Spacer(1, max(0, (_haut_utile - _h_e) / 2)))
        lim += _epi
    """

BLOC_WORD = '''def _convert_word_pdf(docx_path, pdf_path):
    import win32com.client
    word = win32com.client.Dispatch('Word.Application'); word.Visible = False
    try: word.DisplayAlerts = 0; word.AutomationSecurity = 3
    except Exception: pass
    d = word.Documents.Open(os.path.abspath(docx_path), False, True, False)
    try: d.ExportAsFixedFormat(os.path.abspath(pdf_path), 17, False, 0)
    finally: d.Close(False); word.Quit()

def generer_depuis_word():
    import glob, subprocess, time
    c = glob.glob(os.path.join(BASE, '*_KDP.docx')) + glob.glob(os.path.join(BASE, 'export', '*_KDP.docx'))
    if not c: return False
    dx = max(c, key=os.path.getmtime); pf = os.path.splitext(dx)[0] + '.pdf'
    for _ in range(5):
        try:
            if os.path.exists(pf): os.remove(pf)
            break
        except Exception: time.sleep(0.3)
    print('   🖨 Conversion Word→PDF (vos modifs incluses)…')
    try:
        r = subprocess.run([sys.executable, os.path.abspath(__file__), '--convert', dx, pf],
                           timeout=180, capture_output=True, text=True)
        if r.returncode == 0 and os.path.isfile(pf):
            print('✅ PDF KDP prêt (converti depuis Word) : ' + pf); return True
        return False
    except subprocess.TimeoutExpired:
        print('   ⚠️ Word trop lent (>180 s) → génération directe.'); return False

'''

REGISTRE_BLOC = """    # ── CORRECTIF registre : régénère registre.json pour l'interface ──
    try:
        import json as _json_reg
        _r = lire_infos()
        _inf = _r[0] if isinstance(_r, tuple) else _r
        _sty = lire_style()
        _org = lire_organisation()
        _mots = _nchap = _nill = 0
        for _b in (_org or []):
            if str(_b.get('type', '')) == 'chapitre':
                _nchap += 1
                try:
                    _txt = open(os.path.join(BASE, 'Chapitres', str(_b.get('fichier_source', '')) + '.md'), encoding='utf-8').read()
                    _mots += len(_txt.split())
                    _nill += _txt.count('![')
                except Exception:
                    pass
            elif str(_b.get('type', '')) == 'image':
                _nill += 1
        _reg = {
            'ouvrage': (_inf.get('titre complet du roman', '') or titre_aff),
            'format': (_sty.get('format_livre') or _sty.get('format') or ''),
            'mots': _mots,
            'pages': int(pages_reelles or 0),
            'chapitres': _nchap,
            'illustrations': _nill,
            'source': 'JSON',
        }
        for _chem_reg in (os.path.join(BASE, 'registre.json'),
                          os.path.join(BASE, 'statistiques.json'),
                          os.path.join(BASE, 'export', 'registre.json'),
                          os.path.join(BASE, 'export', 'statistiques.json')):
            try:
                os.makedirs(os.path.dirname(_chem_reg), exist_ok=True)
                with open(_chem_reg, 'w', encoding='utf-8') as _f:
                    _json_reg.dump(_reg, _f, ensure_ascii=False, indent=2)
            except Exception:
                pass
    except Exception:
        pass
"""

REGISTRE_V2 = '''
# ── CORRECTIF registre v2 : clés/valeurs propres pour l'interface ──
def _normaliser_registre():
    import json as _j
    for _p in (os.path.join(BASE, 'registre.json'), os.path.join(BASE, 'export', 'registre.json')):
        try:
            if not os.path.isfile(_p):
                continue
            _d = _j.load(open(_p, encoding='utf-8'))
            _c = {str(k).strip(): (str(v).strip() if isinstance(v, str) else v) for k, v in _d.items()}
            if not _c.get('mots'):
                _tot = 0
                _dch = os.path.join(BASE, 'Chapitres')
                if os.path.isdir(_dch):
                    for _f in os.listdir(_dch):
                        if _f.lower().endswith('.md'):
                            try:
                                _tot += len(open(os.path.join(_dch, _f), encoding='utf-8').read().split())
                            except Exception:
                                pass
                _c['mots'] = _tot
            _j.dump(_c, open(_p, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
        except Exception:
            pass
import atexit as _atx_reg
_atx_reg.register(_normaliser_registre)
'''

FN_RESUME = '''  Future<void> _lireResume() async {
    try {
      Directory dir = Directory.current;
      Directory? backend;
      for (int i = 0; i < 6; i++) {
        final cand = Directory(path.join(dir.path, 'backend'));
        if (cand.existsSync() && File(path.join(cand.path, 'generer_roman.py')).existsSync()) { backend = cand; break; }
        final parent = dir.parent;
        if (parent.path == dir.path) break;
        dir = parent;
      }
      if (backend == null) { _afficherMsg('Dossier backend introuvable'); return; }
      final resumePath = path.join(backend.path, 'Résumé.md');
      if (!await File(resumePath).exists()) {
        _afficherMsg('« Résumé.md » introuvable. Générez-le via « Résumés IA » (menu Production).');
        return;
      }
      if (Platform.isWindows) { await Process.run('cmd', ['/c', 'start', '', resumePath]); }
      else if (Platform.isMacOS) { await Process.run('open', [resumePath]); }
      else { await Process.run('xdg-open', [resumePath]); }
    } catch (e) { _afficherMsg('Erreur ouverture résumé : ' + e.toString()); }
  }

'''

FN_AFFICHER = '''  void _afficherMsg(String msg) {
    if (!mounted) return;
    ScaffoldMessenger.of(context).showSnackBar(SnackBar(content: Text(msg)));
  }

'''

DART_REGISTRE = '''  String _valeurRegistre(String cle) {
    Directory? backend;
    Directory dir = Directory.current;
    for (int i = 0; i < 6; i++) {
      final cand = Directory(path.join(dir.path, 'backend'));
      if (cand.existsSync() && File(path.join(cand.path, 'generer_roman.py')).existsSync()) { backend = cand; break; }
      final parent = dir.parent;
      if (parent.path == dir.path) break;
      dir = parent;
    }
    final cands = <String>[];
    if (backend != null) {
      cands.add(path.join(backend.path, 'registre.json'));
      cands.add(path.join(backend.path, 'export', 'registre.json'));
      cands.add(path.join(backend.path, 'statistiques.json'));
      cands.add(path.join(backend.path, 'export', 'statistiques.json'));
    }
    cands.add('registre.json');
    for (final c in cands) {
      try {
        final f = File(c);
        if (!f.existsSync()) continue;
        final d = jsonDecode(f.readAsStringSync());
        if (d is Map) {
          for (final k in d.keys) {
            if (k.toString().trim() == cle) {
              final v = d[k].toString().trim();
              if (v.isNotEmpty) return v;
            }
          }
        }
      } catch (_) {}
    }
    return '—';
  }

'''


def ecrire_partages():
    with open(JSONR, 'w', encoding='utf-8') as f:
        json.dump(REGLES, f, ensure_ascii=False, indent=2)
    with open(PYR, 'w', encoding='utf-8') as f:
        f.write(MODULE)
    print('   ✅ regles_mise_en_page.json + regles.py écrits')


def normaliser_registres():
    """Nettoie immédiatement les registre.json existants (clés/valeurs)."""
    _touche = False
    for _p in (os.path.join(BASE, 'registre.json'), os.path.join(BASE, 'export', 'registre.json')):
        try:
            if not os.path.isfile(_p): continue
            _d = json.load(open(_p, encoding='utf-8'))
            _c = {str(k).strip(): (str(v).strip() if isinstance(v, str) else v) for k, v in _d.items()}
            if not _c.get('mots'):
                _tot = 0
                _dch = os.path.join(BASE, 'Chapitres')
                if os.path.isdir(_dch):
                    for _f in os.listdir(_dch):
                        if _f.lower().endswith('.md'):
                            try: _tot += len(open(os.path.join(_dch, _f), encoding='utf-8').read().split())
                            except Exception: pass
                _c['mots'] = _tot
            json.dump(_c, open(_p, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
            _touche = True
        except Exception:
            pass
    if _touche:
        print("   ✅ registre.json normalisé (clés propres pour l'interface)")


# ═══════════════ 2. generer_roman.py ═══════════════
def patcher_roman():
    if not os.path.isfile(ROMAN): return
    s = open(ROMAN, encoding='utf-8').read(); mod = False
    if 'import regles as _R' not in s:
        s = s.replace('from docx.oxml.ns import qn, nsdecls',
                      'from docx.oxml.ns import qn, nsdecls\nimport regles as _R', 1); mod = True
    if 'def _lire_json_brut' not in s:
        s = s.replace('def lire_infos():',
            "def _lire_json_brut():\n    try:\n        import json\n"
            "        return json.load(open(CHEMIN_CONFIG_JSON, encoding='utf-8'))\n"
            "    except Exception:\n        return None\n\ndef lire_infos():", 1)
        A = "def lire_infos():\n    infos = {l: '' for l in CHAMPS_LABELS}\n    source = 'défaut'\n"
        if A in s:
            s = s.replace(A, A +
                "    ji = (_lire_json_brut() or {}).get('informations', {}) or {}\n"
                "    for label in CHAMPS_LABELS:\n        for k, v in ji.items():\n"
                "            if v and nettoyer(k).lower().startswith(CORRESPONDANCES[label]):\n"
                "                infos[label] = nettoyer(v); source = 'JSON'; break\n", 1)
        mod = True
    if 'if nom.lower() in s or s == nom' not in s:
        s, r = entre(s, 'def dimensions_pour_format(label):', 'def verifier_formats():',
            "def dimensions_pour_format(label):\n"
            "    s = str(label).lower().strip().replace(',', '.')\n"
            "    for nom, w, h in FORMATS_LIVRE:\n"
            "        if nom.lower() in s or s == nom.lower().replace(' po', ''):\n"
            "            return w, h\n"
            "    if 'a4' in s:\n        return 21.0, 29.7\n"
            "    m = re.search(r'(\\d+(?:\\.\\d+)?)\\s*x\\s*(\\d+(?:\\.\\d+)?)', s)\n"
            "    if m:\n        w, h = float(m.group(1)), float(m.group(2))\n"
            "        if 'mm' in s: return w / 10.0, h / 10.0\n"
            "        if 'cm' in s: return w, h\n"
            "        return w * 2.54, h * 2.54\n"
            "    return 17.78, 25.4\n\n")
        mod = mod or r
    if 'pages = d.ComputeStatistics(2)' not in s:
        s = s.replace("        d.Save()\n        d.Close(False)\n        d = None\n"
            "        print('   ✅ TDM et champs mis à jour dans le fichier final.')",
            "        d.Save()\n        pages = d.ComputeStatistics(2)\n        d.Close(False)\n        d = None\n"
            "        print('   ✅ TDM et champs mis à jour dans le fichier final.')\n        return pages", 1)
        mod = True
    if 'pages_reelles = maj_champs_word' not in s:
        s, r = entre(s, '    for _passe in range(2):', '    afficher_statistiques(pages_reelles, titre_aff)',
            "    print('   📑 Comptage des pages réelles via Word…')\n"
            "    pages_reelles = maj_champs_word(nom_fichier)\n"
            "    if not pages_reelles:\n        pages_reelles = compter_pages_reelles(nom_fichier)\n"
            "    if pages_reelles:\n        m_cible = marge_kdp_pour(pages_reelles)\n"
            "        if abs(m_cible - gutter) >= 0.001:\n            gutter = m_cible\n"
            "            for sec in doc.sections:\n                sec.left_margin = Inches(gutter)\n"
            "                sec.right_margin = Inches(gutter)\n"
            "            nom_fichier = enregistrer_docx_securise(doc, nom_fichier)\n"
            "            pages_reelles = maj_champs_word(nom_fichier) or pages_reelles\n")
        mod = mod or r
    if "max_l = max(6.0, min(14.0, STYLE['largeur_cm'] - 4.0))" not in s:
        s = s.replace('max_l = largeur_illustration()',
                      "max_l = max(6.0, min(14.0, STYLE['largeur_cm'] - 4.0))", 1); mod = True
    if 'CORRECTIF registre' not in s:
        if '    afficher_statistiques(pages_reelles, titre_aff)' in s:
            s = s.replace('    afficher_statistiques(pages_reelles, titre_aff)',
                          '    afficher_statistiques(pages_reelles, titre_aff)\n' + REGISTRE_BLOC, 1)
            mod = True
            print("   • registre.json rétabli (statistiques pour l'interface)")
    if 'CORRECTIF registre v2' not in s:
        s += REGISTRE_V2
        mod = True
        print("   • Normalisation registre v2 activée (atexit)")
    if mod:
        bak(ROMAN); open(ROMAN, 'w', encoding='utf-8').write(s)
        print('   ✅ generer_roman.py patché')
    else:
        print('   ℹ️ generer_roman.py déjà à jour')


# ═══════════════ 3. generer_pdf_direct.py ═══════════════
def patcher_pdf():
    if not os.path.isfile(PDF): return
    s = open(PDF, encoding='utf-8').read(); mod = False
    if 'import regles as _R' not in s:
        s = s.replace('from pypdf import PdfReader, PdfWriter',
                      'from pypdf import PdfReader, PdfWriter\nimport regles as _R', 1); mod = True
    if 'if not infos:' not in s:
        s, r = entre(s, 'def lire_infos():', 'def lire_organisation():', LIRE_INFOS_PDF); mod = mod or r
    if '_R.corrections_pour(fichier)' not in s:
        s = s.replace('for a, b in CORRECTIONS_COMMUNES: texte = texte.replace(a, b)',
                      'for a, b in CORRECTIONS_COMMUNES: texte = texte.replace(a, b)\n'
                      '    for a, b in _R.corrections_pour(fichier): texte = texte.replace(a, b)', 1); mod = True
    if "img.hAlign = 'CENTER'" not in s:
        s = s.replace('st.append(RLImage(ch_hd, width=w_cm * cm, height=h_cm * cm))',
                      "img = RLImage(ch_hd, width=w_cm * cm, height=h_cm * cm); img.hAlign = 'CENTER'; st.append(img)", 1)
        mod = True
    if 'm_sym = gut + 0.125' not in s:
        s = s.replace('MARGES = (gut * 72, out * 72, 1.9 * cm, 1.9 * cm)',
                      'm_sym = gut + 0.125\n    MARGES = (m_sym * 72, m_sym * 72, 1.9 * cm, 1.9 * cm)', 1); mod = True
    if 'p if p < (debut_num or 1)' not in s:
        s = s.replace('{p - debut_num + 1}', '{p if p < (debut_num or 1) else p - (debut_num or 1) + 1}')
        s = s.replace('{p - (debut_num or 1) + 1}', '{p if p < (debut_num or 1) else p - (debut_num or 1) + 1}')
        mod = True
    if '_cop_kt.wrap(' in s:
        s = s.replace("    from reportlab.platypus import KeepTogether as _KT\n", "")
        s = s.replace("    _cop_kt = _KT(_cop)\n    _w_cop, _h_cop = _cop_kt.wrap(_larg_utile, _haut_utile)",
                      "    _h_cop = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _cop)")
        s = s.replace("        _ded_kt = _KT(_ded)\n        _w_d, _h_d = _ded_kt.wrap(_larg_utile, _haut_utile)",
                      "        _h_d = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _ded)")
        s = s.replace("        _epi_kt = _KT(_epi)\n        _w_e, _h_e = _epi_kt.wrap(_larg_utile, _haut_utile)",
                      "        _h_e = sum(_f.wrap(_larg_utile, _haut_utile)[1] for _f in _epi)")
        mod = True
        print('   • Mesure des blocs liminaires réparée (sans KeepTogether.wrap)')
    if 'texte en BAS de page' not in s:
        _m = re.search(r"lim = \[Paragraph\(escape\(\(INFOS\.get\('titre complet du roman'", s)
        if not _m:
            _m = re.search(r"_tit = INFOS\.get\('titre complet du roman'", s)
        _e = re.search(r"segments\.append\(\{'type': 'lim'", s)
        if _m and _e and _e.start() > _m.start():
            s = s[:_m.start()] + LIM_NOUVEAU + s[_e.start():]
            mod = True
            print('   • Pages liminaires : copyright en bas + dédicace/épigraphe centrées')
    _lignes = s.split('\n')
    _net = [_l for _l in _lignes if '# CORRECTIF: espace titre-chapitre' not in _l]
    if len(_net) != len(_lignes):
        s = '\n'.join(_net); mod = True
        print('   • Ligne défectueuse (espace titre) supprimée')
    if 'spaceAfter=1.2 * cm' not in s and '# espace-titre-v2' not in s:
        _lignes = s.split('\n'); _touche = False
        for _i, _l in enumerate(_lignes):
            if 'ParagraphStyle(' in _l and 'TA_CENTER' in _l and 'spaceAfter=' in _l:
                _m = re.search(r"spaceAfter=([^,)]+)", _l)
                if _m and '(' not in _m.group(1):
                    _lignes[_i] = _l[:_m.start()] + 'spaceAfter=1.2 * cm' + _l[_m.end():]
                    _touche = True
        if _touche:
            s = '\n'.join(_lignes); mod = True
            print('   • Espace titre→texte augmenté (spaceAfter = 1,2 cm)')
        else:
            _idx = None
            for _i, _l in enumerate(_lignes):
                if 'st.append(Paragraph(' in _l and 'chapitre_ligne' in _l:
                    _idx = _i
            if _idx is not None:
                _ind = None
                for _j in range(_idx + 1, min(_idx + 6, len(_lignes))):
                    if _lignes[_j].strip():
                        _ls = _lignes[_j]
                        _ind = _ls[:len(_ls) - len(_ls.lstrip())]
                        break
                if _ind is not None:
                    _lignes.insert(_idx + 1, _ind + "st.append(Spacer(1, 0.8 * cm))  # espace-titre-v2")
                    s = '\n'.join(_lignes); mod = True
                    print('   • Espace titre→texte augmenté (Spacer après titre)')
    if 'def generer_depuis_word' not in s:
        s = s.replace('def main():', BLOC_WORD +
            "def main():\n    if '--direct' not in sys.argv[1:] and generer_depuis_word():\n        return\n", 1)
        mod = True
    if "'--convert' in sys.argv" not in s:
        s = s.replace("if __name__ == '__main__':\n    main()",
            "if __name__ == '__main__':\n    if '--convert' in sys.argv[1:]:\n"
            "        _i = sys.argv.index('--convert')\n"
            "        _convert_word_pdf(sys.argv[_i + 1], sys.argv[_i + 2])\n"
            "    else:\n        main()", 1)
        mod = True
    if mod:
        bak(PDF); open(PDF, 'w', encoding='utf-8').write(s)
        print('   ✅ generer_pdf_direct.py patché')
    else:
        print('   ℹ️ generer_pdf_direct.py déjà à jour')


# ═══════════════ 4. generer_ebook.py ═══════════════
def patcher_ebook():
    if not os.path.isfile(EBOOK): return
    s = open(EBOOK, encoding='utf-8').read()
    _s0 = s
    if "'export', '*_KDP.docx'" not in s:
        s = s.replace(
            "cands = glob.glob(os.path.join(BASE, '*_KDP.docx'))",
            "cands = (glob.glob(os.path.join(BASE, '*_KDP.docx'))\n"
            "             + glob.glob(os.path.join(BASE, 'export', '*_KDP.docx')))", 1)
    if "dossier_sortie = os.path.join(BASE, 'export')" not in s:
        s = s.replace(
            "sortie = os.path.join(BASE, f'{slug_titre}_ebook.epub')",
            "dossier_sortie = os.path.join(BASE, 'export')\n"
            "    os.makedirs(dossier_sortie, exist_ok=True)\n"
            "    sortie = os.path.join(dossier_sortie, f'{slug_titre}_ebook.epub')", 1)
    if 'def _lire_json_brut' not in s:
        s = s.replace(
            'def lire_infos():',
            "def _lire_json_brut():\n"
            "    try:\n"
            "        import json\n"
            "        return json.load(open(os.path.join(BASE, 'Configuration_roman.json'), encoding='utf-8'))\n"
            "    except Exception:\n"
            "        return None\n\n"
            "def lire_infos():", 1)
    if "ji = (_lire_json_brut() or {}).get('informations'" not in s:
        _A = "def lire_infos():\n    infos = {l: '' for l in CHAMPS_LABELS}\n    source = 'défaut'\n"
        if _A in s:
            s = s.replace(_A, _A +
                "    ji = (_lire_json_brut() or {}).get('informations', {}) or {}\n"
                "    for label in CHAMPS_LABELS:\n"
                "        for k, v in ji.items():\n"
                "            kl = str(k).strip().lower()\n"
                "            if v and any(kl.startswith(p) for p in CORRESPONDANCES[label]):\n"
                "                infos[label] = str(v).strip(); source = 'JSON'; break\n", 1)
    if 'if v and not infos[label]:' not in s:
        s = s.replace(
            "                if v:\n                    infos[label] = str(v).strip()",
            "                if v and not infos[label]:\n                    infos[label] = str(v).strip()", 1)
    if s != _s0:
        bak(EBOOK); open(EBOOK, 'w', encoding='utf-8').write(s)
        print('   ✅ generer_ebook.py patché (export + infos du menu Informations)')
    else:
        print('   ℹ️ generer_ebook.py déjà à jour')


# ═══════════════ 5. home_screen.dart ═══════════════
def _dart_page(nom, corps):
    return ("  Widget " + nom + "() {\n    return SingleChildScrollView(\n"
            "      padding: const EdgeInsets.fromLTRB(22, 20, 18, 18),\n"
            "      child: Column(crossAxisAlignment: CrossAxisAlignment.start, children: [\n        "
            + ",\n        ".join(corps) + ",\n      ]),\n    );\n  }\n\n")


def patcher_dart():
    if not os.path.isfile(DART): return
    s = open(DART, encoding='utf-8').read(); mod = False
    if "import 'dart:io';" not in s:
        s = s.replace('import ', "import 'dart:io';\nimport ", 1); mod = True
    if "import 'dart:convert';" not in s:
        s = s.replace("import 'dart:io';", "import 'dart:io';\nimport 'dart:convert';", 1); mod = True
    if 'package:path/path.dart' not in s:
        s = s.replace("import 'dart:io';", "import 'dart:io';\nimport 'package:path/path.dart' as path;", 1); mod = True
    if 'if (_progressTarget < 99) {' not in s:
        s = s.replace('if (_progressTarget < 95) {', 'if (_progressTarget < 99) {', 1)
        s = s.replace('(_progressTarget + 0.3).clamp(0, 95)', '(_progressTarget + 0.2).clamp(0, 99)', 1); mod = True
    # ── REGISTRE v4 : champs mutables + affectations lisant registre.json ──
    _cles = {'_statOuvrage': 'ouvrage', '_statFormat': 'format', '_statMots': 'mots',
             '_statPages': 'pages', '_statChapitres': 'chapitres', '_statIllus': 'illustrations'}
    if '_valeurRegistre' not in s:
        if 'Widget _pageRegistre() {' in s:
            s = s.replace('Widget _pageRegistre() {', DART_REGISTRE + '  Widget _pageRegistre() {', 1)
        else:
            s, _ = inserer_avant_lire_doc(s, DART_REGISTRE)
        mod = True
        print("   • Registre : lecteur registre.json ajouté (_valeurRegistre)")
    # 1) supprime les getters/méthodes _statX insérés par v3 (conflit setter)
    for _nom in _cles:
        if re.search(r'^[ \t]*String\s+(?:get\s+)?' + re.escape(_nom) + r'\s*(?:\([^)]*\))?\s*=>', s, re.M):
            s, _ = suppr_methode_dart(s, _nom)
            mod = True
    # 2) rétablit les champs mutables
    _champs = ''
    for _nom in _cles:
        if not re.search(r'^[ \t]*String\s+' + re.escape(_nom) + r'\s*=', s, re.M):
            _champs += "  String " + _nom + " = '—';\n"
    if _champs:
        if 'Widget _pageRegistre() {' in s:
            s = s.replace('Widget _pageRegistre() {', _champs + '  Widget _pageRegistre() {', 1)
        else:
            s += '\n' + _champs
        mod = True
    # 3) réécrit les affectations existantes pour lire registre.json
    for _nom, _cle in _cles.items():
        s, _n = re.subn(r'^([ \t]*)' + re.escape(_nom) + r'\s*=\s*[^;\n]+;',
                        (lambda m, _nm=_nom, _cl=_cle: m.group(1) + _nm + " = _valeurRegistre('" + _cl + "');"),
                        s, flags=re.M)
        if _n:
            mod = True
    if mod:
        print("   • Registre v4 : champs rétablis + valeurs lues depuis registre.json")
    pages = {
      '_pageReglages': ["_titrePage('RÉGLAGES')", "_actionPage('⚙', 'Paramètres', _ouvrirParametres)",
        "_actionPage('📤', 'Dossier export', () => _ouvrirDossier('export'))",
        "_actionPage('🖼', 'Dossier des images', () => _ouvrirDossier('Images'))",
        "_actionPage('📁', 'Dossier des chapitres', () => _ouvrirDossier('Chapitres'))",
        "_actionPage('🌐', 'Dossier traductions', () => _ouvrirDossier('Traductions'))",
        "_actionPage('📥', 'Journal des erreurs', _ouvrirJournal)"],
      '_pageCorrection': ["_titrePage('CORRECTION')", "_chapitrePicker()", "const SizedBox(height: 6)",
        "_actionPage('📝', 'Corriger le chapitre', _corriger)"],
      '_pageProduction': ["_titrePage('PRODUCTION')", "_modeSelector()", "const SizedBox(height: 10)",
        "_actionPage('▶', 'Générer le livre', _genererLivre)",
        "_actionPage('🖨', 'PDF KDP noir & blanc', _exportPdf)",
        "_actionPage('📱', 'Ebook KDP (EPUB)', _genererEbook)", "_actionPage('🤖', 'Résumés IA', _resumesIA)"],
      '_pageLecture': ["_titrePage('LECTURE')", "_actionPage('📖', 'Lire le Word', () => _lireDocument('word'))",
        "_actionPage('📄', 'Lire le PDF', () => _lireDocument('pdf'))",
        "_actionPage('📚', \"Lire l'EPUB\", () => _lireDocument('epub'))",
        "_actionPage('📝', 'Lire le résumé', _lireResume)"],
      '_pageRegistre': ["_titrePage('REGISTRE')", "_ligneRegistre('Ouvrage', _statOuvrage)",
        "_ligneRegistre('Format', _statFormat)", "_ligneRegistre('Mots', _statMots)",
        "_ligneRegistre('Pages', _statPages)", "_ligneRegistre('Chapitres', _statChapitres)",
        "_ligneRegistre('Illustr.', _statIllus)"],
    }
    fins = {'_pageReglages': '  // ══ 1. INFORMATIONS', '_pageCorrection': '  // ══ 4. PRODUCTION',
            '_pageProduction': '  // ══ 5. LECTURE', '_pageLecture': '  // ══ 7. CONTACT',
            '_pageRegistre': 'Widget _ligneRegistre'}
    for nom, corps in pages.items():
        a = 'Widget ' + nom + '() {'
        if a in s and 'return SingleChildScrollView' not in s[s.index(a):s.index(a) + 200]:
            s, r = entre(s, a, fins[nom], _dart_page(nom, corps)); mod = mod or r
    if "Lire le résumé" not in s:
        a_epub = "_actionPage('📚', \"Lire l'EPUB\", () => _lireDocument('epub')),"
        if a_epub in s:
            s = s.replace(a_epub, a_epub + "\n        _actionPage('📝', 'Lire le résumé', _lireResume),", 1); mod = True
    if '_trouverDossierBackend' in s or '_afficherErreur' in s:
        s, _ = suppr_fonction_dart(s, 'Future<void> _lireResume()')
        s, ok = inserer_avant_lire_doc(s, FN_RESUME); mod = mod or ok
        if 'void _afficherMsg(' not in s:
            s, ok2 = inserer_avant_lire_doc(s, FN_AFFICHER); mod = mod or ok2
        print('   🔧 _lireResume() réparé (version autonome)')
    elif 'Future<void> _lireResume()' not in s:
        s, ok = inserer_avant_lire_doc(s, FN_RESUME); mod = mod or ok
        if 'void _afficherMsg(' not in s:
            s, ok2 = inserer_avant_lire_doc(s, FN_AFFICHER); mod = mod or ok2
    if mod:
        bak(DART); open(DART, 'w', encoding='utf-8').write(s)
        print('   ✅ home_screen.dart patché')
    else:
        print('   ℹ️ home_screen.dart déjà à jour')


# ═══════════════ 6. EXÉCUTION + CONTRÔLES ═══════════════
def main():
    print('📐 appliquer_correctif.py – application des correctifs…')
    ecrire_partages()
    patcher_roman()
    patcher_pdf()
    patcher_ebook()
    patcher_dart()
    normaliser_registres()
    ok = True
    for fic in (PYR, ROMAN, PDF, EBOOK):
        if os.path.isfile(fic):
            try:
                ast.parse(open(fic, encoding='utf-8').read())
                print('   ✅ syntaxe OK : ' + os.path.basename(fic))
            except SyntaxError as e:
                ok = False
                print(f'   ⚠️  SYNTAXE {os.path.basename(fic)} ligne {e.lineno} : {e.msg}')
    try:
        import importlib, sys as _sys
        if BASE not in _sys.path: _sys.path.insert(0, BASE)
        if 'regles' in _sys.modules: del _sys.modules['regles']
        r = importlib.import_module('regles')
        for attr in ('FORMATS_LIVRE', 'dimensions_pour_format', 'BAREME_KDP', 'GARDE',
                     'MARGES_SYMETRIQUES', 'gouttiere_kdp_pour', 'marge_kdp_pour',
                     'CORRECTIONS_COMMUNES', 'CORR_PAR_PREFIXE', 'REGEX_PENSEES',
                     'corrections_pour', 'TYPO', 'MISE_EN_PAGE'):
            getattr(r, attr)
        print('   ✅ regles.py expose tous les attributs requis')
    except Exception as e:
        ok = False
        print('   ⚠️  regles.py : ' + str(e))
    print('🎯 Terminé.' if ok else '🎯 Terminé avec avertissements.')


if __name__ == '__main__':
    main()